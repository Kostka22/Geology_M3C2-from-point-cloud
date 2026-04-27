import os
import numpy as np

# ── OpenMP thread count — must be set BEFORE py4dgeo is imported ─────
# py4dgeo uses OpenMP internally in C++. Setting OMP_NUM_THREADS here
# controls the actual thread count used during M3C2 distance computation.
_N_THREADS = max(1, (os.cpu_count() or 4) - 1)
os.environ['OMP_NUM_THREADS'] = str(_N_THREADS)

import py4dgeo
import laspy
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from scipy.spatial import cKDTree
from scipy import stats
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

# ── Lock matplotlib to the non-interactive Agg backend BEFORE anything
#    else can import it.  If matplotlib is initialised with TkAgg first
#    (which happens when ProfilePicker imports FigureCanvasTkAgg), any
#    subsequent matplotlib.use("Agg") call in the worker thread silently
#    fails and all figure creation in that thread still uses TkAgg —
#    allocating tk.Variable / tk.PhotoImage objects in the wrong thread,
#    which causes the Tcl_AsyncDelete crash on exit.
#    Setting it here, at import time, before *any* matplotlib symbol is
#    touched, guarantees worker-thread figures always use Agg.
import matplotlib
matplotlib.use("Agg")          # must be before any other matplotlib import
import matplotlib.pyplot as _mpl_plt   # pre-initialise in main thread
import matplotlib.colors as _mpl_colors
import matplotlib.cm as _mpl_cm
del _mpl_plt, _mpl_colors, _mpl_cm    # don't pollute the namespace

# ── Thread-safe UI helpers ────────────────────────────────────────────
# Pipeline functions run in a worker thread; all tkinter calls must be
# marshalled back to the main thread via root.after().
_root_ref = []   # filled in after root = tk.Tk()

def _ui_error(title, msg):
    """Show an error dialog safely from any thread."""
    if _root_ref:
        _root_ref[0].after(0, lambda t=title, m=msg: messagebox.showerror(t, m))
    else:
        messagebox.showerror(title, msg)

def _ui_info(title, msg):
    """Show an info dialog safely from any thread."""
    if _root_ref:
        _root_ref[0].after(0, lambda t=title, m=msg: messagebox.showinfo(t, m))
    else:
        messagebox.showinfo(title, msg)

# -----------------------------
# Progress Bar Manager
# -----------------------------
class ProgressManager:
    """
    All public methods are safe to call from any thread.
    Widget updates are marshalled to the main thread via root.after().
    """
    def __init__(self, current_bar, current_label, overall_bar, overall_label):
        self.current_bar   = current_bar
        self.current_label = current_label
        self.overall_bar   = overall_bar
        self.overall_label = overall_label
        self.total_steps   = 9
        self.current_step  = 0

    def _schedule(self, fn):
        """Run fn() on the main thread. Safe to call from any thread."""
        if _root_ref:
            _root_ref[0].after(0, fn)
        else:
            fn()   # fallback: already on main thread (startup)

    def update_current(self, value, text):
        v, t = value, text
        def _do():
            self.current_bar['value'] = v
            self.current_label.config(text=t)
        self._schedule(_do)

    def update_overall(self, step_name):
        self.current_step += 1
        pct  = (self.current_step / self.total_steps) * 100
        step = self.current_step
        name = step_name
        total = self.total_steps
        def _do():
            self.overall_bar['value'] = pct
            self.overall_label.config(
                text=f"Overall Progress: {name} ({step}/{total})")
        self._schedule(_do)

    def reset(self):
        self.current_step = 0
        def _do():
            self.current_bar['value']  = 0
            self.overall_bar['value']  = 0
            self.current_label.config(text="Ready")
            self.overall_label.config(text="Overall Progress: Ready")
        self._schedule(_do)




# ═══════════════════════════════════════════════════════════════════════
# Interactive Profile Picker  —  hillshaded terrain raster + profile lines
# ═══════════════════════════════════════════════════════════════════════
class ProfilePicker:
    """
    Bins the reference cloud to a Z-grid, computes an analytical hillshade
    (light from NW, 45°), blends it with the terrain colormap, and displays
    the result as an imshow — identical to what QGIS does with a DEM.
    Edges and ridges are immediately visible.

    Slider controls grid resolution (coarser = faster, finer = more detail).
    Left-click: first point → second point → line saved.
    Done button always visible (packed before canvas).
    """

    def __init__(self, parent, las_file1, las_file2, existing_lines=None):
        self.parent     = parent
        self.las_file1  = las_file1
        self.lines      = list(existing_lines or [])
        self.pending_pt = None
        self.pts_xy     = None   # raw (N,2) XY — kept for resolution rebin
        self.pts_z      = None   # raw (N,)  Z
        self._pa        = None   # pending-point artist
        self._res_var   = tk.DoubleVar(value=1.0)   # grid resolution [m]
        self._img_cache = {}     # res → (rgba, extent)

        # ── Window ──────────────────────────────────────────────────
        self.win = tk.Toplevel(parent)
        self.win.title(
            "Profile Picker  —  hillshaded terrain  |  click two points per line")
        self.win.geometry("1200x880")
        self.win.minsize(800, 640)
        self.win.grab_set()
        self.win.protocol("WM_DELETE_WINDOW", self._cancel)

        # ── Status bar (top) ────────────────────────────────────────
        self.status_var = tk.StringVar(value="Loading reference cloud…")
        tk.Label(self.win, textvariable=self.status_var,
                 fg="#1565c0", font=("Arial", 9, "bold"), anchor="w",
                 bg="#e8f0fb", relief="flat", padx=8, pady=4
                 ).pack(side="top", fill="x")

        # ══ Bottom bar packed FIRST — guaranteed never obscured ══
        bot = tk.Frame(self.win, bg="#f0f0f0", bd=1, relief="raised", pady=5)
        bot.pack(side="bottom", fill="x")

        # Right: Done  Cancel
        self.btn_done = tk.Button(
            bot, text="✓   Done — use these lines",
            command=self._done, state="disabled",
            font=("Arial", 10, "bold"), relief="flat",
            bg="#2e7d32", fg="white", padx=18, pady=7,
            activebackground="#1b5e20", activeforeground="white",
            cursor="hand2")
        self.btn_done.pack(side="right", padx=12, pady=2)

        tk.Button(bot, text="✕  Cancel", command=self._cancel,
                  font=("Arial", 9), relief="flat",
                  bg="#e0e0e0", padx=10, pady=5
                  ).pack(side="right", padx=4, pady=2)

        # Left: count  undo  clear  |  resolution slider
        self.lbl_count = tk.Label(bot, text="Lines: 0",
                                   font=("Arial", 9, "bold"), fg="#333333",
                                   bg="#f0f0f0")
        self.lbl_count.pack(side="left", padx=(12, 8), pady=2)

        self.btn_undo = tk.Button(bot, text="↩ Undo",
                                   command=self._undo, state="disabled",
                                   font=("Arial", 9), relief="flat",
                                   bg="#e0e0e0", padx=7, pady=5)
        self.btn_undo.pack(side="left", padx=3, pady=2)

        self.btn_clear = tk.Button(bot, text="✕ Clear all",
                                    command=self._clear, state="disabled",
                                    font=("Arial", 9), relief="flat",
                                    bg="#e0e0e0", padx=7, pady=5)
        self.btn_clear.pack(side="left", padx=3, pady=2)

        tk.Frame(bot, width=2, bg="#cccccc").pack(
            side="left", fill="y", padx=10, pady=4)

        tk.Label(bot, text="Grid res [m]:", font=("Arial", 9),
                 bg="#f0f0f0", fg="#444444").pack(side="left", padx=(0, 4))
        self.lbl_res = tk.Label(bot, text="1.0 m",
                                 font=("Arial", 9, "bold"),
                                 bg="#f0f0f0", fg="#1565c0", width=6)
        self.lbl_res.pack(side="left", padx=(0, 4))
        tk.Scale(bot, from_=0.25, to=5.0, resolution=0.25,
                 orient="horizontal", length=180, showvalue=False,
                 variable=self._res_var, command=self._on_res_change,
                 bg="#f0f0f0", highlightthickness=0,
                 troughcolor="#cccccc", activebackground="#1565c0"
                 ).pack(side="left", pady=2)
        tk.Label(bot, text="← finer    coarser →",
                 font=("Arial", 7), fg="#888888", bg="#f0f0f0"
                 ).pack(side="left", padx=(4, 0))

        # ── Canvas + toolbar fill remaining space ────────────────────
        from matplotlib.backends.backend_tkagg import (
            FigureCanvasTkAgg, NavigationToolbar2Tk)
        import matplotlib.pyplot as plt
        self._plt = plt

        cf = tk.Frame(self.win)
        cf.pack(side="top", fill="both", expand=True, padx=4, pady=(2, 0))

        self.fig, self.ax = plt.subplots(figsize=(11, 7.5))
        self.fig.patch.set_facecolor('#1a1a1a')
        self.ax.set_facecolor('#111111')
        self.ax.set_aspect('equal')
        self._style_ax()

        self.canvas = FigureCanvasTkAgg(self.fig, master=cf)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)

        tbf = tk.Frame(cf, bg="#2a2a2a")
        tbf.pack(fill="x")
        self.toolbar = NavigationToolbar2Tk(self.canvas, tbf)
        self.toolbar.update()

        self._cid   = self.canvas.mpl_connect('button_press_event', self._on_click)
        self.result = None

        threading.Thread(target=self._load_cloud, daemon=True).start()

    # ── Core raster computation ─────────────────────────────────────

    @staticmethod
    def _hillshade_rgba(x, y, z, res, azimuth=315.0, altitude=45.0):
        """
        Bin XYZ to a grid at *res* metres, compute analytical hillshade,
        blend with terrain colormap, return (rgba uint8, extent).

        Algorithm (identical to QGIS / GDAL hillshade):
          1. Build mean-Z grid
          2. np.gradient(z_grid, res) → dz/dx, dz/dy
          3. Compute slope and aspect
          4. hs = sin(alt)*cos(slope) + cos(alt)*sin(slope)*cos(az - aspect)
          5. Multiply terrain-colourmap RGB by hillshade factor
        """
        import matplotlib.colors as mcolors
        import matplotlib.pyplot as plt

        x_min, x_max = x.min(), x.max()
        y_min, y_max = y.min(), y.max()

        nx = max(2, int(np.ceil((x_max - x_min) / res)) + 1)
        ny = max(2, int(np.ceil((y_max - y_min) / res)) + 1)

        xi = np.clip(((x - x_min) / res).astype(int), 0, nx - 1)
        yi = np.clip(((y - y_min) / res).astype(int), 0, ny - 1)

        z_sum   = np.zeros((ny, nx), dtype=np.float64)
        z_cnt   = np.zeros((ny, nx), dtype=np.int32)
        np.add.at(z_sum, (yi, xi), z)
        np.add.at(z_cnt, (yi, xi), 1)

        valid   = z_cnt > 0
        z_grid  = np.where(valid, z_sum / np.where(z_cnt > 0, z_cnt, 1),
                           np.nan).astype(np.float32)

        # Fill NaN holes with nearest valid neighbour (simple 2-pass)
        if np.isnan(z_grid).any():
            from scipy.ndimage import distance_transform_edt
            nan_mask = np.isnan(z_grid)
            idx = distance_transform_edt(nan_mask,
                                         return_distances=False,
                                         return_indices=True)
            z_grid = z_grid[idx[0], idx[1]]

        # ── Hillshade ───────────────────────────────────────────────
        dzdx, dzdy = np.gradient(z_grid, res, res)
        slope  = np.arctan(np.hypot(dzdx, dzdy))
        aspect = np.arctan2(-dzdy, dzdx)   # GDAL convention

        az_rad  = np.radians(360.0 - azimuth + 90.0)
        alt_rad = np.radians(altitude)

        hs = (np.sin(alt_rad) * np.cos(slope)
              + np.cos(alt_rad) * np.sin(slope) * np.cos(az_rad - aspect))
        hs = np.clip(hs, 0.0, 1.0)

        # ── Terrain colour × hillshade ───────────────────────────────
        z_lo = float(np.percentile(z_grid, 2))
        z_hi = float(np.percentile(z_grid, 98))
        if z_hi <= z_lo:
            z_hi = z_lo + 1.0

        norm     = mcolors.Normalize(vmin=z_lo, vmax=z_hi)
        cmap     = plt.cm.terrain
        rgb      = cmap(norm(z_grid))[:, :, :3].astype(np.float32)

        # Blend: darken terrain by hillshade (0=dark, 1=full colour)
        # blend = terrain * (0.4 + 0.6 * hs)  — keeps colour even in shadow
        blend_hs = (0.4 + 0.6 * hs)[:, :, np.newaxis]
        rgb_out  = np.clip(rgb * blend_hs, 0.0, 1.0)

        alpha    = valid.astype(np.float32)
        rgba     = (np.dstack([rgb_out, alpha[:, :, np.newaxis]])
                    * 255).astype(np.uint8)

        extent = (x_min - res * 0.5, x_max + res * 0.5,
                  y_min - res * 0.5, y_max + res * 0.5)
        return rgba, extent, z_lo, z_hi

    # ── Load / plot ─────────────────────────────────────────────────

    def _load_cloud(self):
        try:
            self.status_var.set("Loading reference cloud…")
            las   = laspy.read(self.las_file1)
            x_all = np.asarray(las.x, dtype=np.float32)
            y_all = np.asarray(las.y, dtype=np.float32)
            z_all = np.asarray(las.z, dtype=np.float32)
            n = len(x_all)
            self.status_var.set(
                f"Building hillshade raster  ({n:,} points)…")
            self.pts_xy = np.column_stack([x_all, y_all])
            self.pts_z  = z_all
            self.win.after(0, self._plot_cloud)
        except Exception as e:
            import traceback; traceback.print_exc()
            self.win.after(0, lambda: self.status_var.set(f"Error: {e}"))

    def _get_raster(self, res):
        """Return cached (rgba, extent, z_lo, z_hi) for this resolution."""
        if res not in self._img_cache:
            self.status_var.set(f"Rasterising at {res} m…")
            result = self._hillshade_rgba(
                self.pts_xy[:, 0], self.pts_xy[:, 1], self.pts_z, res)
            self._img_cache[res] = result
        return self._img_cache[res]

    def _plot_cloud(self, keep_view=False):
        """Full redraw. If keep_view=True, restores zoom level after draw."""
        if self.pts_xy is None:
            return

        try:
            xlim = self.ax.get_xlim() if keep_view else None
            ylim = self.ax.get_ylim() if keep_view else None
            had  = keep_view and self.ax.has_data()
        except Exception:
            had = False

        res = float(self._res_var.get())
        rgba, extent, z_lo, z_hi = self._get_raster(res)

        self.ax.cla()
        self._style_ax()
        self.ax.set_facecolor('#111111')

        # Main hillshaded image
        self.ax.imshow(rgba, extent=extent, origin='lower',
                       aspect='equal', interpolation='bilinear',
                       zorder=2)

        # Invisible scatter just for a mappable → colorbar
        import matplotlib.colors as mcolors
        import matplotlib.cm as mcm
        norm = mcolors.Normalize(vmin=z_lo, vmax=z_hi)
        sm   = mcm.ScalarMappable(cmap='terrain', norm=norm)
        sm.set_array([])
        for _a in self.fig.get_axes()[1:]:
            _a.remove()
        cb = self.fig.colorbar(sm, ax=self.ax,
                               fraction=0.025, pad=0.01, aspect=35)
        cb.set_label("Z  [m]", color='#aaaaaa', fontsize=8)
        cb.ax.tick_params(colors='#aaaaaa', labelsize=7)
        cb.outline.set_edgecolor('#555555')

        if had and xlim and xlim[0] != xlim[1]:
            self.ax.set_xlim(xlim)
            self.ax.set_ylim(ylim)

        for i, ln in enumerate(self.lines):
            self._draw_line_on_ax(ln, i + 1)

        self.canvas.draw()
        ny, nx = rgba.shape[:2]
        self.status_var.set(
            f"Hillshade  {nx}×{ny} cells @ {res} m  "
            "·  Left-click: pt 1 → pt 2 = line  "
            "·  Deactivate pan/zoom before clicking")
        self._refresh_buttons()

    def _on_res_change(self, _=None):
        res = float(self._res_var.get())
        self.lbl_res.config(text=f"{res:.2f} m")
        if self.pts_xy is not None:
            # Recompute in background, then redraw
            def _worker():
                self._get_raster(res)
                self.win.after(0, lambda: self._plot_cloud(keep_view=True))
            threading.Thread(target=_worker, daemon=True).start()

    # ── Axis style ──────────────────────────────────────────────────

    def _style_ax(self):
        self.ax.tick_params(colors='#aaaaaa', labelsize=8)
        self.ax.set_xlabel("X  [m]", color='#aaaaaa', fontsize=9)
        self.ax.set_ylabel("Y  [m]", color='#aaaaaa', fontsize=9)
        for sp in self.ax.spines.values():
            sp.set_edgecolor('#555555')

    # ── Line drawing ────────────────────────────────────────────────

    def _draw_line_on_ax(self, line, idx):
        x0, y0 = float(line[0, 0]), float(line[0, 1])
        x1, y1 = float(line[1, 0]), float(line[1, 1])
        length  = float(np.linalg.norm(line[1] - line[0]))
        self.ax.plot([x0, x1], [y0, y1], '-',
                     color='#FF4400', linewidth=2.4, zorder=9,
                     solid_capstyle='round')
        self.ax.plot([x0, x1], [y0, y1], 'o',
                     color='#FF4400', markersize=7, zorder=10,
                     markeredgecolor='white', markeredgewidth=0.8)
        mx, my = (x0 + x1) / 2, (y0 + y1) / 2
        self.ax.annotate(
            f' P{idx}  {length:.1f} m',
            xy=(mx, my), fontsize=9, color='white', fontweight='bold',
            zorder=11,
            bbox=dict(boxstyle='round,pad=0.35', facecolor='#CC3300',
                      alpha=0.88, edgecolor='none'))

    # ── Click handler ───────────────────────────────────────────────

    def _on_click(self, event):
        if self.toolbar and self.toolbar.mode:
            return
        if event.inaxes != self.ax:
            return
        if event.button != 1:
            return
        if event.xdata is None or event.ydata is None:
            return
        if self.pts_xy is None:
            return

        x, y = event.xdata, event.ydata

        if self.pending_pt is None:
            self.pending_pt = (x, y)
            self._pa = self.ax.plot(
                x, y, 'o', color='white', markersize=10, zorder=12,
                markeredgecolor='#FF4400', markeredgewidth=2.0)[0]
            self.canvas.draw_idle()
            self.status_var.set(
                f"Point 1: ({x:.2f}, {y:.2f})  —  now click second point")
        else:
            x0, y0 = self.pending_pt
            ln = np.array([[x0, y0], [x, y]], dtype=np.float64)
            self.lines.append(ln)
            if self._pa is not None:
                self._pa.remove()
                self._pa = None
            self.pending_pt = None
            self._draw_line_on_ax(ln, len(self.lines))
            self.canvas.draw_idle()
            length = float(np.linalg.norm(ln[1] - ln[0]))
            self.status_var.set(
                f"Line {len(self.lines)} added  ({length:.1f} m)  "
                "—  click to add another,  or click  ✓ Done")
            self._refresh_buttons()

    # ── Buttons ─────────────────────────────────────────────────────

    def _refresh_buttons(self):
        n  = len(self.lines)
        st = "normal" if n > 0 else "disabled"
        self.lbl_count.config(text=f"Lines: {n}")
        self.btn_undo.config(state=st)
        self.btn_clear.config(state=st)
        self.btn_done.config(state=st)

    def _undo(self):
        if not self.lines:
            return
        if self.pending_pt is not None:
            if self._pa:
                self._pa.remove()
                self._pa = None
            self.pending_pt = None
        self.lines.pop()
        self._plot_cloud(keep_view=True)

    def _clear(self):
        if not messagebox.askyesno("Clear all",
                "Remove all defined profile lines?", parent=self.win):
            return
        self.lines.clear()
        self.pending_pt = None
        self._pa = None
        self._plot_cloud(keep_view=True)

    def _done(self):
        if self.pending_pt is not None:
            if not messagebox.askyesno(
                    "Unfinished line",
                    "First point set but no second point yet.\n"
                    "Discard it and finish?", parent=self.win):
                return
        self.result = list(self.lines)
        self.canvas.mpl_disconnect(self._cid)
        self._plt.close(self.fig)
        self.win.grab_release()
        self.win.destroy()

    def _cancel(self):
        self.result = None
        self.canvas.mpl_disconnect(self._cid)
        self._plt.close(self.fig)
        self.win.grab_release()
        self.win.destroy()

    def wait(self):
        self.parent.wait_window(self.win)
        return self.result


# -----------------------------
# Subsampling
# -----------------------------
def subsample_by_distance(points, min_dist, progress_mgr=None):
    if progress_mgr:
        progress_mgr.update_current(0, "Subsampling: Building KD-Tree...")
    selected = []
    tree = cKDTree(points)
    mask = np.ones(len(points), dtype=bool)
    total_points = len(points)
    for i, p in enumerate(points):
        if mask[i]:
            selected.append(i)
            idx = tree.query_ball_point(p, r=min_dist)
            mask[idx] = False
        if progress_mgr and i % 1000 == 0:
            progress_mgr.update_current((i / total_points) * 100,
                                        f"Subsampling: {i}/{total_points} points processed")
    if progress_mgr:
        progress_mgr.update_current(100, f"Subsampling: Complete - {len(selected)} corepoints selected")
    return points[selected]


# -----------------------------
# Significance
# -----------------------------
def calculate_significance_improved(distances, uncertainties=None, confidence_level=0.95,
                                    sigma1_manual=None, sigma2_manual=None,
                                    sigma_mode='per_point',
                                    registration_error=0.0,
                                    progress_mgr=None):
    """
    Three mutually exclusive modes:

    A — per_point (default):
        LoD(i) = t × √(σ1²(i)/n1 + σ2²(i)/n2 + reg_error²)
        Per-point LoD from py4dgeo + user-supplied registration error.

    B — global_auto:
        sigma1 = mean(SPREAD1), sigma2 = mean(SPREAD2)  [from py4dgeo]
        LoD = t × √(σ1² + σ2²)  — one threshold, py4dgeo computes sigma.

    C — manual_sigma:
        User supplies σ1, σ2 (point cloud precision from calibration).
        LoD = t × √(σ1² + σ2²)  — one threshold.
    """
    if progress_mgr:
        progress_mgr.update_current(0, "Calculating significance...")

    if confidence_level == 0.95:   t_critical = 1.96
    elif confidence_level == 0.99: t_critical = 2.576
    elif confidence_level == 0.90: t_critical = 1.645
    else: t_critical = stats.t.ppf((1 + confidence_level) / 2, df=1000)

    if progress_mgr:
        progress_mgr.update_current(33, "Calculating detection limits...")

    # ── helper: safe array extraction ────────────────────────────────
    def _arr(key):
        try:
            return np.asarray(uncertainties[key], dtype=np.float64).ravel()
        except (TypeError, KeyError, ValueError):
            return None

    if sigma_mode == 'manual_sigma' and sigma1_manual is not None and sigma2_manual is not None:
        # ── Mode C: user-supplied σ1, σ2 ─────────────────────────────
        sigma1 = float(sigma1_manual)
        sigma2 = float(sigma2_manual)
        detection_limit = t_critical * np.sqrt(sigma1**2 + sigma2**2)
        significant = np.abs(distances) > detection_limit
        print(f"Manuálne sigma: σ1={sigma1:.4f} m  σ2={sigma2:.4f} m  "
              f"LoD={detection_limit:.4f} m  → {int(significant.sum()):,} significant")

    elif sigma_mode == 'global_auto':
        # ── Mode B: global LoD from py4dgeo mean spread ──────────────
        s1_arr = _arr('spread1')
        s2_arr = _arr('spread2')
        if s1_arr is not None and s2_arr is not None:
            sigma1 = float(np.nanmean(s1_arr))
            sigma2 = float(np.nanmean(s2_arr))
        else:
            sigma1 = sigma2 = 0.05
            print("  Fallback sigma=0.05 (spread nedostupný)")
        detection_limit = t_critical * np.sqrt(sigma1**2 + sigma2**2)
        significant = np.abs(distances) > detection_limit
        print(f"Globálny auto LoD: σ1={sigma1:.4f} m  σ2={sigma2:.4f} m  "
              f"LoD={detection_limit:.4f} m  → {int(significant.sum()):,} significant")

    else:
        # ── Mode A: per-point LoD from py4dgeo ───────────────────────
        lod_arr = _arr('lodetection')
        s1_arr  = _arr('spread1')
        s2_arr  = _arr('spread2')
        if lod_arr is not None and not np.all(np.isnan(lod_arr)):
            if registration_error > 0:
                lod_arr = np.sqrt(lod_arr**2 + registration_error**2)
            sigma1 = float(np.nanmean(s1_arr)) if s1_arr is not None else 0.0
            sigma2 = float(np.nanmean(s2_arr)) if s2_arr is not None else 0.0
            detection_limit = float(np.nanmean(lod_arr))
            significant = np.abs(distances) > lod_arr
            print(f"Per-point LoD (reg={registration_error:.4f} m): "
                  f"mean={detection_limit:.4f}  "
                  f"min={float(np.nanmin(lod_arr)):.4f}  "
                  f"max={float(np.nanmax(lod_arr)):.4f}  "
                  f"→ {int(significant.sum()):,} significant")
        else:
            # Fallback (PDAL — no per-point uncertainties)
            sigma1 = sigma2 = 0.05
            detection_limit = t_critical * np.sqrt(sigma1**2 + sigma2**2 + registration_error**2)
            significant = np.abs(distances) > detection_limit
            print(f"Fallback LoD={detection_limit:.4f} m (uncertainties nedostupné)")

    if progress_mgr:
        progress_mgr.update_current(100, "Significance calculation complete")

    return significant, detection_limit, sigma1, sigma2


# ═══════════════════════════════════════════════════════════════════════
# Quiver export  —  two modes
# ═══════════════════════════════════════════════════════════════════════
def export_quiver(corepoints, distances, normals, significant, output_file,
                  grid_size=1.0, export_png=True, export_html=True,
                  export_gpkg=True, progress_mgr=None,
                  mode='normal', dx=None, dy=None):
    if mode not in ('normal', 'motion'):
        raise ValueError(f"mode must be 'normal' or 'motion', got: {mode!r}")

    mode_label  = "Normal-direction quiver" if mode == 'normal' else "Motion vector (DX · DY)"
    mode_suffix = "_quiver_normal"           if mode == 'normal' else "_quiver_motion"

    if progress_mgr:
        progress_mgr.update_current(0, f"Quiver [{mode}]: binning onto grid...")

    norm_mag = np.linalg.norm(normals, axis=1, keepdims=True)
    norm_mag[norm_mag == 0] = 1.0
    n_norm = normals / norm_mag

    xs   = corepoints[:, 0]
    ys   = corepoints[:, 1]
    dist = np.asarray(distances, dtype=np.float64)

    valid = ~np.isnan(dist)
    xs, ys, dist = xs[valid], ys[valid], dist[valid]
    n_norm = n_norm[valid]

    if mode == 'motion':
        if dx is None or dy is None:
            _ui_error("Chyba", "Motion-vector quiver requires DX/DY arrays.")
            return []
        dx_v = np.asarray(dx, dtype=np.float64)[valid]
        dy_v = np.asarray(dy, dtype=np.float64)[valid]

    x_min, x_max = xs.min(), xs.max()
    y_min, y_max = ys.min(), ys.max()
    cols    = int(np.ceil((x_max - x_min) / grid_size)) + 1
    col_idx = np.floor((xs - x_min) / grid_size).astype(int)
    row_idx = np.floor((ys - y_min) / grid_size).astype(int)
    cell_key = row_idx * cols + col_idx

    sort_order  = np.argsort(cell_key, kind="stable")
    sorted_keys = cell_key[sort_order]
    sorted_dist = dist[sort_order]
    sorted_nx   = n_norm[sort_order, 0]
    sorted_ny   = n_norm[sort_order, 1]

    unique_keys, first_idx, n_points = np.unique(
        sorted_keys, return_index=True, return_counts=True)

    mean_d  = np.add.reduceat(sorted_dist, first_idx) / n_points
    mean_nx = np.add.reduceat(sorted_nx,   first_idx) / n_points
    mean_ny = np.add.reduceat(sorted_ny,   first_idx) / n_points

    if mode == 'motion':
        sorted_dx = dx_v[sort_order]
        sorted_dy = dy_v[sort_order]
        mean_dx   = np.add.reduceat(sorted_dx, first_idx) / n_points
        mean_dy   = np.add.reduceat(sorted_dy, first_idx) / n_points

    if progress_mgr:
        progress_mgr.update_current(20, f"Quiver [{mode}]: {len(unique_keys)} grid cells...")

    cell_rows = unique_keys // cols
    cell_cols = unique_keys  % cols
    gx = x_min + (cell_cols + 0.5) * grid_size
    gy = y_min + (cell_rows + 0.5) * grid_size

    if mode == 'normal':
        horiz_mag = np.sqrt(mean_nx**2 + mean_ny**2)
        safe = horiz_mag > 1e-9
        nx_n = np.where(safe, mean_nx / horiz_mag, 0.0)
        ny_n = np.where(safe, mean_ny / horiz_mag, 0.0)
        u = nx_n * np.abs(mean_d)
        v = ny_n * np.abs(mean_d)
    else:
        u = mean_dx
        v = mean_dy

    c  = mean_d
    gx = np.array(gx); gy = np.array(gy)
    u  = np.array(u);  v  = np.array(v); c = np.array(c)

    arrow_mag  = np.sqrt(u**2 + v**2)
    max_mag    = arrow_mag.max() if arrow_mag.max() > 0 else 1.0
    plot_scale = (0.9 * grid_size) / max_mag
    u_plot = u * plot_scale
    v_plot = v * plot_scale

    abs_max = float(np.nanpercentile(np.abs(c), 98))
    if abs_max == 0:
        abs_max = 1.0
    n_arrows = len(gx)
    saved_paths = []

    if export_png:
        try:
            import matplotlib.pyplot as plt
            import matplotlib.colors as mcolors
            from matplotlib.cm import ScalarMappable
            if progress_mgr:
                progress_mgr.update_current(40, f"Quiver [{mode}]: rendering PNG...")
            fig, ax = plt.subplots(figsize=(14, 11), dpi=150)
            ax.set_aspect("equal")
            ax.set_facecolor("#1a1a1a"); fig.patch.set_facecolor("#1a1a1a")
            norm_c = mcolors.TwoSlopeNorm(vmin=-abs_max, vcenter=0, vmax=abs_max)
            cmap   = plt.cm.RdBu_r
            ax.quiver(gx, gy, u_plot, v_plot, color=cmap(norm_c(c)),
                      angles="xy", scale_units="xy", scale=1,
                      width=0.0015, headwidth=5, headlength=6, alpha=0.92)
            sm = ScalarMappable(cmap=cmap, norm=norm_c); sm.set_array([])
            cbar = fig.colorbar(sm, ax=ax, fraction=0.03, pad=0.02)
            cbar.set_label("Mean M3C2 Distance [m]", color="white", fontsize=11)
            cbar.ax.yaxis.set_tick_params(color="white")
            plt.setp(cbar.ax.yaxis.get_ticklabels(), color="white")
            subtitle = ("Direction = normal XY projection  |  Length & colour = mean M3C2"
                        if mode == 'normal' else
                        "Direction & length = true horizontal displacement (DX, DY)  |  colour = mean M3C2")
            ax.set_title(f"M3C2 {mode_label}  (grid={grid_size} m, {n_arrows} cells)\n{subtitle}",
                         color="white", fontsize=10, pad=12)
            ax.set_xlabel("X [m]", color="white"); ax.set_ylabel("Y [m]", color="white")
            ax.tick_params(colors="white")
            for sp in ax.spines.values(): sp.set_edgecolor("#444444")
            png_path = (output_file.replace(".las", f"{mode_suffix}.png")
                                   .replace(".laz", f"{mode_suffix}.png"))
            fig.savefig(png_path, bbox_inches="tight", facecolor=fig.get_facecolor())
            plt.close(fig)
            print(f"Quiver PNG [{mode}]: {png_path}")
            saved_paths.append(png_path)
        except ImportError:
            _ui_error("Chyba", "Chýba matplotlib.\npip install matplotlib")

    if export_html:
        try:
            import plotly.graph_objects as go
            if progress_mgr:
                progress_mgr.update_current(65, f"Quiver [{mode}]: building HTML...")
            head_x = gx + u_plot; head_y = gy + v_plot
            seg_x = np.empty(n_arrows * 3); seg_y = np.empty(n_arrows * 3)
            seg_x[0::3]=gx;  seg_x[1::3]=head_x; seg_x[2::3]=np.nan
            seg_y[0::3]=gy;  seg_y[1::3]=head_y; seg_y[2::3]=np.nan
            fig_html = go.Figure()
            fig_html.add_trace(go.Scatter(
                x=seg_x, y=seg_y, mode="lines",
                line=dict(color="rgba(200,200,200,0.35)", width=1),
                hoverinfo="skip", showlegend=False, name="shafts"))
            hover_extra = ("Azimuth: %{customdata[3]:.1f}°<extra></extra>"
                           if mode == 'normal' else
                           "DX: %{customdata[4]:.4f} m<br>DY: %{customdata[5]:.4f} m<extra></extra>")
            cd_cols = [c, gx, gy, np.degrees(np.arctan2(v_plot, u_plot))]
            if mode == 'motion':
                cd_cols += [u, v]
            fig_html.add_trace(go.Scatter(
                x=head_x, y=head_y, mode="markers",
                marker=dict(symbol="arrow", size=9,
                            angle=np.degrees(np.arctan2(v_plot, u_plot)),
                            color=c, colorscale="RdBu_r", reversescale=False,
                            cmin=-abs_max, cmax=abs_max,
                            colorbar=dict(
                                title=dict(text="Mean M3C2<br>Distance [m]", font=dict(color="white")),
                                tickfont=dict(color="white"), thickness=16, len=0.75),
                            line=dict(width=0)),
                customdata=np.stack(cd_cols, axis=-1),
                hovertemplate=("Cell: (%{customdata[1]:.2f}, %{customdata[2]:.2f})<br>"
                               "Mean M3C2: %{customdata[0]:.4f} m<br>" + hover_extra),
                showlegend=False, name="heads"))
            subtitle_html = ("Direction = normal XY projection<br><sup>Length & colour = mean M3C2 distance</sup>"
                             if mode == 'normal' else
                             "Direction & length = true horizontal displacement (DX, DY)<br><sup>Colour = mean M3C2 distance</sup>")
            fig_html.update_layout(
                title=dict(text=f"M3C2 {mode_label} — grid={grid_size} m | {n_arrows} cells<br>{subtitle_html}",
                           font=dict(color="white", size=13)),
                xaxis=dict(title="X [m]", scaleanchor="y", scaleratio=1,
                           color="white", gridcolor="#2a2a2a", zerolinecolor="#444444"),
                yaxis=dict(title="Y [m]", color="white",
                           gridcolor="#2a2a2a", zerolinecolor="#444444"),
                paper_bgcolor="#1a1a1a", plot_bgcolor="#111111",
                font=dict(color="white"), width=1300, height=950,
                margin=dict(l=70, r=70, t=90, b=60))
            html_path = (output_file.replace(".las", f"{mode_suffix}.html")
                                    .replace(".laz", f"{mode_suffix}.html"))
            fig_html.write_html(html_path, include_plotlyjs="cdn")
            print(f"Quiver HTML [{mode}]: {html_path}")
            saved_paths.append(html_path)
        except ImportError:
            _ui_error("Chyba", "Chýba plotly.\npip install plotly")

    if export_gpkg:
        try:
            import geopandas as gpd
            from shapely.geometry import LineString
            if progress_mgr:
                progress_mgr.update_current(85, f"Quiver [{mode}]: writing GPKG...")
            geometries = [
                LineString([(float(gx[i]), float(gy[i])),
                            (float(gx[i]) + float(u_plot[i]),
                             float(gy[i]) + float(v_plot[i]))])
                for i in range(len(gx))
            ]
            records = []
            for i in range(len(gx)):
                r = {"mean_m3c2": float(c[i]), "abs_m3c2": float(abs(c[i])),
                     "cell_x": float(gx[i]), "cell_y": float(gy[i]),
                     "azimuth_deg": float(np.degrees(np.arctan2(float(v[i]), float(u[i])))),
                     "n_points": int(n_points[i]), "quiver_mode": mode}
                if mode == 'motion':
                    r["dx_mean"]   = float(u[i])
                    r["dy_mean"]   = float(v[i])
                    r["horiz_mag"] = float(arrow_mag[i])
                records.append(r)
            gdf = gpd.GeoDataFrame(records, geometry=geometries)
            gpkg_path = (output_file.replace(".las", f"{mode_suffix}.gpkg")
                                    .replace(".laz", f"{mode_suffix}.gpkg"))
            gdf.to_file(gpkg_path, driver="GPKG", layer=f"m3c2_{mode}")
            print(f"Quiver GPKG [{mode}]: {gpkg_path}  ({len(gdf)} arrows)")
            saved_paths.append(gpkg_path)
        except ImportError:
            _ui_error("Chyba", "Chýba geopandas/shapely.\npip install geopandas shapely")

    if progress_mgr:
        progress_mgr.update_current(100, f"Quiver [{mode}] export complete")
    return saved_paths


# -----------------------------
# Save LAS/LAZ + optional TXT
# -----------------------------
def save_results_simplified(corepoints, distances, nz, nxy, significant,
                            output_file, save_txt=False, progress_mgr=None,
                            uncertainties=None, dx=None, dy=None, dz=None,
                            normal_scale=None, las_fields=None):
    if las_fields is None:
        las_fields = {"Nz","Nxy","DX","DY","DZ","LOD","SPREAD1","SPREAD2","NORMAL_SCALE"}
    n = len(corepoints)
    if progress_mgr:
        progress_mgr.update_current(0, "Creating LAS file...")
    print("Ukladám výsledky do LAS/LAZ...")
    core_las = laspy.LasData(laspy.LasHeader(point_format=3, version="1.4"))
    core_las.x = corepoints[:, 0]; core_las.y = corepoints[:, 1]; core_las.z = corepoints[:, 2]
    if progress_mgr:
        progress_mgr.update_current(25, "Adding fields...")
    for name, arr, dtype in [("M3C2_DISTANCE", distances, np.float32),
                               ("SIGNIFICANT",   significant, np.int8)]:
        core_las.add_extra_dim(laspy.ExtraBytesParams(name=name, type=dtype))
        core_las[name] = arr.astype(dtype)
    txt_cols = [corepoints[:, 0], corepoints[:, 1], corepoints[:, 2],
                distances.astype(np.float32), significant.astype(np.float32)]
    txt_hdrs = ["X", "Y", "Z", "M3C2_DISTANCE", "SIGNIFICANT"]
    optional_fields = []
    if "Nz"  in las_fields: optional_fields.append(("Nz",  nz,  np.float32))
    if "Nxy" in las_fields: optional_fields.append(("Nxy", nxy, np.float32))
    if "DX"  in las_fields and dx is not None: optional_fields.append(("DX", dx, np.float32))
    if "DY"  in las_fields and dy is not None: optional_fields.append(("DY", dy, np.float32))
    if "DZ"  in las_fields and dz is not None: optional_fields.append(("DZ", dz, np.float32))
    if "NORMAL_SCALE" in las_fields and normal_scale is not None:
        optional_fields.append(("NORMAL_SCALE", normal_scale, np.float32))
    for name, arr, dtype in optional_fields:
        core_las.add_extra_dim(laspy.ExtraBytesParams(name=name, type=dtype))
        core_las[name] = arr.astype(dtype)
        txt_cols.append(arr.astype(dtype)); txt_hdrs.append(name)
    if progress_mgr:
        progress_mgr.update_current(50, "Adding LOD / SPREAD1 / SPREAD2...")
    for py4d_key, col_name in [("lodetection","LOD"),("spread1","SPREAD1"),("spread2","SPREAD2")]:
        if col_name not in las_fields:
            continue
        arr = None
        if uncertainties is not None:
            try:
                raw = uncertainties[py4d_key]
                arr = np.asarray(raw, dtype=np.float32).ravel()
                if len(arr) != n: arr = None
            except Exception: arr = None
        if arr is None:
            arr = np.full(n, np.nan, dtype=np.float32)
        core_las.add_extra_dim(laspy.ExtraBytesParams(name=col_name, type=np.float32))
        core_las[col_name] = arr
        txt_cols.append(arr); txt_hdrs.append(col_name)
    if progress_mgr:
        progress_mgr.update_current(70, "Writing LAS/LAZ file...")
    out_dir = os.path.dirname(output_file)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    try:
        core_las.write(output_file, do_compress=output_file.lower().endswith(".laz"))
        print(f"Uložený: {output_file}")
    except Exception as e:
        _ui_error("Chyba", f"Chyba pri ukladaní LAS: {e}")
        return False
    if save_txt:
        if progress_mgr:
            progress_mgr.update_current(85, "Exporting TXT...")
        txt_file = output_file.replace(".las","_results.txt").replace(".laz","_results.txt")
        try:
            np.savetxt(txt_file, np.column_stack(txt_cols), delimiter="\t",
                       header="\t".join(txt_hdrs), comments="", fmt="%.6f")
            print(f"TXT: {txt_file}")
        except Exception as e:
            print(f"  ! TXT failed: {e}")
    if progress_mgr:
        progress_mgr.update_current(100, "Export complete")
    return True


# -----------------------------
# Clip raster by vector mask
# -----------------------------
def clip_raster_by_mask(raster_path, mask_path, progress_mgr=None):
    """
    Clip a GeoTIFF in-place using a vector mask (SHP or GPKG).
    Uses rasterio.mask.  The first layer of the GPKG / SHP is used.
    Overwrites the input file with the clipped version.
    Returns True on success, False on error.
    """
    try:
        import rasterio
        from rasterio.mask import mask as rio_mask
        import geopandas as gpd
    except ImportError as e:
        _ui_error("Chyba",
                  f"Clip requires rasterio + geopandas:\npip install rasterio geopandas\n\n{e}")
        return False

    if progress_mgr:
        progress_mgr.update_current(0, f"Clipping: {os.path.basename(raster_path)}...")

    try:
        # Load vector mask
        gdf = gpd.read_file(mask_path)
        if gdf.empty:
            _ui_error("Chyba", f"Clip mask is empty: {mask_path}")
            return False

        with rasterio.open(raster_path) as src:
            # Reproject mask to raster CRS if needed
            if gdf.crs and src.crs and gdf.crs != src.crs:
                gdf = gdf.to_crs(src.crs)

            shapes = [geom.__geo_interface__ for geom in gdf.geometry if geom is not None]
            if not shapes:
                _ui_error("Chyba", "Clip mask has no valid geometries.")
                return False

            out_image, out_transform = rio_mask(src, shapes, crop=True,
                                                nodata=src.nodata, filled=True)
            out_meta = src.meta.copy()
            out_meta.update({
                'height':    out_image.shape[1],
                'width':     out_image.shape[2],
                'transform': out_transform,
            })

        # Write clipped raster back to same path
        tmp_path = raster_path + '.clip_tmp.tif'
        with rasterio.open(tmp_path, 'w', **out_meta) as dst:
            dst.write(out_image)

        os.replace(tmp_path, raster_path)
        print(f"  Clipped: {os.path.basename(raster_path)}")

    except Exception as e:
        _ui_error("Chyba", f"Clip error on {os.path.basename(raster_path)}:\n{e}")
        # Clean up temp file if it exists
        tmp = raster_path + '.clip_tmp.tif'
        if os.path.exists(tmp):
            try: os.remove(tmp)
            except Exception: pass
        return False

    if progress_mgr:
        progress_mgr.update_current(100, f"Clipped: {os.path.basename(raster_path)}")
    return True



def export_raster(corepoints, values, output_raster_path, resolution,
                  field_name="LOD", nodata=-9999.0, interp_method="IDW",
                  epsg=None, progress_mgr=None):
    try:
        import rasterio
        from rasterio.transform import from_origin
        from scipy.interpolate import griddata, NearestNDInterpolator
    except ImportError as e:
        _ui_error("Chyba", f"Chýba knižnica: {e}\npip install rasterio scipy")
        return False
    if progress_mgr:
        progress_mgr.update_current(0, f"Raster [{interp_method}]: {field_name}...")
    xs, ys = corepoints[:, 0], corepoints[:, 1]
    zv = np.asarray(values, dtype=np.float64)
    valid = ~np.isnan(zv)
    if valid.sum() < 4:
        _ui_error("Chyba", "Príliš málo platných bodov.")
        return False
    x_min, x_max = xs[valid].min(), xs[valid].max()
    y_min, y_max = ys[valid].min(), ys[valid].max()
    grid_x = np.arange(x_min, x_max + resolution, resolution)
    grid_y = np.arange(y_max, y_min - resolution, -resolution)
    gx, gy = np.meshgrid(grid_x, grid_y)
    if progress_mgr:
        progress_mgr.update_current(30, f"Raster: interpolácia ({interp_method})...")
    pts  = np.column_stack([xs[valid], ys[valid]])
    vals = zv[valid]
    m    = interp_method.upper()
    if m == "TIN":
        grid_z = griddata(pts, vals, (gx, gy), method="linear", fill_value=np.nan)
    elif m == "NEAREST":
        grid_z = NearestNDInterpolator(pts, vals)(gx, gy)
    elif m == "IDW":
        tree  = cKDTree(pts)
        k     = min(12, len(pts))
        dists, idx = tree.query(np.column_stack([gx.ravel(), gy.ravel()]), k=k, workers=_N_THREADS)
        dists = np.where(dists == 0, 1e-10, dists)
        w     = 1.0 / dists**2
        grid_z = (np.sum(w * vals[idx], axis=1) / np.sum(w, axis=1)).reshape(gx.shape)
    elif m == "KRIGING":
        try:
            from pykrige.ok import OrdinaryKriging
        except ImportError:
            _ui_error("Chyba", "pip install pykrige")
            return False
        ok = OrdinaryKriging(xs[valid], ys[valid], vals,
                             variogram_model="linear", verbose=False, enable_plotting=False)
        z_krig, _ = ok.execute("grid", grid_x, grid_y[::-1])
        grid_z = np.flipud(z_krig.data)
    else:
        _ui_error("Chyba", f"Neznáma metóda: {interp_method}")
        return False
    grid_z = np.where(np.isnan(grid_z), nodata, grid_z).astype(np.float32)
    out_dir = os.path.dirname(output_raster_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    try:
        transform = from_origin(x_min, y_max, resolution, resolution)
        crs = None
        if epsg:
            try:
                # Force PROJ to use pyproj's bundled proj.db, not any
                # system install (e.g. PostgreSQL/PostGIS) that may be stale.
                import pyproj
                import os as _os
                _proj_data = pyproj.datadir.get_data_dir()
                _old_proj  = _os.environ.get('PROJ_DATA')
                _os.environ['PROJ_DATA'] = _proj_data
                try:
                    crs = rasterio.crs.CRS.from_epsg(int(epsg))
                finally:
                    if _old_proj is None:
                        _os.environ.pop('PROJ_DATA', None)
                    else:
                        _os.environ['PROJ_DATA'] = _old_proj
            except Exception as e:
                print(f"  ! EPSG {epsg} not recognised: {e} — writing without CRS")
        with rasterio.open(output_raster_path, 'w', driver='GTiff',
                           height=grid_z.shape[0], width=grid_z.shape[1],
                           count=1, dtype='float32', nodata=nodata,
                           transform=transform,
                           crs=crs) as dst:
            dst.write(grid_z, 1)
        crs_str = f"  EPSG:{epsg}" if crs else "  (no CRS)"
        print(f"Raster: {output_raster_path}{crs_str}")
    except Exception as e:
        _ui_error("Chyba", f"Chyba rastra: {e}")
        return False
    if progress_mgr:
        progress_mgr.update_current(100, "Raster complete")
    return True


# ─────────────────────────────────────────────────────────────────────
# Batch CSV export
# ─────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────
# Excel auto-populate (reads batch_stats.csv → fills template → saves)
# ─────────────────────────────────────────────────────────────────────
EXCEL_TEMPLATE_NAME = 'm3c2_batch_analyza_template.xlsx'
EXCEL_OUTPUT_NAME   = 'batch_stats_analyza.xlsx'

# CSV key → Excel column index (1-based) for RAW_DATA sheet
_EXCEL_COL_MAP = {
    'run_name':     1,  'timestamp':    2,  'd':            3,
    'normal_radii': 4,  'core_spacing': 5,  'reg_error':    6,
    'sigma_mode':   7,  'sigma1':       8,  'sigma2':       9,
    'LoD95':       10,  'total_cp':    11,  'valid_cp':    12,
    'nan_cp':      13,  'sig_count':   14,  'sig_pct':     15,
    'mean_dist':   16,  'median_dist': 17,  'std_dist':    18,
    'min_dist':    19,  'max_dist':    20,  'rms_dist':    21,
    'nz_mean':     22,  'nz_min':      23,
}
_EXCEL_NUMERIC_COLS = {
    'd', 'core_spacing', 'reg_error', 'sigma1', 'sigma2', 'LoD95',
    'total_cp', 'valid_cp', 'nan_cp', 'sig_count', 'sig_pct',
    'mean_dist', 'median_dist', 'std_dist', 'min_dist', 'max_dist',
    'rms_dist', 'nz_mean', 'nz_min',
}

def _populate_excel_from_csv(csv_path):
    """
    Read batch_stats.csv and populate m3c2_batch_analyza_template.xlsx
    (located next to this .py file). Saves result as batch_stats_analyza.xlsx
    next to the CSV. Preserves template charts, formulas, and styling.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  ! openpyxl not installed — Excel auto-populate skipped")
        print("    Install with:  pip install openpyxl")
        return False

    import csv as _csv

    # Locate template next to the .py file
    script_dir    = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, EXCEL_TEMPLATE_NAME)

    if not os.path.exists(template_path):
        print(f"  ! Excel template not found: {template_path}")
        print(f"    → Place {EXCEL_TEMPLATE_NAME} next to m3c2_checkpoint14.py")
        print(f"    → CSV export still OK, only Excel skipped")
        return False

    if not os.path.exists(csv_path):
        return False

    # Read CSV
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            rows = list(_csv.DictReader(f))
    except Exception as e:
        print(f"  ! CSV read failed: {e}")
        return False

    if not rows:
        return False

    # Load template
    try:
        wb = load_workbook(template_path)
    except Exception as e:
        print(f"  ! Cannot load Excel template: {e}")
        return False

    if 'RAW_DATA' not in wb.sheetnames:
        print(f"  ! Template missing 'RAW_DATA' sheet — aborting Excel populate")
        return False
    ws = wb['RAW_DATA']

    # Clear any old data rows (4..103) to avoid stale values when CSV shrinks
    for rw in range(4, 104):
        for ci in range(1, 24):
            cell = ws.cell(row=rw, column=ci)
            if cell.value is not None:
                cell.value = None

    # Write rows starting at row 4
    for i, row in enumerate(rows):
        excel_row = 4 + i
        if excel_row > 103:
            print(f"  ! CSV has {len(rows)} rows — template supports max 100; truncated")
            break
        for key, ci in _EXCEL_COL_MAP.items():
            val = row.get(key, '')
            if key in _EXCEL_NUMERIC_COLS and val not in ('', None):
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=excel_row, column=ci, value=val)

    # Save output Excel next to CSV
    output_xlsx = os.path.join(os.path.dirname(csv_path), EXCEL_OUTPUT_NAME)
    try:
        wb.save(output_xlsx)
        print(f"  ✓ Excel auto-populated ({len(rows)} rows) → {output_xlsx}")
        return True
    except PermissionError:
        print(f"  ! {EXCEL_OUTPUT_NAME} is open in Excel/LibreOffice — close it and re-run")
        return False
    except Exception as e:
        print(f"  ! Excel save failed: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────
# Batch CSV export
# ─────────────────────────────────────────────────────────────────────
def _append_batch_csv_row(output_file, report_data, csv_dir_override=None):
    """
    Append one summary row to  <output_dir>/batch_stats.csv.
    Creates the file with a header row if it does not yet exist.
    Called after every run when 'Export batch stats to CSV' is ticked.

    csv_dir_override : if provided, batch_stats.csv is written there instead
                       of next to output_file. Used by batch worker to keep
                       all N runs aggregated in one CSV (parent folder).

    After CSV write, _populate_excel_from_csv() is called to regenerate
    batch_stats_analyza.xlsx from the template (if template is present).
    """
    import csv, datetime as _dt
    import numpy as np

    d          = report_data.get('distances')
    sig        = report_data.get('significant')
    vm         = report_data.get('valid_mask')
    nz_arr     = report_data.get('nz')
    lod        = report_data.get('detection_limit', float('nan'))
    sigma1     = report_data.get('sigma1', float('nan'))
    sigma2     = report_data.get('sigma2', float('nan'))
    cp_arr     = report_data.get('corepoints')

    # Derived stats
    total_cp = len(d) if d is not None else 0
    valid_mask_full = (vm & ~np.isnan(d)) if (vm is not None and d is not None) else np.zeros(total_cp, dtype=bool)
    vd       = d[valid_mask_full] if d is not None else np.array([])
    valid_cp = int(valid_mask_full.sum())
    nan_cp   = total_cp - valid_cp
    sig_valid = sig[valid_mask_full].astype(bool) if sig is not None else np.zeros(valid_cp, dtype=bool)
    sig_count = int(sig_valid.sum())
    sig_pct   = (sig_count / valid_cp * 100) if valid_cp > 0 else float('nan')

    def _s(a): return float(a) if len(a) > 0 else float('nan')
    mean_d   = float(np.mean(vd))   if len(vd) > 0 else float('nan')
    median_d = float(np.median(vd)) if len(vd) > 0 else float('nan')
    std_d    = float(np.std(vd))    if len(vd) > 0 else float('nan')
    min_d    = float(np.min(vd))  if len(vd) > 0 else float('nan')
    max_d    = float(np.max(vd))  if len(vd) > 0 else float('nan')
    rms_d    = float(np.sqrt(np.mean(vd**2))) if len(vd) > 0 else float('nan')
    nz_mean  = float(np.mean(nz_arr))  if nz_arr is not None and len(nz_arr) > 0 else float('nan')
    nz_min   = float(np.min(nz_arr))   if nz_arr is not None and len(nz_arr) > 0 else float('nan')

    nr = report_data.get('normal_radii', [])
    if hasattr(nr, 'tolist'):          # numpy array → list
        nr = nr.tolist()
    normal_radii_str = ";".join(str(r) for r in nr) if isinstance(nr, (list, tuple)) else str(nr)

    row = {
        'run_name':       os.path.splitext(os.path.basename(output_file))[0],
        'timestamp':      _dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'd':              report_data.get('cyl_radius', ''),
        'normal_radii':   normal_radii_str,
        'core_spacing':   report_data.get('core_spacing', ''),
        'reg_error':      report_data.get('registration_error', ''),
        'sigma_mode':     report_data.get('sigma_mode', ''),
        'sigma1':         round(float(sigma1), 6),
        'sigma2':         round(float(sigma2), 6),
        'LoD95':          round(float(lod), 6),
        'total_cp':       total_cp,
        'valid_cp':       valid_cp,
        'nan_cp':         nan_cp,
        'sig_count':      sig_count,
        'sig_pct':        round(sig_pct, 4),
        'mean_dist':      round(mean_d, 6),
        'median_dist':    round(median_d, 6),
        'std_dist':       round(std_d, 6),
        'min_dist':       round(min_d, 6),
        'max_dist':       round(max_d, 6),
        'rms_dist':       round(rms_d, 6),
        'nz_mean':        round(nz_mean, 6),
        'nz_min':         round(nz_min, 6),
    }

    csv_dir = csv_dir_override if csv_dir_override else os.path.dirname(output_file)
    csv_path = os.path.join(csv_dir, 'batch_stats.csv')
    write_header = not os.path.exists(csv_path)
    try:
        with open(csv_path, 'a', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=list(row.keys()))
            if write_header:
                writer.writeheader()
            writer.writerow(row)
        print(f"  ✓ Batch CSV row appended → {csv_path}")
    except Exception as e:
        print(f"  ✗ Batch CSV export failed: {e}")
        return

    # ── Excel auto-populate (regenerate from template + current CSV) ──
    _populate_excel_from_csv(csv_path)


# -----------------------------
# Statistics
# -----------------------------
def print_comprehensive_stats(distances, significant, valid_mask, detection_limit,
                               max_depth, nz, nxy):
    valid_distances = distances[valid_mask & ~np.isnan(distances)]
    if len(valid_distances) == 0:
        print("Žiadne platné vzdialenosti!")
        return
    print(f"\n=== ŠTATISTIKA VÝSLEDKOV ===")
    print(f"Celkový počet corepoints : {len(distances)}")
    print(f"Platných bodov           : {len(valid_distances)}")
    print(f"Významných zmien         : {np.sum(significant[valid_mask])}")
    print(f"Detekčný limit           : {float(detection_limit):.4f} m")
    print(f"\nVZDIALENOSTI:")
    print(f"  Min={np.min(valid_distances):.4f}  Max={np.max(valid_distances):.4f}  "
          f"Mean={np.mean(valid_distances):.4f}  Median={np.median(valid_distances):.4f}  "
          f"SD={np.std(valid_distances):.4f}")
    print(f"\nNORMÁLOVÉ ZLOŽKY:")
    print(f"  Nz  min={nz.min():.3f}  max={nz.max():.3f}  mean={nz.mean():.3f}")
    print(f"  Nxy min={nxy.min():.3f}  max={nxy.max():.3f}  mean={nxy.mean():.3f}")
    bins   = [0, 0.1, 0.5, 1.0, 2.0, 5.0, 10.0, np.inf]
    labels = ["<0.1m","0.1-0.5m","0.5-1.0m","1.0-2.0m","2.0-5.0m","5.0-10.0m",">10.0m"]
    hist, _ = np.histogram(np.abs(valid_distances), bins=bins)
    print(f"\nROZDELENIE ZMIEN:")
    for lbl, cnt in zip(labels, hist):
        if cnt > 0:
            print(f"  {lbl}: {cnt} ({cnt/len(valid_distances)*100:.1f}%)")
    if max_depth > 0:
        print(f"\nMax depth {max_depth} m → odfiltrovaných: {len(distances)-len(valid_distances)}")


# ─────────────────────────────────────────────────────────────────────
# Profile helpers
# ─────────────────────────────────────────────────────────────────────

def export_profile_lines(profile_lines, output_stem,
                         export_gpkg=True, export_txt=True):
    """
    Export interactively picked profile lines to:
      <stem>_profiles.gpkg  — GeoPackage LineString layer (loadable in QGIS)
      <stem>_profiles.txt   — plain Y X text, one vertex per line, lines
                               separated by a blank line  (loadable back into
                               the tool via the TXT import field)

    profile_lines : list of (label, np.ndarray shape (2,2)) tuples
    Returns list of saved paths.
    """
    if not profile_lines:
        return []

    saved = []

    # ── GPKG ────────────────────────────────────────────────────────
    if export_gpkg:
        try:
            import geopandas as gpd
            from shapely.geometry import LineString
            geoms   = []
            records = []
            for i, (lbl, pxy) in enumerate(profile_lines):
                x0, y0 = float(pxy[0, 0]), float(pxy[0, 1])
                x1, y1 = float(pxy[1, 0]), float(pxy[1, 1])
                length  = float(np.linalg.norm(pxy[1] - pxy[0]))
                geoms.append(LineString([(x0, y0), (x1, y1)]))
                records.append({
                    'line_id':  i + 1,
                    'label':    lbl,
                    'x_start':  x0, 'y_start': y0,
                    'x_end':    x1, 'y_end':   y1,
                    'length_m': round(length, 3),
                })
            gdf  = gpd.GeoDataFrame(records, geometry=geoms)
            path = f"{output_stem}_profiles.gpkg"
            gdf.to_file(path, driver='GPKG', layer='profile_lines')
            print(f"Profile lines GPKG: {path}")
            saved.append(path)
        except ImportError:
            _ui_error("Chyba", "GPKG export requires geopandas.\npip install geopandas")

    # ── TXT ─────────────────────────────────────────────────────────
    if export_txt:
        try:
            path = f"{output_stem}_profiles.txt"
            with open(path, 'w') as fh:
                fh.write("# M3C2 profile lines  —  format: Y (Northing)  X (Easting)\n")
                fh.write("# Each line pair separated by blank line.\n")
                fh.write("# Load via 'Profile TXT' field in the M3C2 Tool.\n\n")
                for i, (lbl, pxy) in enumerate(profile_lines):
                    fh.write(f"# Line {i+1}: {lbl}\n")
                    for pt in pxy:
                        # Y X order (Northing Easting) — matches load_profile_line
                        fh.write(f"{pt[1]:.6f}  {pt[0]:.6f}\n")
                    fh.write("\n")
            print(f"Profile lines TXT: {path}")
            saved.append(path)
        except Exception as e:
            _ui_error("Chyba", f"Profile TXT export error: {e}")

    return saved



def load_profile_line(txt_path):
    """
    Read one or more profile polylines from a TXT file.

    Supports two formats:
    1. Simple: one vertex per line, Y X (Northing Easting).
       Returns a single (2,2) array — backwards compatible.
    2. Multi-line (as exported by export_profile_lines):
       Groups separated by blank lines, each group = one line.
       Returns a list of (label, array) tuples if multiple groups found,
       OR a single array if only one group.

    Lines starting with '#' are treated as comments.
    The '#' comment immediately before a group is used as the label.
    """
    groups   = []   # list of (label, list_of_coords)
    current  = []
    last_lbl = None

    with open(txt_path, 'r', encoding='utf-8', errors='replace') as fh:
        for raw in fh:
            line = raw.strip()
            if not line:
                # blank line = group separator
                if current:
                    groups.append((last_lbl or f"Line {len(groups)+1}", current[:]))
                    current  = []
                    last_lbl = None
                continue
            if line.startswith('#'):
                # capture label comment e.g. "# Line 2: Line 2"
                last_lbl = line.lstrip('#').strip().split(':', 1)[-1].strip() or None
                continue
            parts = line.replace(',', ' ').replace(';', ' ').split()
            if len(parts) >= 2:
                try:
                    y_val, x_val = float(parts[0]), float(parts[1])
                    current.append([x_val, y_val])   # stored as (X, Y)
                except ValueError:
                    pass

    # flush last group
    if current:
        groups.append((last_lbl or f"Line {len(groups)+1}", current[:]))

    if not groups:
        raise ValueError("Profile TXT contains no valid coordinate pairs.")

    # Build result
    result = []
    for lbl, coords in groups:
        arr = np.array(coords, dtype=np.float64)
        if len(arr) < 2:
            continue
        result.append((lbl, arr))

    if not result:
        raise ValueError("Profile TXT: no group has ≥ 2 vertices.")

    # Single group → return bare array (legacy behaviour)
    if len(result) == 1:
        return result[0][1]

    # Multiple groups → return list of (label, array) tuples
    return result


def load_profile_lines_from_txt(txt_path):
    """
    Wrapper that always returns list of (label, array) regardless of count.
    Used by start_m3c2_threaded so multi-line TXT files work correctly.
    """
    raw = load_profile_line(txt_path)
    if isinstance(raw, np.ndarray):
        lbl = os.path.splitext(os.path.basename(txt_path))[0]
        return [(lbl, raw)]
    return raw   # already list of (label, array)





def _project_to_polyline(pts_xy, profile_xy):
    n          = len(pts_xy)
    best_along = np.zeros(n, dtype=np.float64)
    best_perp  = np.zeros(n, dtype=np.float64)
    best_abs   = np.full(n, np.inf, dtype=np.float64)
    cumlen     = 0.0

    for i in range(len(profile_xy) - 1):
        p0, p1  = profile_xy[i], profile_xy[i + 1]
        seg     = p1 - p0
        seg_len = float(np.linalg.norm(seg))
        if seg_len < 1e-10:
            continue
        tdir = seg / seg_len
        v    = pts_xy - p0
        t    = v @ tdir
        tc   = np.clip(t, 0.0, seg_len)
        foot = p0 + np.outer(tc, tdir)
        dp   = pts_xy - foot
        absd = np.linalg.norm(dp, axis=1)
        signed = tdir[0] * dp[:, 1] - tdir[1] * dp[:, 0]
        closer = absd < best_abs
        best_along[closer] = cumlen + tc[closer]
        best_perp[closer]  = signed[closer]
        best_abs[closer]   = absd[closer]
        cumlen += seg_len

    return best_along, best_perp, best_abs, cumlen


def export_profile_png(corepoints, distances, significant, profile_xy,
                       corridor_width, output_path,
                       detection_limit=None, profile_label="", progress_mgr=None,
                       ref_cloud=None, comp_cloud=None):
    """
    Two-panel profile PNG.

    Upper panel  — reference cloud (blue) and comparison cloud (red),
                   both projected into the corridor.  If the raw clouds are
                   not supplied the panel falls back to corepoints coloured
                   by M3C2 distance.

    Lower panel  — M3C2 distance along the profile with LoD threshold lines
                   (unchanged).

    Parameters
    ----------
    ref_cloud  : (N, 3) float array  — full reference point cloud XYZ
    comp_cloud : (M, 3) float array  — full comparison point cloud XYZ
    """
    try:
        import matplotlib.pyplot as plt
        import matplotlib.colors as mcolors
    except ImportError:
        _ui_error("Chyba", "Chýba matplotlib.\npip install matplotlib")
        return False

    if progress_mgr:
        progress_mgr.update_current(10, f"Profile '{profile_label}': projecting onto line...")

    # ── Project M3C2 corepoints (used for lower panel + title count) ──
    pts_xy = corepoints[:, :2].astype(np.float64)
    dist_along_cp, _perp_cp, abs_perp_cp, total_len = _project_to_polyline(
        pts_xy, profile_xy)

    half_w  = corridor_width / 2.0
    mask_cp = abs_perp_cp <= half_w
    n_in    = int(mask_cp.sum())

    if n_in == 0:
        _ui_error(
            "Chyba",
            f"No corepoints found within {corridor_width} m of profile '{profile_label}'.\n"
            "Check coordinates and corridor width.")
        return False

    d_al_cp = dist_along_cp[mask_cp]
    m3c2    = distances[mask_cp].astype(np.float64)
    sig     = (significant[mask_cp].astype(bool)
               if significant is not None else np.ones(n_in, dtype=bool))

    order   = np.argsort(d_al_cp)
    d_al_cp = d_al_cp[order]
    m3c2    = m3c2[order]
    sig     = sig[order]

    # ── Project raw clouds for the upper panel ────────────────────────
    MAX_RAW = 500_000   # cap per cloud to keep projection fast

    def _project_cloud(cloud):
        """Return (dist_along, z) arrays for points inside the corridor."""
        if cloud is None or len(cloud) == 0:
            return None, None
        xyz = np.asarray(cloud, dtype=np.float64)
        if len(xyz) > MAX_RAW:
            idx = np.random.choice(len(xyz), MAX_RAW, replace=False)
            xyz = xyz[idx]
        da, _perp, ab_perp, _ = _project_to_polyline(xyz[:, :2], profile_xy)
        m = ab_perp <= half_w
        if m.sum() == 0:
            return None, None
        return da[m], xyz[m, 2]

    if progress_mgr:
        progress_mgr.update_current(20, f"Profile '{profile_label}': projecting raw clouds...")

    use_raw = (ref_cloud is not None) or (comp_cloud is not None)
    if use_raw:
        da_ref,  z_ref  = _project_cloud(ref_cloud)
        da_comp, z_comp = _project_cloud(comp_cloud)
    else:
        # Fallback: corepoints coloured by M3C2 (legacy behaviour)
        da_ref  = d_al_cp
        z_ref   = corepoints[mask_cp, 2][order]
        da_comp = None
        z_comp  = None

    if progress_mgr:
        progress_mgr.update_current(35, f"Profile '{profile_label}': building figure...")

    # ── Colour scale (lower panel only) ──────────────────────────────
    valid_m = m3c2[~np.isnan(m3c2)]
    abs_max = float(np.nanpercentile(np.abs(valid_m), 98)) if len(valid_m) > 0 else 1.0
    if abs_max == 0:
        abs_max = 1.0
    norm_c = mcolors.TwoSlopeNorm(vmin=-abs_max, vcenter=0, vmax=abs_max)
    cmap   = plt.cm.RdBu_r

    # ── Figure ────────────────────────────────────────────────────────
    fig = plt.figure(figsize=(16, 9), dpi=150, facecolor='white')
    gs  = fig.add_gridspec(2, 1, height_ratios=[2, 1], hspace=0.42)
    ax1 = fig.add_subplot(gs[0])
    ax2 = fig.add_subplot(gs[1])

    for ax in (ax1, ax2):
        ax.set_facecolor('#f7f7f7')
        ax.grid(True, which='major', color='white',   linewidth=1.0, zorder=0)
        ax.grid(True, which='minor', color='#eeeeee', linewidth=0.5, zorder=0)
        ax.minorticks_on()
        for sp in ax.spines.values(): sp.set_color('#bbbbbb')
        ax.tick_params(axis='both', colors='#444444', labelsize=8, length=4)
        ax.set_xlabel("Distance along profile [m]", fontsize=9,
                      color='#333333', labelpad=6)

    # ── Upper panel — reference (blue) + comparison (red) ────────────
    if use_raw:
        # Plot comparison first (behind), then reference on top
        if da_comp is not None:
            ax1.scatter(da_comp, z_comp,
                        s=3, color='#e53935', alpha=0.55, linewidths=0,
                        rasterized=True, zorder=2, label=f'Comparison  ({len(da_comp):,} pts)')
        if da_ref is not None:
            ax1.scatter(da_ref, z_ref,
                        s=3, color='#1565c0', alpha=0.65, linewidths=0,
                        rasterized=True, zorder=3, label=f'Reference  ({len(da_ref):,} pts)')

        leg1 = ax1.legend(fontsize=8, framealpha=0.9, edgecolor='#cccccc',
                          facecolor='white', loc='upper right', markerscale=3)
        for txt in leg1.get_texts():
            txt.set_color('#333333')
    else:
        # Legacy: corepoints coloured by M3C2
        sc = ax1.scatter(da_ref, z_ref, c=m3c2, cmap=cmap, norm=norm_c,
                         s=7, linewidths=0, zorder=3, rasterized=True)
        cb = fig.colorbar(sc, ax=ax1, fraction=0.022, pad=0.02, aspect=32)
        cb.set_label("M3C2 Distance [m]", fontsize=8, color='#333333', labelpad=6)
        cb.ax.tick_params(labelsize=7, colors='#444444')
        cb.outline.set_edgecolor('#bbbbbb')

    ax1.set_ylabel("Elevation Z [m]", fontsize=9, color='#333333', labelpad=6)

    title_str = "M3C2 Profile Cut"
    if profile_label:
        title_str += f"  ·  {profile_label}"
    title_str += (f"  ·  corridor {corridor_width} m  ·  {n_in} corepoints"
                  f"  ·  length {total_len:.1f} m")
    ax1.set_title(title_str, fontsize=10, color='#111111', pad=10,
                  fontweight='bold', loc='left')

    xlim = (-total_len * 0.01, total_len * 1.01)
    ax1.set_xlim(xlim)

    # ── Lower panel — M3C2 distance (unchanged) ───────────────────────
    not_sig = ~sig & ~np.isnan(m3c2)
    is_sig  =  sig & ~np.isnan(m3c2)
    if not_sig.any():
        ax2.scatter(d_al_cp[not_sig], m3c2[not_sig], color='#aaaaaa',
                    s=5, linewidths=0, zorder=3, label='Not significant',
                    rasterized=True)
    if is_sig.any():
        ax2.scatter(d_al_cp[is_sig], m3c2[is_sig],
                    c=m3c2[is_sig], cmap=cmap, norm=norm_c,
                    s=6, linewidths=0, zorder=4, label='Significant',
                    rasterized=True)

    ax2.axhline(0, color='#555555', linewidth=0.9, zorder=5)
    if detection_limit is not None and detection_limit > 0:
        ax2.axhspan(-detection_limit, +detection_limit,
                    alpha=0.10, color='#888888', zorder=1,
                    label=f'Below LoD  (\xb1{detection_limit:.3f} m)')
        ax2.axhline(+detection_limit, color='#c62828', linewidth=1.1,
                    linestyle='--', zorder=5, label=f'+LoD  {detection_limit:.3f} m')
        ax2.axhline(-detection_limit, color='#1565c0', linewidth=1.1,
                    linestyle='--', zorder=5, label=f'\u2212LoD  {detection_limit:.3f} m')

    ax2.set_ylabel("M3C2 Distance [m]", fontsize=9, color='#333333', labelpad=6)
    ax2.set_xlim(xlim)
    leg2 = ax2.legend(fontsize=7, framealpha=0.95, edgecolor='#cccccc',
                      facecolor='white', loc='upper right', ncol=2)
    for txt in leg2.get_texts():
        txt.set_color('#333333')

    # ── Footer ────────────────────────────────────────────────────────
    n_ref_str  = f"{len(da_ref):,}"  if da_ref  is not None else "–"
    n_comp_str = f"{len(da_comp):,}" if da_comp is not None else "–"
    fig.text(0.01, 0.005,
             f"Profile length: {total_len:.2f} m  ·  corridor ±{half_w:.1f} m  "
             f"·  ref: {n_ref_str} pts  ·  comp: {n_comp_str} pts  "
             f"·  M3C2 Tool – CHECKPOINT13",
             fontsize=6.5, color='#888888', va='bottom')
    fig.text(0.99, 0.005,
             f"Start: ({profile_xy[0,0]:.1f}, {profile_xy[0,1]:.1f})  "
             f"End: ({profile_xy[-1,0]:.1f}, {profile_xy[-1,1]:.1f})",
             fontsize=6.5, color='#888888', va='bottom', ha='right')

    if progress_mgr:
        progress_mgr.update_current(85, f"Profile '{profile_label}': saving PNG...")

    try:
        fig.savefig(output_path, bbox_inches='tight', facecolor='white', dpi=150)
        plt.close(fig)
        print(f"Profile PNG: {output_path}")
    except Exception as e:
        _ui_error("Chyba", f"Chyba pri ukladaní profilu: {e}")
        plt.close(fig)
        return False

    if progress_mgr:
        progress_mgr.update_current(100, "Profile export complete")
    return True


# ─────────────────────────────────────────────────────────────────────
# Wind Rose Export
# ─────────────────────────────────────────────────────────────────────

def export_windrose(corepoints, dx, dy, dz, significant, output_file,
                   n_bins=16, sig_only=False, suffix='_windrose',
                   sign_filter=None,
                   progress_mgr=None):
    """
    Export a two-panel wind rose PNG.

    sig_only=False   → all valid points (grey = not sig, coloured = sig)
    sig_only=True    → only significant points (all coloured by magnitude)
    sign_filter=None → no sign filter
    sign_filter='pos'→ only significant positive (gain/deposition)
    sign_filter='neg'→ only significant negative (loss/erosion)
    suffix           → appended to stem before .png
    """
    try:
        import matplotlib.pyplot as plt
        import matplotlib.colors as mcolors
        import matplotlib.patches as mpatches
    except ImportError:
        _ui_error("Chyba", "Chýba matplotlib.\npip install matplotlib")
        return None

    if progress_mgr:
        progress_mgr.update_current(5, "Wind rose: preparing data...")

    # ── Filter: always use all geometrically valid points ─────────────
    dx_a  = np.asarray(dx,          dtype=np.float64)
    dy_a  = np.asarray(dy,          dtype=np.float64)
    dz_a  = np.asarray(dz,          dtype=np.float64)
    sig_a = np.asarray(significant, dtype=bool)
    dist_a = np.asarray(
        getattr(significant, 'distances', None) or
        np.zeros(len(dx_a)), dtype=np.float64)

    valid = ~(np.isnan(dx_a) | np.isnan(dy_a) | np.isnan(dz_a))
    if valid.sum() < 3:
        _ui_error("Chyba", "Wind rose: not enough valid points (need ≥ 3).")
        return None

    dx_v  = dx_a[valid]
    dy_v  = dy_a[valid]
    dz_v  = dz_a[valid]
    sig_v = sig_a[valid]

    # ── If sig_only: keep only significant points ─────────────────────
    if sig_only:
        if sig_v.sum() < 3:
            _ui_error("Chyba", "Wind rose (sig only): not enough significant points.")
            return None
        dx_v  = dx_v[sig_v]
        dy_v  = dy_v[sig_v]
        dz_v  = dz_v[sig_v]
        sig_v = np.ones(len(dx_v), dtype=bool)

    # ── sign_filter: pos = gain only, neg = loss only ─────────────────
    if sign_filter in ('pos', 'neg'):
        # DZ is the vertical component of displacement; use it for sign
        # positive DZ = surface rose = deposition/gain
        sign_mask = (dz_v > 0) if sign_filter == 'pos' else (dz_v < 0)
        # Intersect with significant
        sign_mask = sign_mask & sig_v
        if sign_mask.sum() < 3:
            label = 'positive' if sign_filter == 'pos' else 'negative'
            _ui_error("Chyba", f"Wind rose ({label} only): not enough points.")
            return None
        dx_v  = dx_v[sign_mask]
        dy_v  = dy_v[sign_mask]
        dz_v  = dz_v[sign_mask]
        sig_v = np.ones(len(dx_v), dtype=bool)

    # ── 3-D displacement magnitude (significant points only for bounds) ─
    mag3d     = np.sqrt(dx_v**2 + dy_v**2 + dz_v**2)
    mag3d_sig = mag3d[sig_v]

    if len(mag3d_sig) < 2:
        mag3d_sig = mag3d

    # ── Magnitude class boundaries from significant points ────────────
    p25, p50, p75 = np.percentile(mag3d_sig, [25, 50, 75])

    def _round2(v):
        if v == 0: return 0.0
        from math import log10, floor
        d = floor(log10(abs(v)))
        return round(v, -d + 1)

    # sig_only: scale starts at min significant magnitude, not 0
    # all-points: scale starts at 0 (grey class begins there)
    b_min = _round2(float(mag3d_sig.min())) if sig_only else 0.0

    bounds = sorted(set([b_min,
                         _round2(p25),
                         _round2(p50),
                         _round2(p75),
                         mag3d_sig.max() * 1.001]))
    # ensure at least 3 boundaries for 2 classes
    if len(bounds) < 3:
        bounds = [b_min, mag3d_sig.mean(), mag3d_sig.max() * 1.001]

    # ── Colour scheme ─────────────────────────────────────────────────
    sig_colors    = ['#4CAF50', '#CDDC39', '#FF9800', '#F44336', '#9C27B0']
    n_sig_classes = len(bounds) - 1

    if sig_only:
        # All bars coloured — no grey class
        n_classes = n_sig_classes
        colors    = sig_colors[:n_sig_classes]

        def _class_idx(mag, is_sig):
            for ci in range(n_sig_classes - 1, -1, -1):
                if mag >= bounds[ci]:
                    return ci
            return 0
    else:
        # Class 0 = NOT significant → grey; classes 1..n = significant coloured
        n_classes = 1 + n_sig_classes
        colors    = ['#AAAAAA'] + sig_colors[:n_sig_classes]

        def _class_idx(mag, is_sig):
            if not is_sig:
                return 0
            for ci in range(n_sig_classes - 1, -1, -1):
                if mag >= bounds[ci]:
                    return ci + 1
            return 1

    class_idx = np.array([_class_idx(m, s) for m, s in zip(mag3d, sig_v)])

    if progress_mgr:
        progress_mgr.update_current(20, "Wind rose: computing azimuths...")

    # ── Horizontal azimuth: geographic bearing 0°=N, clockwise ───────
    # arctan2(DY, DX) is math angle (0°=E, CCW); convert to bearing:
    # bearing = 90° − math_angle  (mod 360°)
    math_az = np.degrees(np.arctan2(dy_v, dx_v))
    bearing  = (90.0 - math_az) % 360.0   # 0=N, 90=E, 180=S, 270=W

    # ── Vertical angle: 0°=horizontal, 90°=straight down/up ──────────
    horiz_mag = np.sqrt(dx_v**2 + dy_v**2)
    vert_angle = np.degrees(np.arctan2(np.abs(dz_v), horiz_mag))  # 0=horiz,90=vert
    # Map to 0–180° sector (like the reference panel c)
    vert_sector = vert_angle  # keep 0–90 range; mirror for display: plot 0..90..0

    # ── Bin helper ────────────────────────────────────────────────────
    def _polar_bins(angles_deg, n_b):
        bin_edges = np.linspace(0, 360, n_b + 1)
        bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
        bin_w = np.radians(360.0 / n_b)
        # stacked counts per class
        stacks = np.zeros((n_b, n_classes), dtype=int)
        for i in range(n_b):
            lo, hi = bin_edges[i], bin_edges[i+1]
            mask = (angles_deg >= lo) & (angles_deg < hi)
            for ci in range(n_classes):
                stacks[i, ci] = int(np.sum(mask & (class_idx == ci)))
        total = stacks.sum()
        freqs = stacks / max(total, 1)   # fraction per bin per class
        thetas = np.radians(bin_centers)
        return thetas, freqs, bin_w

    # Vertical bins use 0–90 range (two quadrants shown as 0–180 sector)
    def _sector_bins(angles_deg, n_b):
        """0–90° sector, mirrored to produce 0–180° display."""
        n_half = n_b // 2
        bin_edges = np.linspace(0, 90, n_half + 1)
        bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
        stacks = np.zeros((n_half, n_classes), dtype=int)
        for i in range(n_half):
            lo, hi = bin_edges[i], bin_edges[i+1]
            mask = (angles_deg >= lo) & (angles_deg < hi)
            for ci in range(n_classes):
                stacks[i, ci] = int(np.sum(mask & (class_idx == ci)))
        total = stacks.sum()
        freqs = stacks / max(total, 1)
        # Mirror: produce angles from 0° to 180°
        # left side (0→90°) + right side (90°→0°) displayed as a semicircle
        return bin_centers, freqs, bin_edges

    if progress_mgr:
        progress_mgr.update_current(40, "Wind rose: rendering figure...")

    thetas_h, freqs_h, bin_w_h = _polar_bins(bearing, n_bins)

    # ── Figure ────────────────────────────────────────────────────────
    fig = plt.figure(figsize=(14, 7), dpi=150, facecolor='white')
    fig.patch.set_facecolor('white')

    ax_h = fig.add_subplot(121, projection='polar')
    ax_v = fig.add_subplot(122, projection='polar')

    # ── Styling helper ───────────────────────────────────────────────
    def _style_polar(ax, title):
        ax.set_facecolor('#f8f8f8')
        ax.spines['polar'].set_color('#cccccc')
        ax.grid(color='#cccccc', linewidth=0.5, linestyle='-')
        ax.set_title(title, fontsize=11, fontweight='bold', color='#222222', pad=18)
        ax.tick_params(axis='both', labelsize=7, colors='#555555')

    # ══ Panel a: Horizontal wind rose ════════════════════════════════
    _style_polar(ax_h, 'a)  Horizontal movement direction')

    # North = top: rotate so 0°(N) is at top and goes clockwise
    ax_h.set_theta_zero_location('N')
    ax_h.set_theta_direction(-1)   # clockwise

    max_freq_h = freqs_h.sum(axis=1).max()
    r_ticks_h = np.linspace(0, max_freq_h, 5)[1:]
    ax_h.set_yticks(r_ticks_h)
    ax_h.set_yticklabels([f'{v*100:.0f}%' for v in r_ticks_h], fontsize=7)
    ax_h.set_xticks(np.radians([0, 45, 90, 135, 180, 225, 270, 315]))
    ax_h.set_xticklabels(['N','NE','E','SE','S','SW','W','NW'], fontsize=8)

    bottoms_h = np.zeros(n_bins)
    for ci in range(n_classes):
        heights = freqs_h[:, ci]
        bars = ax_h.bar(thetas_h, heights, width=bin_w_h * 0.92,
                        bottom=bottoms_h, color=colors[ci],
                        edgecolor='white', linewidth=0.4,
                        alpha=0.92, zorder=3)
        bottoms_h += heights

    # ══ Panel b: Vertical wind rose (semicircle) ═════════════════════
    _style_polar(ax_v, 'b)  Vertical movement angle')

    # 0° = horizontal (left), 90° = straight down (bottom), shown as semicircle
    ax_v.set_theta_zero_location('W')   # 0° on left
    ax_v.set_theta_direction(-1)
    ax_v.set_thetamin(0)
    ax_v.set_thetamax(180)

    n_vbins = n_bins // 2
    bin_edges_v = np.linspace(0, 90, n_vbins + 1)
    bin_centers_v = (bin_edges_v[:-1] + bin_edges_v[1:]) / 2
    bin_w_v = np.radians(90.0 / n_vbins)

    stacks_v = np.zeros((n_vbins, n_classes), dtype=int)
    for i in range(n_vbins):
        lo, hi = bin_edges_v[i], bin_edges_v[i+1]
        mask = (vert_angle >= lo) & (vert_angle < hi)
        for ci in range(n_classes):
            stacks_v[i, ci] = int(np.sum(mask & (class_idx == ci)))
    total_v = stacks_v.sum()
    freqs_v = stacks_v / max(total_v, 1)

    # Mirror: plot same bars on both halves (left 0–90, right 90–0)
    # Angles for left side (0→90°) and right side (90→180°)
    thetas_vL = np.radians(bin_centers_v)           # 0° → 90°
    thetas_vR = np.radians(180 - bin_centers_v)     # 90° → 0° mirrored

    max_freq_v = freqs_v.sum(axis=1).max()
    r_ticks_v = np.linspace(0, max_freq_v, 5)[1:]
    ax_v.set_yticks(r_ticks_v)
    ax_v.set_yticklabels([f'{v*100:.0f}%' for v in r_ticks_v], fontsize=7)
    ax_v.set_xticks(np.radians([0, 45, 90, 135, 180]))
    ax_v.set_xticklabels(['0°\n(horiz.)', '45°', '90°\n(vert.)', '45°', '0°\n(horiz.)'],
                          fontsize=7)

    bottoms_vL = np.zeros(n_vbins)
    bottoms_vR = np.zeros(n_vbins)
    for ci in range(n_classes):
        heights = freqs_v[:, ci]
        ax_v.bar(thetas_vL, heights, width=bin_w_v * 0.88,
                 bottom=bottoms_vL, color=colors[ci],
                 edgecolor='white', linewidth=0.4, alpha=0.92, zorder=3)
        ax_v.bar(thetas_vR, heights, width=bin_w_v * 0.88,
                 bottom=bottoms_vR, color=colors[ci],
                 edgecolor='white', linewidth=0.4, alpha=0.92, zorder=3)
        bottoms_vL += heights
        bottoms_vR += heights

    # ══ Shared legend ════════════════════════════════════════════════
    patches = []
    if sig_only:
        # All bars coloured — no grey entry
        for ci in range(n_classes):
            lo_lbl = f'{bounds[ci]:.2f}'
            hi_lbl = (f'{bounds[ci+1]:.2f}' if ci < n_sig_classes - 1
                      else f'> {bounds[-2]:.2f}')
            lbl = (f'{lo_lbl} \u2013 {hi_lbl} m' if ci < n_sig_classes - 1
                   else f'> {bounds[-2]:.2f} m')
            n_cls = int((class_idx == ci).sum())
            patches.append(mpatches.Patch(color=colors[ci],
                                           label=f'{lbl}  ({n_cls:,})'))
    else:
        # Grey = not significant
        n_not_sig = int((class_idx == 0).sum())
        patches.append(mpatches.Patch(color='#AAAAAA',
                                       label=f'Not significant  ({n_not_sig:,})'))
        for ci in range(1, n_classes):
            lo_lbl = f'{bounds[ci-1]:.2f}'
            hi_lbl = (f'{bounds[ci]:.2f}' if ci < n_sig_classes
                      else f'> {bounds[-2]:.2f}')
            lbl = (f'{lo_lbl} \u2013 {hi_lbl} m' if ci < n_sig_classes
                   else f'> {bounds[-2]:.2f} m')
            n_cls = int((class_idx == ci).sum())
            patches.append(mpatches.Patch(color=colors[ci],
                                           label=f'{lbl}  ({n_cls:,})'))

    fig.legend(handles=patches, title='3-D displacement [m]',
               title_fontsize=8, fontsize=8,
               loc='lower center', ncol=min(n_classes, 4),
               framealpha=0.95, edgecolor='#cccccc',
               bbox_to_anchor=(0.5, -0.04))

    n_total  = int(valid.sum())
    n_sig_ct = int(sig_a[valid].sum())
    n_shown  = int(len(dx_v))

    if sign_filter == 'pos':
        mode_str = f'positive changes only (deposition / gain)  ·  {n_shown:,} points'
    elif sign_filter == 'neg':
        mode_str = f'negative changes only (erosion / loss)  ·  {n_shown:,} points'
    elif sig_only:
        mode_str = f'significant only  ·  {n_shown:,} points'
    else:
        mode_str = (f'{n_sig_ct:,} significant, '
                    f'{n_total-n_sig_ct:,} not significant  ·  {n_total:,} points')

    fig.text(0.5, 1.01,
             f"M3C2 Wind Rose  \u00b7  {mode_str}  \u00b7  {n_bins} azimuth bins",
             ha='center', fontsize=10, fontweight='bold', color='#222222')

    fig.text(0.01, -0.02,
             f"M3C2 Tool \u2013 CHECKPOINT13  \u00b7  Samuel Leb\xf3",
             fontsize=7, color='#aaaaaa', va='bottom')

    plt.tight_layout(rect=[0, 0.06, 1, 1])

    stem     = os.path.splitext(output_file)[0]
    out_path = f"{stem}{suffix}.png"

    if progress_mgr:
        progress_mgr.update_current(90, "Wind rose: saving PNG...")

    try:
        fig.savefig(out_path, bbox_inches='tight', facecolor='white', dpi=150)
        plt.close(fig)
        print(f"Wind rose PNG: {out_path}")
    except Exception as e:
        _ui_error("Chyba", f"Wind rose save error: {e}")
        plt.close(fig)
        return None

    if progress_mgr:
        progress_mgr.update_current(100, "Wind rose complete")
    return out_path


# ═══════════════════════════════════════════════════════════════════════
# Volume Change Estimation
# ═══════════════════════════════════════════════════════════════════════

def compute_volume_change(corepoints, distances, significant, nz,
                          cell_size, las_file1=None, las_file2=None,
                          dem_resolution=0.5, dem_interp='IDW',
                          detection_limit=0.0,
                          progress_mgr=None):
    """
    Compute volume change by two DoD methods:

    Method A — DoD all cells
        Both clouds rasterised → dZ per cell, sum all non-NaN cells.

    Method B — DoD significant cells only
        Same raster, only cells where |dZ| > LoD contribute.
    """
    valid_mask = ~np.isnan(distances)
    sig_mask   = significant.astype(bool) & valid_mask

    result = {
        'dod_all_gain':   None, 'dod_all_loss': None, 'dod_all_net': None,
        'dod_sig_gain':   None, 'dod_sig_loss': None, 'dod_sig_net': None,
        'dod_res':        dem_resolution,
        'dod_lod':        detection_limit,
        'n_all_cells':    None,
        'n_sig_cells':    None,
        'dod_ok':         False,
    }

    if las_file1 is None or las_file2 is None:
        return result

    try:
        from scipy.spatial import cKDTree as _KD
        from scipy.interpolate import griddata as _gd

        if progress_mgr:
            progress_mgr.update_current(10, "Volume: reading point clouds...")

        las1 = laspy.read(las_file1)
        las2 = laspy.read(las_file2)
        x1 = np.asarray(las1.x, np.float64)
        y1 = np.asarray(las1.y, np.float64)
        z1 = np.asarray(las1.z, np.float64)
        x2 = np.asarray(las2.x, np.float64)
        y2 = np.asarray(las2.y, np.float64)
        z2 = np.asarray(las2.z, np.float64)

        # Overlapping extent
        x_min = max(x1.min(), x2.min())
        x_max = min(x1.max(), x2.max())
        y_min = max(y1.min(), y2.min())
        y_max = min(y1.max(), y2.max())

        if x_min >= x_max or y_min >= y_max:
            print("  Volume DoD: cloud extents do not overlap.")
            return result

        # ── Thin raw clouds to DoD grid resolution ────────────────────
        # Running IDW on 2M+ raw points for every grid cell is the main
        # bottleneck. One point per dem_resolution cell is more than
        # sufficient for volume estimation — thin both clouds first.
        if progress_mgr:
            progress_mgr.update_current(18, "Volume: thinning clouds to grid resolution...")

        def _thin(x, y, z, res):
            """Keep one point per res×res cell (highest Z — avoids voids)."""
            xi = np.floor((x - x_min) / res).astype(np.int32)
            yi = np.floor((y - y_min) / res).astype(np.int32)
            key = xi.astype(np.int64) * 1_000_000 + yi.astype(np.int64)
            order = np.argsort(key)
            key_s, z_s = key[order], z[order]
            x_s, y_s   = x[order], y[order]
            # For each unique cell keep the index with max Z
            _, first = np.unique(key_s, return_index=True)
            last      = np.append(first[1:], len(key_s))
            best = np.array([
                first[i] + np.argmax(z_s[first[i]:last[i]])
                for i in range(len(first))
            ])
            return x_s[best], y_s[best], z_s[best]

        x1, y1, z1 = _thin(x1, y1, z1, dem_resolution)
        x2, y2, z2 = _thin(x2, y2, z2, dem_resolution)
        print(f"  Volume DoD: thinned cloud1={len(x1):,}  cloud2={len(x2):,} pts "
              f"(res={dem_resolution} m)")

        gx_v = np.arange(x_min, x_max, dem_resolution)
        gy_v = np.arange(y_min, y_max, dem_resolution)
        gx, gy = np.meshgrid(gx_v, gy_v)
        pts_grid = np.column_stack([gx.ravel(), gy.ravel()])

        def _interp_z(x, y, z, pts, method):
            m = method.upper()
            if m == 'IDW':
                tree = _KD(np.column_stack([x, y]))
                k    = min(12, len(x))
                d, idx = tree.query(pts, k=k, workers=_N_THREADS)
                d = np.where(d == 0, 1e-10, d)
                w = 1.0 / d ** 2
                return (np.sum(w * z[idx], axis=1) /
                        np.sum(w, axis=1)).reshape(gx.shape)
            else:
                return _gd(np.column_stack([x, y]), z, (gx, gy),
                           method='linear', fill_value=np.nan)

        # Interpolate both epochs simultaneously in separate threads
        if progress_mgr:
            progress_mgr.update_current(30, "Volume: interpolating both epochs in parallel...")
        with ThreadPoolExecutor(max_workers=2) as ex:
            fut1 = ex.submit(_interp_z, x1, y1, z1, pts_grid, dem_interp)
            fut2 = ex.submit(_interp_z, x2, y2, z2, pts_grid, dem_interp)
            z1_grid = fut1.result()
            z2_grid = fut2.result()

        if progress_mgr:
            progress_mgr.update_current(85, "Volume: summing DoD cells...")

        dz     = z2_grid - z1_grid          # positive = surface rose
        cell_a = dem_resolution ** 2
        valid  = ~np.isnan(dz)

        # ── Method A: all valid cells ─────────────────────────────────
        all_pos = valid & (dz > 0)
        all_neg = valid & (dz < 0)
        dod_all_gain = float(np.nansum(dz[all_pos]) * cell_a)
        dod_all_loss = float(np.nansum(dz[all_neg]) * cell_a)
        dod_all_net  = dod_all_gain + dod_all_loss

        # ── Method B: significant cells only (|dZ| > LoD) ────────────
        lod = float(detection_limit) if detection_limit and detection_limit > 0 else 0.0
        sig_cells = valid & (np.abs(dz) > lod)
        sig_pos   = sig_cells & (dz > 0)
        sig_neg   = sig_cells & (dz < 0)
        dod_sig_gain = float(np.nansum(dz[sig_pos]) * cell_a)
        dod_sig_loss = float(np.nansum(dz[sig_neg]) * cell_a)
        dod_sig_net  = dod_sig_gain + dod_sig_loss

        # ── Percent difference ────────────────────────────────────────
        denom = (abs(dod_all_net) + abs(dod_sig_net)) / 2
        pct   = abs(dod_all_net - dod_sig_net) / denom * 100 if denom > 0 else None

        # ── Volume uncertainty via error propagation (Bernard et al. 2021) ──
        # σ_V = √[Σ (LoD × A_cell)²] = LoD × A_cell × √N_cells
        # Assumes per-cell errors are independent — propagate in quadrature.
        if lod > 0:
            n_all_int = int(valid.sum())
            n_sig_int = int(sig_cells.sum())
            cell_err    = lod * cell_a                                # per-cell uncertainty [m³]
            sigma_V_all = float(np.sqrt(n_all_int) * cell_err)
            sigma_V_sig = float(np.sqrt(n_sig_int) * cell_err)
            rel_err_all = abs(sigma_V_all / dod_all_net) * 100 if abs(dod_all_net) > 0 else None
            rel_err_sig = abs(sigma_V_sig / dod_sig_net) * 100 if abs(dod_sig_net) > 0 else None
        else:
            sigma_V_all = sigma_V_sig = rel_err_all = rel_err_sig = None

        result.update({
            'dod_all_gain': dod_all_gain,
            'dod_all_loss': dod_all_loss,
            'dod_all_net':  dod_all_net,
            'dod_sig_gain': dod_sig_gain,
            'dod_sig_loss': dod_sig_loss,
            'dod_sig_net':  dod_sig_net,
            'sigma_V_all':  sigma_V_all,
            'sigma_V_sig':  sigma_V_sig,
            'rel_err_all':  rel_err_all,
            'rel_err_sig':  rel_err_sig,
            'n_all_cells':  int(valid.sum()),
            'n_sig_cells':  int(sig_cells.sum()),
            'pct_diff':     pct,
            'dod_ok':       True,
        })

        if sigma_V_all is not None:
            print(f"Volume DoD (all cells):  {dod_all_net:+.2f} \u00b1 {sigma_V_all:.2f} m\u00b3  "
                  f"(\u00b1{rel_err_all:.1f} %)  [{int(valid.sum()):,} cells]")
            print(f"Volume DoD (sig cells):  {dod_sig_net:+.2f} \u00b1 {sigma_V_sig:.2f} m\u00b3  "
                  f"(\u00b1{rel_err_sig:.1f} %)  [{int(sig_cells.sum()):,} cells, |dZ|>{lod:.4f} m]")
        else:
            print(f"Volume DoD (all cells):  gain={dod_all_gain:+.2f} m\u00b3  "
                  f"loss={dod_all_loss:+.2f} m\u00b3  net={dod_all_net:+.2f} m\u00b3  "
                  f"({int(valid.sum()):,} cells)")
            print(f"Volume DoD (sig cells):  gain={dod_sig_gain:+.2f} m\u00b3  "
                  f"loss={dod_sig_loss:+.2f} m\u00b3  net={dod_sig_net:+.2f} m\u00b3  "
                  f"({int(sig_cells.sum()):,} cells, |dZ|>{lod:.4f} m)")
        if pct is not None:
            print(f"\u0394 DoD all vs sig: {pct:.1f} %")

    except Exception as e:
        print(f"  Volume DoD failed: {e}")

    if progress_mgr:
        progress_mgr.update_current(100, "Volume estimation complete")

    return result


# ═══════════════════════════════════════════════════════════════════════
# Histogram Export
# ═══════════════════════════════════════════════════════════════════════

def export_histogram_png(distances, significant, detection_limit,
                         output_file, progress_mgr=None):
    """
    Publication-ready M3C2 distance histogram PNG.

    Layout:
      - Full distribution as a grey-filled histogram (all valid points)
      - Significant changes overlaid as a coloured stacked histogram
        (negative = blue, positive = red)
      - Vertical ±LoD dashed lines
      - Annotations: n_total, n_significant, mean, median, LoD
    """
    try:
        import matplotlib.pyplot as plt
        import matplotlib.colors as mcolors
        import matplotlib.patches as mpatches
        from matplotlib.ticker import AutoMinorLocator
    except ImportError:
        _ui_error("Chyba", "Chýba matplotlib.\npip install matplotlib")
        return None

    if progress_mgr:
        progress_mgr.update_current(10, "Histogram: preparing data...")

    d = np.asarray(distances, dtype=np.float64)
    valid   = ~np.isnan(d)
    d_valid = d[valid]
    sig_v   = significant[valid].astype(bool)

    if len(d_valid) == 0:
        return None

    # Clip display range at 1st/99th percentile — avoids giant tails
    lo = float(np.percentile(d_valid, 0.5))
    hi = float(np.percentile(d_valid, 99.5))
    if abs(lo) > abs(hi):   # make symmetric around 0
        hi =  abs(lo)
    else:
        lo = -abs(hi)
    lo = min(lo, -detection_limit * 2.5)
    hi = max(hi,  detection_limit * 2.5)

    # Bin width: Freedman–Diaconis rule on clipped data
    clipped = d_valid[(d_valid >= lo) & (d_valid <= hi)]
    iqr  = float(np.subtract(*np.percentile(clipped, [75, 25])))
    bw   = max(2.0 * iqr * len(clipped) ** (-1/3), detection_limit / 10)
    bins = np.arange(lo, hi + bw, bw)

    if progress_mgr:
        progress_mgr.update_current(30, "Histogram: rendering figure...")

    fig, ax = plt.subplots(figsize=(12, 6), dpi=150)
    ax.set_facecolor('#f7f7f7')
    fig.patch.set_facecolor('white')
    ax.grid(True, which='major', color='white', linewidth=1.0, zorder=0)
    ax.grid(True, which='minor', color='#eeeeee', linewidth=0.4, zorder=0)
    ax.minorticks_on()
    for sp in ax.spines.values():
        sp.set_color('#bbbbbb')
    ax.tick_params(colors='#444444', labelsize=8)

    # ── All valid: grey background ─────────────────────────────────────
    ax.hist(d_valid, bins=bins, color='#cccccc', edgecolor='white',
            linewidth=0.3, zorder=2, label='Not significant')

    # ── Significant negative (loss, blue) ─────────────────────────────
    d_sig_neg = d_valid[sig_v & (d_valid < 0)]
    if len(d_sig_neg):
        ax.hist(d_sig_neg, bins=bins, color='#1565C0', edgecolor='white',
                linewidth=0.3, alpha=0.85, zorder=3, label='Significant (loss)')

    # ── Significant positive (gain, red) ──────────────────────────────
    d_sig_pos = d_valid[sig_v & (d_valid > 0)]
    if len(d_sig_pos):
        ax.hist(d_sig_pos, bins=bins, color='#C62828', edgecolor='white',
                linewidth=0.3, alpha=0.85, zorder=3, label='Significant (gain)')

    # ── LoD lines ─────────────────────────────────────────────────────
    ymax = ax.get_ylim()[1]
    if detection_limit > 0:
        ax.axvline(+detection_limit, color='#C62828', linewidth=1.3,
                   linestyle='--', zorder=6,
                   label=f'+LoD  {detection_limit:.4f} m')
        ax.axvline(-detection_limit, color='#1565C0', linewidth=1.3,
                   linestyle='--', zorder=6,
                   label=f'\u2212LoD  {detection_limit:.4f} m')
    ax.axvline(0, color='#444444', linewidth=0.8, zorder=5)

    # ── Shade LoD zone ─────────────────────────────────────────────────
    if detection_limit > 0:
        ax.axvspan(-detection_limit, detection_limit,
                   alpha=0.06, color='#888888', zorder=1, label='Below LoD')

    # ── Annotations ───────────────────────────────────────────────────
    n_sig_tot  = int(sig_v.sum())
    n_tot      = int(len(d_valid))
    mean_all   = float(np.mean(d_valid))
    med_all    = float(np.median(d_valid))
    mean_sig   = float(np.mean(d_valid[sig_v])) if n_sig_tot > 0 else 0.0

    stats_txt = (f"n valid = {n_tot:,}\n"
                 f"n significant = {n_sig_tot:,}  ({n_sig_tot/n_tot*100:.1f} %)\n"
                 f"mean (all) = {mean_all:.4f} m\n"
                 f"median (all) = {med_all:.4f} m\n"
                 f"mean (sig) = {mean_sig:.4f} m\n"
                 f"LoD = \xb1{detection_limit:.4f} m")
    ax.text(0.97, 0.97, stats_txt,
            transform=ax.transAxes,
            fontsize=7.5, va='top', ha='right', color='#333333',
            bbox=dict(boxstyle='round,pad=0.5', facecolor='white',
                      alpha=0.85, edgecolor='#cccccc'))

    ax.set_xlabel('M3C2 Distance  [m]', fontsize=9, color='#333333', labelpad=6)
    ax.set_ylabel('Number of Corepoints', fontsize=9, color='#333333', labelpad=6)
    ax.set_title('M3C2 Distance Distribution',
                 fontsize=11, fontweight='bold', color='#111111', pad=10, loc='left')
    ax.set_xlim(lo, hi)

    leg = ax.legend(fontsize=8, framealpha=0.95, edgecolor='#cccccc',
                    facecolor='white', loc='upper left', ncol=2)
    for txt in leg.get_texts():
        txt.set_color('#333333')

    fig.text(0.01, 0.005,
             'M3C2 Tool \u2013 CHECKPOINT13 \u2013 Samuel Leb\xf3',
             fontsize=7, color='#aaaaaa', va='bottom')

    plt.tight_layout()

    stem     = os.path.splitext(output_file)[0]
    out_path = f"{stem}_histogram.png"

    if progress_mgr:
        progress_mgr.update_current(85, "Histogram: saving PNG...")

    try:
        fig.savefig(out_path, bbox_inches='tight', dpi=150, facecolor='white')
        plt.close(fig)
        print(f"Histogram PNG: {out_path}")
    except Exception as e:
        _ui_error("Chyba", f"Histogram save error: {e}")
        plt.close(fig)
        return None

    if progress_mgr:
        progress_mgr.update_current(100, "Histogram complete")
    return out_path




def _get_system_info():
    """Collect CPU / RAM / GPU info. Returns a dict."""
    import platform
    info = {
        'os':        platform.system() + ' ' + platform.release(),
        'python':    platform.python_version(),
        'machine':   platform.machine(),
        'cpu_name':  platform.processor() or 'N/A',
        'cpu_cores': 'N/A', 'cpu_freq': 'N/A',
        'ram_total': 'N/A', 'ram_avail': 'N/A',
        'gpu':       'N/A',
    }
    try:
        import psutil
        info['cpu_cores'] = (f"{psutil.cpu_count(logical=False)} physical / "
                             f"{psutil.cpu_count(logical=True)} logical")
        freq = psutil.cpu_freq()
        if freq:
            info['cpu_freq'] = f"{freq.current:.0f} MHz  (max {freq.max:.0f} MHz)"
        vm = psutil.virtual_memory()
        info['ram_total'] = f"{vm.total / 2**30:.1f} GB"
        info['ram_avail'] = f"{vm.available / 2**30:.1f} GB"
    except Exception:
        pass
    try:
        import subprocess
        r = subprocess.run(
            ['nvidia-smi',
             '--query-gpu=name,memory.total,driver_version',
             '--format=csv,noheader,nounits'],
            capture_output=True, text=True, timeout=5)
        if r.returncode == 0 and r.stdout.strip():
            parts = r.stdout.strip().split(',')
            name = parts[0].strip()
            mem  = f"{int(float(parts[1].strip()))/1024:.1f} GB" if len(parts)>1 else '?'
            drv  = parts[2].strip() if len(parts)>2 else '?'
            info['gpu'] = f"{name}  ({mem})  driver {drv}"
    except Exception:
        pass
    return info


def _get_cloud_info(path):
    """Return dict with point count, file size, bbox for a LAS/LAZ file."""
    info = {'points':'N/A','size_mb':'N/A',
            'x_range':'N/A','y_range':'N/A','z_range':'N/A'}
    try:
        info['size_mb'] = f"{os.path.getsize(path)/2**20:.2f} MB"
        las = laspy.read(path)
        info['points']  = f"{len(las.x):,}"
        info['x_range'] = f"{float(las.x.min()):.2f} \u2013 {float(las.x.max()):.2f} m"
        info['y_range'] = f"{float(las.y.min()):.2f} \u2013 {float(las.y.max()):.2f} m"
        info['z_range'] = f"{float(las.z.min()):.2f} \u2013 {float(las.z.max()):.2f} m"
    except Exception:
        pass
    return info


def _make_profile_map_png(corepoints, distances, significant,
                           profile_lines, profile_width, tmp_path):
    """Top-down M3C2 map with profile corridors. Returns True on success."""
    try:
        import matplotlib.pyplot as plt
        import matplotlib.colors as mcolors
        import matplotlib.patches as mpatches
        from matplotlib.patches import Polygon as MplPolygon
    except ImportError:
        return False

    fig, ax = plt.subplots(figsize=(10, 8), dpi=150)
    ax.set_facecolor('#f4f4f4')
    ax.set_aspect('equal')

    valid = ~np.isnan(distances)
    not_s = valid & ~significant.astype(bool)
    is_s  = valid &  significant.astype(bool)
    d_sig = distances[is_s]

    abs_max = float(np.nanpercentile(np.abs(d_sig), 98)) if is_s.any() else 1.0
    if abs_max == 0: abs_max = 1.0
    norm_c = mcolors.TwoSlopeNorm(vmin=-abs_max, vcenter=0, vmax=abs_max)
    cmap   = plt.cm.RdBu_r

    if not_s.any():
        ax.scatter(corepoints[not_s,0], corepoints[not_s,1],
                   s=0.8, c='#cccccc', linewidths=0, rasterized=True, zorder=2)
    if is_s.any():
        ax.scatter(corepoints[is_s,0], corepoints[is_s,1],
                   s=1.2, c=d_sig, cmap=cmap, norm=norm_c,
                   linewidths=0, rasterized=True, zorder=3)

    if profile_lines:
        hw = profile_width / 2.0
        for lbl, pxy in profile_lines:
            x0,y0 = float(pxy[0,0]),float(pxy[0,1])
            x1,y1 = float(pxy[1,0]),float(pxy[1,1])
            dl = np.sqrt((x1-x0)**2+(y1-y0)**2)
            if dl < 1e-6: continue
            px = -(y1-y0)/dl*hw;  py = (x1-x0)/dl*hw
            ax.add_patch(MplPolygon(
                [[x0+px,y0+py],[x1+px,y1+py],[x1-px,y1-py],[x0-px,y0-py]],
                closed=True, facecolor='#FF4400', alpha=0.12,
                edgecolor='none', zorder=4))
            ax.plot([x0,x1],[y0,y1],'-',color='#FF4400',linewidth=2.0,zorder=5)
            ax.plot([x0,x1],[y0,y1],'o',color='#FF4400',markersize=5,zorder=6,
                    markeredgecolor='white',markeredgewidth=0.6)
            mx,my = (x0+x1)/2,(y0+y1)/2
            ax.annotate(f' {lbl}',xy=(mx,my),fontsize=7,color='#CC2200',
                        fontweight='bold',zorder=7,
                        bbox=dict(boxstyle='round,pad=0.2',
                                  facecolor='white',alpha=0.75,edgecolor='none'))

    sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm_c); sm.set_array([])
    cb = fig.colorbar(sm, ax=ax, fraction=0.025, pad=0.02, aspect=30)
    cb.set_label('M3C2 Distance [m]', fontsize=8)
    cb.ax.tick_params(labelsize=7)
    ax.set_xlabel('X [m]', fontsize=8); ax.set_ylabel('Y [m]', fontsize=8)
    ax.tick_params(labelsize=7)
    ax.grid(True, color='white', linewidth=0.5, zorder=0)
    ax.set_title('M3C2 Results Overview \u2013 Profile Locations',
                 fontsize=9, fontweight='bold', pad=8)
    handles = [mpatches.Patch(color='#cccccc',label='Not significant'),
               mpatches.Patch(color='#E24B4A',label='Significant'),
               mpatches.Patch(color='#FF4400',alpha=0.5,label='Profile corridor')]
    ax.legend(handles=handles, fontsize=7, loc='upper right', framealpha=0.9)

    # ── North arrow ───────────────────────────────────────────────────
    # placed in axes-fraction coords (bottom-left corner area)
    ax.annotate('', xy=(0.06, 0.12), xycoords='axes fraction',
                xytext=(0.06, 0.04), textcoords='axes fraction',
                arrowprops=dict(arrowstyle='->', color='#222222',
                                lw=1.5, mutation_scale=12), zorder=20)
    ax.text(0.06, 0.135, 'N', transform=ax.transAxes,
            ha='center', va='bottom', fontsize=9,
            fontweight='bold', color='#222222', zorder=20)

    # ── Scale bar ─────────────────────────────────────────────────────
    # auto-pick a round distance ≈ 15 % of x-span
    xlim_sb = ax.get_xlim()
    x_span  = xlim_sb[1] - xlim_sb[0]
    raw_sb  = x_span * 0.15
    # round to nearest clean value
    mag     = 10 ** int(np.floor(np.log10(raw_sb)))
    sb_len  = round(raw_sb / mag) * mag

    ylim_sb  = ax.get_ylim()
    y_span   = ylim_sb[1] - ylim_sb[0]
    sb_x0    = xlim_sb[0] + x_span * 0.05
    sb_x1    = sb_x0 + sb_len
    sb_y     = ylim_sb[0] + y_span * 0.04
    sb_yt    = ylim_sb[0] + y_span * 0.065
    sb_tick  = y_span * 0.008

    ax.plot([sb_x0, sb_x1], [sb_y, sb_y],
            '-', color='#222222', linewidth=2.0, zorder=20,
            solid_capstyle='butt')
    for xp in [sb_x0, sb_x1]:
        ax.plot([xp, xp], [sb_y - sb_tick, sb_y + sb_tick],
                '-', color='#222222', linewidth=1.5, zorder=20)

    sb_label = (f'{int(sb_len)} m' if sb_len >= 1
                else f'{sb_len*100:.0f} cm')
    ax.text((sb_x0 + sb_x1) / 2, sb_yt, sb_label,
            ha='center', va='bottom', fontsize=7,
            color='#222222', zorder=20,
            bbox=dict(boxstyle='round,pad=0.15',
                      facecolor='white', alpha=0.7, edgecolor='none'))

    plt.tight_layout()
    fig.savefig(tmp_path, bbox_inches='tight', dpi=150, facecolor='white')
    plt.close(fig)
    return True


def generate_report_pdf(output_path, report_data, progress_mgr=None):
    """
    Build a comprehensive multi-page A4 PDF report using reportlab.
    report_data is a dict assembled inside run_m3c2_gui.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, Image as RLImage, KeepTogether)
        from reportlab.platypus.flowables import HRFlowable
        import datetime
    except ImportError as e:
        _ui_error("Chyba", f"Report requires reportlab:\npip install reportlab\n\n{e}")
        return False

    if progress_mgr:
        progress_mgr.update_current(5, "Report: collecting data...")

    rd          = report_data
    sys_info    = _get_system_info()
    cloud1_info = _get_cloud_info(rd.get('las_file1',''))
    cloud2_info = _get_cloud_info(rd.get('las_file2',''))

    corepoints  = rd.get('corepoints')
    distances   = rd.get('distances')
    significant = rd.get('significant')
    valid_mask  = rd.get('valid_mask')
    nz  = rd.get('nz');  nxy = rd.get('nxy')
    dx  = rd.get('dx');  dy  = rd.get('dy'); dz = rd.get('dz')
    detection_limit = float(rd.get('detection_limit', 0))
    profile_lines   = rd.get('profile_lines') or []
    profile_width   = float(rd.get('profile_width', 50.0))
    saved_extras    = rd.get('saved_extras') or []
    start_time      = rd.get('start_time')
    end_time        = rd.get('end_time')

    n_total = int(len(distances)) if distances is not None else 0
    if valid_mask is not None and distances is not None:
        valid_d = distances[valid_mask & ~np.isnan(distances)]
    else:
        valid_d = np.array([])
    n_valid   = int(len(valid_d))
    n_invalid = n_total - n_valid
    n_sig     = 0
    if significant is not None and valid_mask is not None and distances is not None:
        vm_nonan = valid_mask & ~np.isnan(distances)
        n_sig = int(np.sum(significant[vm_nonan]))

    # ── Colours ──────────────────────────────────────────────────────
    BLUE      = colors.HexColor('#1565C0')
    BLUE_LIGHT= colors.HexColor('#E3F2FD')
    GREY1     = colors.HexColor('#333333')
    GREY2     = colors.HexColor('#666666')
    GREY3     = colors.HexColor('#AAAAAA')
    GREY_ROW  = colors.HexColor('#F5F5F5')
    WHITE     = colors.white

    W, H  = A4
    MARGIN = 2.0 * cm
    TW    = W - 2 * MARGIN

    # ── Style helpers ─────────────────────────────────────────────────
    def S(name, **kw):
        return ParagraphStyle(name, **kw)

    TITLE  = S('tt', fontName='Helvetica-Bold', fontSize=30, textColor=BLUE,
                spaceAfter=6, alignment=TA_CENTER)
    H1     = S('h1', fontName='Helvetica-Bold', fontSize=13, textColor=BLUE,
                spaceBefore=14, spaceAfter=5)
    H2     = S('h2', fontName='Helvetica-Bold', fontSize=10, textColor=GREY1,
                spaceBefore=8, spaceAfter=3)
    BODY   = S('bd', fontName='Helvetica', fontSize=9, textColor=GREY1,
                spaceAfter=4, leading=14, alignment=TA_JUSTIFY)
    SMALL  = S('sm', fontName='Helvetica', fontSize=8, textColor=GREY2, spaceAfter=2)
    CENTER = S('ct', fontName='Helvetica', fontSize=9, textColor=GREY1,
                alignment=TA_CENTER, spaceAfter=4)
    FORMULA= S('fm', fontName='Courier-Bold', fontSize=10,
                textColor=colors.HexColor('#1A237E'),
                backColor=colors.HexColor('#EEF2FF'),
                borderPad=8, spaceAfter=8, spaceBefore=8, alignment=TA_CENTER)
    CAPTION= S('cap', fontName='Helvetica-Oblique', fontSize=8, textColor=GREY2,
                alignment=TA_CENTER, spaceAfter=6)
    FOOT_S = S('ft', fontName='Helvetica', fontSize=7, textColor=GREY3,
                alignment=TA_CENTER)

    sp = lambda n=1: Spacer(1, n * 0.35 * cm)

    # ── Table style ───────────────────────────────────────────────────
    def _ts():
        return TableStyle([
            ('BACKGROUND',   (0,0),(-1,0), BLUE),
            ('TEXTCOLOR',    (0,0),(-1,0), WHITE),
            ('FONTNAME',     (0,0),(-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0,0),(-1,0), 9),
            ('ALIGN',        (0,0),(-1,0), 'CENTER'),
            ('BOTTOMPADDING',(0,0),(-1,0), 6),
            ('TOPPADDING',   (0,0),(-1,0), 6),
            ('FONTNAME',     (0,1),(-1,-1),'Helvetica'),
            ('FONTSIZE',     (0,1),(-1,-1), 8),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[WHITE, GREY_ROW]),
            ('GRID',         (0,0),(-1,-1), 0.4, colors.HexColor('#DDDDDD')),
            ('LEFTPADDING',  (0,0),(-1,-1), 6),
            ('RIGHTPADDING', (0,0),(-1,-1), 6),
            ('TOPPADDING',   (0,1),(-1,-1), 4),
            ('BOTTOMPADDING',(0,1),(-1,-1), 4),
        ])

    def _p(text, style=None, bold=False):
        st = style or BODY
        if bold:
            st = S('bx', fontName='Helvetica-Bold', fontSize=st.fontSize,
                   textColor=st.textColor)
        return Paragraph(str(text), st)

    def _tbl(rows_data, widths, hdr_row=True):
        """rows_data: list of lists of strings/Paragraphs."""
        built = []
        for ri, row in enumerate(rows_data):
            built_row = []
            for ci, cell in enumerate(row):
                if isinstance(cell, Paragraph):
                    built_row.append(cell)
                else:
                    fn = 'Helvetica-Bold' if ri == 0 else 'Helvetica'
                    tc = WHITE if ri == 0 else GREY1
                    built_row.append(Paragraph(str(cell),
                        S(f'c{ri}{ci}', fontName=fn, fontSize=8 if ri>0 else 9,
                          textColor=tc)))
            built.append(built_row)
        t = Table(built, colWidths=widths)
        t.setStyle(_ts())
        return t

    def _section(title):
        return [Paragraph(title, H1),
                HRFlowable(width='100%', thickness=1.5,
                           color=BLUE, spaceAfter=5)]

    # ════════════════════════════════════════════════════════════════
    # COVER
    # ════════════════════════════════════════════════════════════════
    story = []
    story += [sp(4), Paragraph('M3C2 Tool', TITLE)]
    story.append(Paragraph('Analysis Report',
        S('sub', fontName='Helvetica-Bold', fontSize=16, textColor=GREY2,
          spaceAfter=4, alignment=TA_CENTER)))
    story.append(sp(0.5))
    story.append(HRFlowable(width='50%', thickness=2, color=BLUE,
                             spaceAfter=10, spaceBefore=0))
    story.append(Paragraph('CHECKPOINT  13',
        S('ck', fontName='Helvetica-Bold', fontSize=11, textColor=WHITE,
          backColor=BLUE, borderPad=6, spaceAfter=8, alignment=TA_CENTER)))
    story.append(sp(0.5))
    story.append(Paragraph('Point Cloud Change Detection  \u00b7  M3C2 Algorithm', CENTER))
    story.append(Paragraph('py4dgeo  \u00b7  laspy  \u00b7  NumPy  \u00b7  SciPy  \u00b7  matplotlib',
        S('libs', fontName='Helvetica', fontSize=8, textColor=GREY3,
          alignment=TA_CENTER, spaceAfter=6)))
    story.append(sp(3))

    now_str = datetime.datetime.now().strftime('%Y-%m-%d  %H:%M')
    dur_str = 'N/A'
    if start_time and end_time:
        secs = int((end_time - start_time).total_seconds())
        dur_str = f"{secs//60} min {secs%60} s"

    cover_rows = [
        ['Author',    'Samuel Leb\xf3'],
        ['Date',       now_str],
        ['Processing time', dur_str],
        ['Version',   'CHECKPOINT 13'],
        ['GitHub',    'github.com/Kostka22/Geology_M3C2-from-point-cloud'],
    ]
    ct = Table(
        [[_p(r[0], S('ck', fontName='Helvetica-Bold', fontSize=9, textColor=BLUE)),
          _p(r[1])] for r in cover_rows],
        colWidths=[4*cm, TW-4*cm])
    ct.setStyle(TableStyle([
        ('ALIGN',          (0,0),(-1,-1),'LEFT'),
        ('TOPPADDING',     (0,0),(-1,-1), 5),
        ('BOTTOMPADDING',  (0,0),(-1,-1), 5),
        ('LEFTPADDING',    (0,0),(-1,-1), 8),
        ('ROWBACKGROUNDS', (0,0),(-1,-1),[BLUE_LIGHT, WHITE]),
        ('BOX',            (0,0),(-1,-1), 0.5, colors.HexColor('#CCCCCC')),
    ]))
    story += [ct, PageBreak()]

    # ════════════════════════════════════════════════════════════════
    # 1. SYSTEM INFO
    # ════════════════════════════════════════════════════════════════
    if progress_mgr:
        progress_mgr.update_current(15, "Report: system info...")

    story += _section('1.  System Information')
    story.append(sp(0.3))
    story.append(_tbl([
        ['Parameter', 'Value'],
        ['Operating System',  sys_info['os']],
        ['Python',            sys_info['python']],
        ['Architecture',      sys_info['machine']],
        ['CPU',               sys_info['cpu_name']],
        ['CPU Cores',         sys_info['cpu_cores']],
        ['CPU Frequency',     sys_info['cpu_freq']],
        ['RAM Total',         sys_info['ram_total']],
        ['RAM Available',     sys_info['ram_avail']],
        ['GPU',               sys_info['gpu']],
    ], [5*cm, TW-5*cm]))
    story.append(sp())

    # ── 2. INPUT DATA ─────────────────────────────────────────────────
    story += _section('2.  Input Data')
    story.append(sp(0.3))

    for cloud_label, path, info in [
        ('Reference Cloud (Epoch 1)', rd.get('las_file1',''), cloud1_info),
        ('Comparison Cloud (Epoch 2)', rd.get('las_file2',''), cloud2_info),
    ]:
        story.append(Paragraph(cloud_label, H2))
        story.append(_tbl([
            ['Property',    'Value'],
            ['File name',    os.path.basename(path)],
            ['Full path',    path],
            ['File size',    info['size_mb']],
            ['Point count',  info['points']],
            ['X range',      info['x_range']],
            ['Y range',      info['y_range']],
            ['Z range',      info['z_range']],
        ], [4*cm, TW-4*cm]))
        story.append(sp(0.4))

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════
    # 3. PARAMETERS
    # ════════════════════════════════════════════════════════════════
    story += _section('3.  Processing Parameters')
    story.append(sp(0.3))

    _mode = rd.get('sigma_mode', 'per_point')
    if _mode == 'manual_sigma':
        sigma_mode = f'C \u2014 Manual \u03c31={rd.get("sigma1","?")} m, \u03c32={rd.get("sigma2","?")} m'
    elif _mode == 'global_auto':
        sigma_mode = 'B \u2014 Global auto (py4dgeo SPREAD1/2)'
    else:
        sigma_mode = f'A \u2014 Per-point LoD (reg_error={rd.get("registration_error", 0.0):.4f} m)'
    cp_mode    = ('All points' if not rd.get('core_spacing')
                  else f"{rd.get('core_spacing')} m greedy distance filter")

    param_rows = [
        ['Parameter', 'Value', 'Notes'],
        ['Computation engine',
         rd.get('engine', 'py4dgeo'),
         'py4dgeo = per-point LoD; PDAL = OpenMP multi-core'],
        ['Cylinder radius',
         f"{rd.get('cyl_radius','N/A')} m",
         'Radius of projection cylinder'],
        ['Normal radii',
         str(rd.get('normal_radii','N/A')),
         'Multi-scale normal estimation'],
        ['Corepoint spacing', cp_mode,
         'Spatial subsampling'],
        ['Max depth filter',
         (f"{rd.get('max_depth')} m" if rd.get('max_depth') else 'Disabled'),
         'Outlier distance threshold'],
        ['Confidence level',
         f"{float(rd.get('confidence_level',0.95))*100:.0f} %",
         'Significance test threshold'],
        ['Sigma mode',      sigma_mode, ''],
        ['Sigma 1',
         f"{float(rd.get('sigma1',0)):.4f} m",
         'Epoch 1 position uncertainty'],
        ['Sigma 2',
         f"{float(rd.get('sigma2',0)):.4f} m",
         'Epoch 2 position uncertainty'],
        ['Level of Detection',
         f"{detection_limit:.4f} m",
         't \xd7 sqrt(\u03c3\xb1\xb2 + \u03c3\xb2\xb2)'],
    ]
    if rd.get('make_raster'):
        param_rows += [
            ['Raster resolution',
             f"{rd.get('raster_resolution','N/A')} m", ''],
            ['Raster interpolation',
             rd.get('raster_interp_method','N/A'), ''],
        ]
    if rd.get('make_dem'):
        param_rows.append(
            ['DEM resolution', f"{rd.get('dem_resolution','N/A')} m", ''])
    if rd.get('clip_mask_path'):
        param_rows.append(
            ['Clip mask', rd.get('clip_mask_path',''), 'Applied to all GeoTIFFs'])
    quiver_modes = []
    if rd.get('quiver_normal'): quiver_modes.append('Normal-direction')
    if rd.get('quiver_motion'): quiver_modes.append('Motion vector')
    if quiver_modes:
        param_rows.append(
            ['Quiver export', ', '.join(quiver_modes),
             f"Grid cell: {rd.get('quiver_grid_size','N/A')} m"])

    story.append(_tbl(param_rows, [4.5*cm, 3.5*cm, TW-8*cm]))
    story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════
    # 4. ALGORITHM EXPLANATION
    # ════════════════════════════════════════════════════════════════
    story += _section('4.  M3C2 Algorithm \u2013 How It Works')
    story.append(sp(0.3))
    story.append(Paragraph(
        'The <b>Multiscale Model to Model Cloud Comparison (M3C2)</b> algorithm '
        '(Lague, Brodu &amp; Leroux, 2013) computes signed 3-D distances between two '
        'point clouds directly along the local surface normal. Unlike raster DoD, '
        'M3C2 works on any surface orientation \u2013 flat terrain, steep slopes, '
        'vertical faces, and overhangs.', BODY))
    story.append(sp(0.3))

    for step_title, step_body in [
        ('1 \u2014 Normal estimation',
         'At each corepoint a plane is fitted to the k nearest neighbours across '
         'multiple scales (the Normal radii list). The scale that minimises local '
         'roughness is automatically selected per point.'),
        ('2 \u2014 Projection cylinder',
         'A cylinder of radius r (Cylinder radius) is projected along the normal. '
         'All points from both clouds that fall inside are collected.'),
        ('3 \u2014 M3C2 distance',
         'The mean position of each cloud inside the cylinder is computed. '
         'The signed distance along the normal between those two means is the '
         'M3C2 distance d. Positive = surface moved away from the reference; '
         'negative = surface moved toward it.'),
        ('4 \u2014 Uncertainty (Level of Detection)',
         'The LoD combines point cloud roughness (SPREAD1, SPREAD2) and '
         'registration error (sigma). A change is flagged significant only if '
         '|d| > t \xd7 sqrt(\u03c31\xb2 + \u03c32\xb2).'),
        ('5 \u2014 Displacement components',
         'DX = d\xd7nx, DY = d\xd7ny, DZ = d\xd7nz decompose the M3C2 distance '
         'onto the Cartesian axes of the unit normal vector, giving physically '
         'interpretable east, north, and vertical components.'),
    ]:
        story.append(KeepTogether([
            Paragraph(step_title, H2),
            Paragraph(step_body, BODY),
        ]))

    story.append(sp(0.3))
    story.append(Paragraph('Significance test:', H2))
    story.append(Paragraph(
        '|M3C2 distance| > t_critical \xd7 sqrt(\u03c31\xb2 + \u03c32\xb2)',
        FORMULA))
    story.append(_tbl([
        ['Confidence Level', 't-critical', 'Interpretation'],
        ['90 %', '1.645', 'Less conservative'],
        ['95 %', '1.960', 'Standard (default)'],
        ['99 %', '2.576', 'Most conservative'],
    ], [4*cm, 3.5*cm, TW-7.5*cm]))
    story.append(sp(0.4))
    story.append(Paragraph(
        '<b>Reference:</b> Lague, D., Brodu, N., Leroux, J. (2013). '
        'Accurate 3D comparison of complex topography with terrestrial laser '
        'scanner: Application to the Rangitikei canyon (N-Z). '
        '<i>ISPRS Journal of Photogrammetry and Remote Sensing</i>, 82, 10\u201326.',
        BODY))
    story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════
    # 5. RESULTS STATISTICS
    # ════════════════════════════════════════════════════════════════
    if progress_mgr:
        progress_mgr.update_current(40, "Report: statistics...")

    story += _section('5.  Results Statistics')
    story.append(sp(0.3))

    pct_sig_val = n_sig/n_valid*100 if n_valid > 0 else 0
    pct_sig_str = f"{pct_sig_val:.1f} %"

    story.append(Paragraph('5.1  Overview', H2))
    story.append(_tbl([
        ['Metric', 'Value'],
        ['Total corepoints',          f"{n_total:,}"],
        ['Valid corepoints',          f"{n_valid:,}"],
        ['Invalid (NaN / filtered)',   f"{n_invalid:,}"],
        ['Significant changes',        f"{n_sig:,}   ({pct_sig_str})"],
        ['Not significant',
         f"{n_valid-n_sig:,}   ({100-pct_sig_val:.1f} %)"],
        ['Level of Detection (LoD)',   f"{detection_limit:.4f} m"],
        ['Confidence level',
         f"{float(rd.get('confidence_level',0.95))*100:.0f} %"],
        ['Sigma 1',  f"{float(rd.get('sigma1',0)):.4f} m"],
        ['Sigma 2',  f"{float(rd.get('sigma2',0)):.4f} m"],
    ], [6*cm, TW-6*cm]))
    story.append(sp(0.4))

    if len(valid_d) > 0:
        sig_d = valid_d[np.abs(valid_d) > detection_limit] if detection_limit > 0 else valid_d

        story.append(Paragraph('5.2  Distance Statistics', H2))

        def _f(arr, fn):
            try: return f"{fn(arr):.4f} m"
            except: return 'N/A'

        story.append(_tbl([
            ['Statistic',  'All valid', 'Significant only'],
            ['Minimum',    _f(valid_d, np.min),    _f(sig_d, np.min)],
            ['Maximum',    _f(valid_d, np.max),    _f(sig_d, np.max)],
            ['Mean',       _f(valid_d, np.mean),   _f(sig_d, np.mean)],
            ['Median',     _f(valid_d, np.median), _f(sig_d, np.median)],
            ['Std Dev',    _f(valid_d, np.std),    _f(sig_d, np.std)],
            ['RMS',
             _f(valid_d, lambda a: np.sqrt(np.mean(a**2))),
             _f(sig_d,   lambda a: np.sqrt(np.mean(a**2)))],
        ], [4.5*cm, (TW-4.5*cm)/2, (TW-4.5*cm)/2]))
        story.append(sp(0.4))

        story.append(Paragraph('5.3  Change Magnitude Distribution', H2))
        bins   = [0, 0.1, 0.25, 0.5, 1.0, 2.0, 5.0, 10.0, np.inf]
        labels = ['< 0.10 m','0.10\u20130.25 m','0.25\u20130.50 m',
                  '0.50\u20131.00 m','1.00\u20132.00 m','2.00\u20135.00 m',
                  '5.00\u201310.0 m','> 10.0 m']
        hist, _ = np.histogram(np.abs(valid_d), bins=bins)
        cum = 0.0
        dist_rows = [['Range','Count','% of valid','Cumulative %']]
        for lbl, cnt in zip(labels, hist):
            pct = cnt/n_valid*100 if n_valid > 0 else 0
            cum += pct
            dist_rows.append([lbl, f"{cnt:,}", f"{pct:.1f} %", f"{cum:.1f} %"])
        story.append(_tbl(dist_rows, [3.5*cm, 3*cm, 3*cm, 3*cm]))
        story.append(sp(0.4))

    if nz is not None:
        story.append(Paragraph('5.4  Normal Vector & Displacement Components', H2))
        nv_rows = [['Field','Min','Max','Mean','Description']]
        for name, arr, desc in [
            ('Nz',  nz,  '1.0=horizontal, 0.0=vertical'),
            ('Nxy', nxy, '0.0=horizontal, 1.0=vertical'),
        ]:
            if arr is not None:
                nv_rows.append([name,
                    f"{float(arr.min()):.3f}", f"{float(arr.max()):.3f}",
                    f"{float(arr.mean()):.3f}", desc])
        for name, arr, desc in [
            ('DX', dx, 'East displacement [m]'),
            ('DY', dy, 'North displacement [m]'),
            ('DZ', dz, 'Vertical displacement [m]'),
        ]:
            if arr is not None:
                a = np.asarray(arr)
                a = a[~np.isnan(a)]
                if len(a) > 0:
                    nv_rows.append([name,
                        f"{float(a.min()):.4f}", f"{float(a.max()):.4f}",
                        f"{float(a.mean()):.4f}", desc])
        story.append(_tbl(nv_rows,
            [1.5*cm, 2*cm, 2*cm, 2*cm, TW-7.5*cm]))

    # ── 5.5 Surface change classification ─────────────────────────────
    if len(valid_d) > 0 and significant is not None:
        story.append(sp(0.4))
        story.append(Paragraph('5.5  Surface Change Classification', H2))
        story.append(Paragraph(
            'Significant changes split into erosion (negative M3C2) and '
            'deposition (positive M3C2) classes. Areas estimated as '
            'n_points \xd7 cell_size\xb2 using the raster resolution parameter.',
            BODY))
        story.append(sp(0.3))

        vm_sig = valid_mask & ~np.isnan(distances) & significant.astype(bool) \
                 if valid_mask is not None else \
                 ~np.isnan(distances) & significant.astype(bool)

        d_sig_all = distances[vm_sig]
        d_ero = d_sig_all[d_sig_all < 0]
        d_dep = d_sig_all[d_sig_all > 0]

        cell_a = float(rd.get('raster_resolution', 1.0)) ** 2

        def _cls(arr, label):
            if len(arr) == 0:
                return [label, '0', '0.00 m²', '—', '—', '—', '—']
            area = len(arr) * cell_a
            return [
                label,
                f"{len(arr):,}",
                f"{area:,.2f} m\xb2",
                f"{float(np.mean(arr)):.4f} m",
                f"{float(np.median(arr)):.4f} m",
                f"{float(np.min(arr)):.4f} m",
                f"{float(np.max(arr)):.4f} m",
            ]

        cls_rows = [
            ['Class', 'Points', 'Est. area', 'Mean dist', 'Median dist', 'Min dist', 'Max dist'],
            _cls(d_ero, 'Erosion (loss)'),
            _cls(d_dep, 'Deposition (gain)'),
            _cls(d_sig_all, 'Total significant'),
        ]
        story.append(_tbl(cls_rows,
            [2.8*cm, 1.8*cm, 2.2*cm, 2.2*cm, 2.2*cm, 2.2*cm, TW-13.4*cm]))

        if len(d_ero) > 0 and len(d_dep) > 0:
            ratio = len(d_dep) / len(d_ero)
            story.append(sp(0.3))
            story.append(Paragraph(
                f'<b>Deposition / erosion point ratio:</b> {ratio:.2f}  '
                f'(\u2014  {"net deposition dominant" if ratio > 1 else "net erosion dominant"})',
                BODY))

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════
    # 6. HISTOGRAM
    # ════════════════════════════════════════════════════════════════
    histogram_pngs = [f for f in saved_extras
                      if '_histogram' in f and f.endswith('.png')]
    if histogram_pngs and os.path.exists(histogram_pngs[0]):
        if progress_mgr:
            progress_mgr.update_current(52, "Report: embedding histogram...")
        story += _section('6.  M3C2 Distance Histogram')
        story.append(Paragraph(
            'Grey bars = all valid corepoints. Blue = significant negative changes '
            '(surface loss). Red = significant positive changes (surface gain). '
            'Dashed lines = \xb1Level of Detection. Shaded band = below LoD zone.',
            BODY))
        story.append(sp(0.3))
        story.append(RLImage(histogram_pngs[0], width=TW, height=TW * 0.52))
        story.append(Paragraph(
            f'Figure.  {os.path.basename(histogram_pngs[0])}', CAPTION))
        story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════
    # 7. VOLUME CHANGE
    # ════════════════════════════════════════════════════════════════
    vol = rd.get('volume_result')
    dem_interp_str = rd.get('volume_dem_interp', 'IDW')
    if vol is not None:
        if progress_mgr:
            progress_mgr.update_current(56, "Report: volume change section...")
        story += _section('7.  Volume Change Estimation')
        story.append(sp(0.3))
        story.append(Paragraph(
            'Two DoD methods computed using the same interpolated raster. '
            'Both clouds are thinned to one point per grid cell before interpolation '
            'for speed, then subtracted cell by cell.',
            BODY))
        story.append(sp(0.3))

        for title, body in [
            ('Method A \u2014 DoD all cells',
             f'Both clouds rasterised to a {vol["dod_res"]} m grid ({dem_interp_str}). '
             'dZ = Z\u2082 \u2212 Z\u2081. All non-NaN cells contribute. '
             'Comparable to CloudCompare\'s Compute 2.5D Volume.'),
            ('Method B \u2014 DoD significant cells only',
             f'Same raster as Method A. Only cells where |dZ| > LoD '
             f'({vol["dod_lod"]:.4f} m) contribute. '
             'Removes sub-noise cells \u2014 DoD equivalent of the M3C2 significance filter.'),
            ('Volume uncertainty \u2014 error propagation',
             'Per-cell uncertainty is propagated into total volume via quadrature: '
             '<b>\u03c3_V = \u221a[\u03a3 (LoD \xd7 A_cell)\xb2] = LoD \xd7 A_cell \xd7 \u221aN_cells</b>. '
             'This gives a statistically valid uncertainty assuming that individual cell errors '
             'are independent. The relative uncertainty (\u03c3_V / |V|) indicates how much of '
             'the measured volume is real signal versus accumulated noise. '
             'Typical values for landslide analysis (Bernard et al., 2021): 10\u201325 %. '
             'Values above 50 % indicate that the measured change is statistically '
             'indistinguishable from the noise floor.'),
        ]:
            story.append(KeepTogether([
                Paragraph(title, H2),
                Paragraph(body, BODY),
            ]))

        story.append(sp(0.4))
        story.append(Paragraph('Results Comparison', H2))

        def _vm(v, unit='m\xb3'):
            if v is None: return 'N/A'
            return f"{v:+.3f} {unit}"

        def _vm_err(v, sig, rel, unit='m\xb3'):
            if v is None: return 'N/A'
            if sig is None: return f"{v:+.2f} {unit}"
            if rel is not None:
                return f"{v:+.1f} \u00b1 {sig:.1f} {unit}  (\u00b1{rel:.1f} %)"
            return f"{v:+.1f} \u00b1 {sig:.1f} {unit}"

        n_dod_all = vol.get('n_all_cells', 0) or 0
        n_dod_sig = vol.get('n_sig_cells', 0) or 0

        vol_rows = [
            ['Metric',                    'DoD  all cells',              'DoD  sig cells only'],
            ['Volume gain  (+)',          _vm(vol.get('dod_all_gain')),  _vm(vol.get('dod_sig_gain'))],
            ['Volume loss  (\u2212)',     _vm(vol.get('dod_all_loss')),  _vm(vol.get('dod_sig_loss'))],
            ['Net volume change',         _vm(vol.get('dod_all_net')),   _vm(vol.get('dod_sig_net'))],
            ['Net  \u00b1  \u03c3_V  (error propagation)',
             _vm_err(vol.get('dod_all_net'), vol.get('sigma_V_all'), vol.get('rel_err_all')),
             _vm_err(vol.get('dod_sig_net'), vol.get('sigma_V_sig'), vol.get('rel_err_sig'))],
            ['Grid cells used',           f"{n_dod_all:,}",              f"{n_dod_sig:,}"],
            ['LoD threshold',             'none',                         f"|dZ| > {vol['dod_lod']:.4f} m"],
        ]
        story.append(_tbl(vol_rows, [5*cm, (TW-5*cm)/2, (TW-5*cm)/2]))
        story.append(sp(0.4))

        # Percent difference
        pct = None
        a, b = vol.get('dod_all_net'), vol.get('dod_sig_net')
        if a is not None and b is not None:
            denom = (abs(a) + abs(b)) / 2
            if denom > 0:
                pct = abs(a - b) / denom * 100
        if pct is not None:
            clr  = '#2E7D32' if pct < 10 else '#E65100' if pct < 25 else '#C62828'
            note = 'good agreement' if pct < 10 else 'moderate' if pct < 25 else 'large discrepancy'
            story.append(Paragraph(
                f'<b>Difference between methods: {pct:.1f} %</b>  \u2014  {note}.',
                S('pvol', fontName='Helvetica-Bold', fontSize=10,
                  textColor=colors.HexColor(clr),
                  backColor=colors.HexColor('#F5F5F5'),
                  borderPad=6, spaceAfter=6, spaceBefore=2,
                  alignment=TA_CENTER)))

        story.append(sp(0.3))
        story.append(Paragraph(
            '<b>Interpretation:</b> Method A is comparable to CloudCompare\'s '
            'Compute 2.5D Volume. Method B removes sub-noise cells using the '
            'same LoD threshold as the M3C2 significance test. '
            'A small difference between methods means most volume change is real signal.',
            BODY))
        story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════
    # 8. PROFILE MAP  (was 6)
    # ════════════════════════════════════════════════════════════════
    if progress_mgr:
        progress_mgr.update_current(60, "Report: generating profile map...")

    story += _section('8.  Profile Location Map')
    story.append(sp(0.3))
    story.append(Paragraph(
        'Top-down overview of all M3C2 corepoints. Grey = not significant. '
        'Coloured = significant (RdBu_r scale). '
        'Orange-red lines and shaded bands mark the profile corridor locations.',
        BODY))
    story.append(sp(0.3))

    tmp_map = output_path.replace('.pdf', '_tmp_map.png')
    map_ok  = (corepoints is not None and distances is not None and
               _make_profile_map_png(corepoints, distances, significant,
                                     profile_lines, profile_width, tmp_map))
    if map_ok and os.path.exists(tmp_map):
        story.append(RLImage(tmp_map, width=TW, height=TW * 0.8))
        story.append(Paragraph(
            f'Figure 1.  M3C2 overview with {len(profile_lines)} profile line(s).',
            CAPTION))
    else:
        story.append(Paragraph('Map not available.', SMALL))

    if profile_lines:
        story.append(sp(0.4))
        story.append(Paragraph('Profile Lines', H2))
        pl_rows = [['#','Label','Start (X, Y)','End (X, Y)','Length [m]','Corridor']]
        for i, (lbl, pxy) in enumerate(profile_lines):
            x0,y0 = float(pxy[0,0]), float(pxy[0,1])
            x1,y1 = float(pxy[1,0]), float(pxy[1,1])
            ln = float(np.linalg.norm(pxy[1]-pxy[0]))
            pl_rows.append([
                str(i+1), lbl,
                f"({x0:.1f},  {y0:.1f})",
                f"({x1:.1f},  {y1:.1f})",
                f"{ln:.2f}",
                f"\xb1{profile_width/2:.1f} m",
            ])
        story.append(_tbl(pl_rows,
            [1*cm, 2.5*cm, 3.5*cm, 3.5*cm, 2*cm, TW-12.5*cm]))

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════
    # 7. PROFILE PNGs
    # ════════════════════════════════════════════════════════════════
    if progress_mgr:
        progress_mgr.update_current(72, "Report: embedding profiles...")

    profile_pngs = [f for f in saved_extras
                    if '_profile_' in f and f.endswith('.png')]
    if profile_pngs:
        story += _section('9.  Profile Cross-Sections')
        story.append(Paragraph(
            'Upper panel: reference cloud (blue) and comparison cloud (red) '
            'projected into the profile corridor. '
            'Lower panel: M3C2 distance with \xb1LoD threshold lines. '
            'Grey points = not significant; coloured = significant.',
            BODY))
        story.append(sp(0.3))
        for i, png in enumerate(profile_pngs):
            if os.path.exists(png):
                story.append(KeepTogether([
                    RLImage(png, width=TW, height=TW * 9/16),
                    Paragraph(f'Figure {i+2}.  {os.path.basename(png)}',
                              CAPTION),
                    sp(0.4),
                ]))
        story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════
    # 8. WIND ROSE
    # ════════════════════════════════════════════════════════════════
    windrose_pngs = [f for f in saved_extras
                     if '_windrose' in f and f.endswith('.png')]
    if windrose_pngs:
        story += _section('10.  Wind Rose \u2013 Displacement Direction & Magnitude')
        story.append(Paragraph(
            'Panel (a): horizontal azimuth rose (0\xb0 = North, clockwise). '
            'Panel (b): vertical angle from horizontal. '
            'Four exports: all points (grey = not significant, coloured = significant), '
            'all significant only, '
            'positive changes only (gain / deposition), '
            'negative changes only (loss / erosion).',
            BODY))
        story.append(sp(0.3))
        captions = {
            '_windrose.png':     'All points \u2013 grey = not significant, coloured = significant',
            '_windrose_sig.png': 'All significant changes only',
            '_windrose_pos.png': 'Positive changes only (gain / deposition)',
            '_windrose_neg.png': 'Negative changes only (loss / erosion)',
        }
        for png in windrose_pngs:
            if os.path.exists(png):
                base = os.path.basename(png)
                lbl  = next((v for k, v in captions.items()
                             if base.endswith(k)), base)
                story.append(RLImage(png, width=TW, height=TW * 0.55))
                story.append(Paragraph(
                    f'Figure.  {lbl}  ({base})', CAPTION))
                story.append(sp(0.3))
        story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════
    # 9. OUTPUT FILES
    # ════════════════════════════════════════════════════════════════
    if progress_mgr:
        progress_mgr.update_current(87, "Report: output file list...")

    story += _section('11.  Output Files')
    story.append(sp(0.3))

    _desc_map = {
        '.laz':                'Corepoint cloud with M3C2 scalar fields (LAZ)',
        '.las':                'Corepoint cloud with M3C2 scalar fields',
        '_M3C2_DISTANCE.tif':  'GeoTIFF \u2013 M3C2 signed distance',
        '_SIGNIFICANT.tif':    'GeoTIFF \u2013 significance flag (1/0)',
        '_Nz.tif':             'GeoTIFF \u2013 normal Z component',
        '_Nxy.tif':            'GeoTIFF \u2013 normal XY magnitude',
        '_DX.tif':             'GeoTIFF \u2013 X displacement component',
        '_DY.tif':             'GeoTIFF \u2013 Y displacement component',
        '_DZ.tif':             'GeoTIFF \u2013 Z displacement component',
        '_NORMAL_SCALE.tif':   'GeoTIFF \u2013 normal estimation radius per corepoint [m]',
        '_LOD.tif':            'GeoTIFF \u2013 Level of Detection',
        '_SPREAD1.tif':        'GeoTIFF \u2013 Epoch 1 roughness',
        '_SPREAD2.tif':        'GeoTIFF \u2013 Epoch 2 roughness',
        '_DEM.tif':            'GeoTIFF \u2013 Digital Elevation Model',
        '_quiver_normal.png':  'Quiver PNG \u2013 normal-direction',
        '_quiver_motion.png':  'Quiver PNG \u2013 motion vector',
        '_quiver_normal.html': 'Quiver interactive HTML \u2013 normal-direction',
        '_quiver_motion.html': 'Quiver interactive HTML \u2013 motion vector',
        '_quiver_normal.gpkg': 'Quiver GPKG \u2013 normal-direction arrows',
        '_quiver_motion.gpkg': 'Quiver GPKG \u2013 motion vector arrows',
        '_windrose.png':       'Wind rose \u2013 all points (grey + coloured)',
        '_windrose_sig.png':   'Wind rose \u2013 all significant changes only',
        '_windrose_pos.png':   'Wind rose \u2013 positive changes only (gain / deposition)',
        '_windrose_neg.png':   'Wind rose \u2013 negative changes only (loss / erosion)',
        '_histogram.png':      'M3C2 distance histogram PNG',
        '_profiles.txt':       'Profile line trajectories \u2013 reloadable into tool',
        '_profiles.gpkg':      'Profile line trajectories \u2013 QGIS GeoPackage',
        '_results.txt':        'Tab-delimited text export of all scalar fields',
        '_report.pdf':         'This analysis report',
    }

    def _fdesc(path):
        base = os.path.basename(path)
        # Sort by suffix length descending so longer matches take priority
        # e.g. _windrose_sig.png before _windrose.png
        for suf, desc in sorted(_desc_map.items(),
                                 key=lambda x: len(x[0]), reverse=True):
            if base.endswith(suf):
                return desc
        if '_profile_' in base and base.endswith('.png'):
            return 'Profile cross-section PNG'
        return '\u2014'

    all_files = ([rd.get('output_file','')] +
                 saved_extras + [output_path])
    all_files  = list(dict.fromkeys(f for f in all_files if f))

    file_rows = [['File name', 'Size', 'Description']]
    for fpath in all_files:
        if not fpath or not os.path.exists(fpath): continue
        kb = os.path.getsize(fpath) / 1024
        sz = f"{kb/1024:.2f} MB" if kb > 1024 else f"{kb:.0f} KB"
        file_rows.append([os.path.basename(fpath), sz, _fdesc(fpath)])

    story.append(_tbl(file_rows, [7*cm, 2*cm, TW-9*cm]))
    story.append(sp())
    story.append(HRFlowable(width='100%', thickness=0.5,
                             color=GREY3, spaceAfter=5))
    story.append(Paragraph(
        'M3C2 Tool \u2013 CHECKPOINT 13 \u2013 Samuel Leb\xf3 \u2013 '
        'github.com/Kostka22/Geology_M3C2-from-point-cloud',
        FOOT_S))

    # ── Build ─────────────────────────────────────────────────────────
    if progress_mgr:
        progress_mgr.update_current(93, "Report: writing PDF...")

    def _page(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(GREY3)
        canvas.drawString(MARGIN, 0.7*cm,
                          'M3C2 Tool \u2013 CHECKPOINT 13 \u2013 Samuel Leb\xf3')
        canvas.drawRightString(W - MARGIN, 0.7*cm,
                               f'Page {doc.page}')
        canvas.restoreState()

    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=2.0*cm, bottomMargin=1.8*cm,
        title='M3C2 Analysis Report \u2013 CHECKPOINT 13',
        author='Samuel Leb\xf3',
        subject='M3C2 Point Cloud Change Detection',
    )
    doc.build(story, onFirstPage=_page, onLaterPages=_page)

    # Clean up temp map
    if os.path.exists(tmp_map):
        try: os.remove(tmp_map)
        except Exception: pass

    print(f"Report PDF: {output_path}")
    if progress_mgr:
        progress_mgr.update_current(100, "Report complete")
    return True


# ═══════════════════════════════════════════════════════════════════════
# PDAL engine
# ═══════════════════════════════════════════════════════════════════════

def _check_pdal():
    """Return (available: bool, version_str: str)."""
    import subprocess
    try:
        r = subprocess.run(['pdal', '--version'],
                           capture_output=True, text=True, timeout=8)
        if r.returncode == 0:
            ver = r.stdout.strip().split('\n')[0]
            return True, ver
        return False, 'pdal returned non-zero exit code'
    except FileNotFoundError:
        return False, 'pdal not found on PATH'
    except Exception as e:
        return False, str(e)


def run_m3c2_pdal(las_file1, las_file2, output_file,
                  cyl_radius, normal_radii, core_spacing,
                  max_depth, confidence_level,
                  use_manual_sigma, sigma1_manual, sigma2_manual,
                  sigma_mode='per_point', registration_error=0.0,
                  progress_mgr=None):
    """
    Run M3C2 via PDAL's filters.m3c2 (OpenMP multi-core).

    Returns the same variables as the py4dgeo path so the rest of
    run_m3c2_gui can proceed identically:
        corepoints, distances, uncertainties (None — PDAL doesn't give per-point),
        ref_points, epoch2_cloud, normals,
        significant, detection_limit, sigma1, sigma2,
        nz, nxy, dx, dy, dz, valid_mask

    Raises RuntimeError on any failure so the caller can fall back.
    """
    import subprocess, json, tempfile, os

    if progress_mgr:
        progress_mgr.update_current(5, "PDAL: building pipeline...")

    # ── Normal radii: PDAL expects a single float for its search radius ──
    # Use the median of the list as a sensible representative value
    nr_list  = [float(x) for x in normal_radii]
    nr_pdal  = float(np.median(nr_list))

    # ── Build pipeline JSON ─────────────────────────────────────────────
    stem       = os.path.splitext(output_file)[0]
    tmp_out    = stem + '_pdal_tmp.laz'

    pipeline = {
        "pipeline": [
            {
                "type": "readers.las",
                "filename": las_file1,
                "tag": "reference"
            },
            {
                "type": "readers.las",
                "filename": las_file2,
                "tag": "comparison"
            },
            {
                "type": "filters.m3c2",
                "inputs": ["reference", "comparison"],
                "normal_radius":   nr_pdal,    # Radius used for normal computation (direction_radii)
                "search_radius":   float(cyl_radius),
                "resolution":      float(core_spacing) if core_spacing > 0 else 0.5,
            },
            {
                "type": "writers.las",
                "filename": tmp_out,
                "extra_dims": "all"
            }
        ]
    }

    with tempfile.NamedTemporaryFile(mode='w', suffix='.json',
                                     delete=False) as fp:
        json.dump(pipeline, fp, indent=2)
        pipe_path = fp.name

    if progress_mgr:
        progress_mgr.update_current(10, "PDAL: running M3C2 (multi-core)...")

    try:
        result = subprocess.run(
            ['pdal', 'pipeline', pipe_path],
            capture_output=True, text=True, timeout=7200)  # 2 hour timeout
    finally:
        try: os.unlink(pipe_path)
        except Exception: pass

    if result.returncode != 0:
        raise RuntimeError(
            f"PDAL pipeline failed (exit {result.returncode}):\n"
            f"{result.stderr[:2000]}")

    if not os.path.exists(tmp_out):
        raise RuntimeError("PDAL ran successfully but output file was not created.")

    if progress_mgr:
        progress_mgr.update_current(70, "PDAL: reading output...")

    # ── Read PDAL output ────────────────────────────────────────────────
    las = laspy.read(tmp_out)
    try: os.unlink(tmp_out)
    except Exception: pass

    # Extract XYZ corepoints
    corepoints = np.column_stack([
        np.asarray(las.x, dtype=np.float64),
        np.asarray(las.y, dtype=np.float64),
        np.asarray(las.z, dtype=np.float64),
    ])

    # ── Find M3C2 distance field ────────────────────────────────────────
    # PDAL names the field 'M3C2Distance' or 'M3C2_distance' depending on version
    distances = None
    for candidate in ['M3C2Distance', 'M3C2_distance', 'm3c2distance',
                      'M3C2', 'distance']:
        try:
            arr = np.asarray(getattr(las, candidate), dtype=np.float64)
            if len(arr) == len(corepoints):
                distances = arr
                print(f"  PDAL: using field '{candidate}' as M3C2 distance")
                break
        except AttributeError:
            continue

    if distances is None:
        # Last resort: check all extra dimensions
        extra_names = [d.name for d in las.point_format.extra_dims]
        print(f"  PDAL output extra dims: {extra_names}")
        for name in extra_names:
            try:
                arr = np.asarray(las[name], dtype=np.float64)
                if len(arr) == len(corepoints) and not np.all(arr == 0):
                    distances = arr
                    print(f"  PDAL: using extra dim '{name}' as M3C2 distance")
                    break
            except Exception:
                continue

    if distances is None:
        raise RuntimeError(
            "Could not find M3C2 distance field in PDAL output.\n"
            f"Available extra dims: {[d.name for d in las.point_format.extra_dims]}")

    # ── Apply max depth filter ──────────────────────────────────────────
    if max_depth > 0:
        distances = np.where(np.abs(distances) > max_depth, np.nan, distances)
    valid_mask = ~np.isnan(distances)

    if progress_mgr:
        progress_mgr.update_current(80, "PDAL: estimating sigma from sample...")

    # ── Estimate sigma via py4dgeo on a small random sample ────────────
    # PDAL doesn't give per-point SPREAD so we estimate from a 50k sample
    sigma1 = sigma2 = None
    if use_manual_sigma and sigma1_manual and sigma2_manual:
        sigma1 = float(sigma1_manual)
        sigma2 = float(sigma2_manual)
        print(f"  PDAL sigma: manual σ1={sigma1:.4f} m, σ2={sigma2:.4f} m")
    else:
        try:
            n_sample = min(50_000, len(corepoints))
            idx      = np.random.choice(len(corepoints), n_sample, replace=False)
            samp_cp  = corepoints[idx]

            e1, e2 = py4dgeo.read_from_las(las_file1, las_file2)
            m3c2_s = py4dgeo.M3C2(
                epochs=(e1, e2), corepoints=samp_cp,
                cyl_radius=cyl_radius, normal_radii=normal_radii)
            _, unc_s = m3c2_s.run()
            sigma1 = float(np.nanmean(unc_s['spread1']))
            sigma2 = float(np.nanmean(unc_s['spread2']))
            print(f"  PDAL sigma (sample n={n_sample}): "
                  f"σ1={sigma1:.4f} m, σ2={sigma2:.4f} m")
        except Exception as e:
            print(f"  PDAL sigma estimation failed ({e}), using fallback 0.05 m")
            sigma1 = sigma2 = 0.05

    # ── Significance ────────────────────────────────────────────────────
    if confidence_level == 0.95: t_crit = 1.96
    elif confidence_level == 0.99: t_crit = 2.576
    else: t_crit = 1.645
    detection_limit = t_crit * np.sqrt(sigma1**2 + sigma2**2)
    significant = np.abs(distances) > detection_limit
    print(f"  PDAL LoD ({confidence_level*100:.0f}%): {detection_limit:.4f} m  "
          f"→ {int(significant.sum()):,} significant")

    # ── Normals + displacement components ──────────────────────────────
    if progress_mgr:
        progress_mgr.update_current(88, "PDAL: computing normals + components...")

    # Load ref cloud for normal computation
    e1_ref = py4dgeo.read_from_las(las_file1, las_file1)[0]
    ref_points  = e1_ref.cloud

    # Try to get epoch2 cloud — re-read las_file2 directly
    e2_ref = py4dgeo.read_from_las(las_file2, las_file2)[0]
    epoch2_cloud = e2_ref.cloud

    # Compute normals via py4dgeo on the PDAL corepoints
    if progress_mgr:
        progress_mgr.update_current(90, "PDAL: computing normals via py4dgeo...")
    m3c2_n = py4dgeo.M3C2(
        epochs=(e1_ref, py4dgeo.read_from_las(las_file2, las_file2)[0]),
        corepoints=corepoints,
        cyl_radius=cyl_radius,
        normal_radii=normal_radii)
    m3c2_n.run()
    normals = np.asarray(m3c2_n.directions(), dtype=np.float64)
    _mag = np.linalg.norm(normals, axis=1, keepdims=True)
    _mag[_mag == 0] = 1.0
    normals = normals / _mag

    nz  = np.abs(normals[:, 2]).astype(np.float32)
    nxy = np.sqrt(normals[:, 0]**2 + normals[:, 1]**2).astype(np.float32)
    dx  = (distances * normals[:, 0]).astype(np.float32)
    dy  = (distances * normals[:, 1]).astype(np.float32)
    dz  = (distances * normals[:, 2]).astype(np.float32)
    # PDAL doesn't expose per-point normal scale selection — use median of normal_radii
    normal_scale_radii = np.full(len(corepoints), float(np.median(normal_radii)),
                                  dtype=np.float32)

    # uncertainties=None signals to the rest of pipeline that SPREAD1/2 unavailable
    uncertainties = None

    if progress_mgr:
        progress_mgr.update_current(100, "PDAL: pipeline complete")

    return (corepoints, distances, uncertainties,
            ref_points, epoch2_cloud, normals,
            significant, detection_limit, sigma1, sigma2,
            nz, nxy, dx, dy, dz, valid_mask, normal_scale_radii)


def run_m3c2_gui(las_file1, las_file2, output_file, cyl_radius, normal_radii,
                 core_spacing, save_txt, export_components=True,
                 confidence_level=0.95, max_depth=0.0,
                 use_manual_sigma=False, sigma1_manual=0.05, sigma2_manual=0.05,
                 progress_mgr=None,
                 sigma_mode='per_point', registration_error=0.0,
                 make_raster=False, raster_resolution=1.0,
                 raster_fields=None, raster_interp_method="IDW",
                 make_dem=False, dem_resolution=0.1, dem_interp_method="IDW",
                 quiver_normal=False, quiver_motion=False,
                 quiver_grid_size=1.0, quiver_png=True, quiver_html=True, quiver_gpkg=True,
                 las_fields=None,
                 profile_lines=None,
                 profile_width=50.0,
                 profile_export_gpkg=False,
                 profile_export_txt=False,
                 # Wind rose
                 make_windrose=False, windrose_bins=16, windrose_sig_only=True,
                 # Raster clip mask (SHP or GPKG); None = no clipping
                 clip_mask_path=None,
                 epsg=None,
                 use_pdal=False,
                 make_report=True,
                 make_histogram=True,
                 make_volume=False,
                 volume_dem_resolution=0.5,
                 volume_dem_interp='IDW',
                 make_batch_csv=False,
                 batch_csv_dir=None):

    if not output_file:
        _ui_error("Chyba", "Zadajte výstupný LAS/LAZ súbor!")
        return

    import datetime as _dt
    _start_time = _dt.datetime.now()

    # ═══════════════════════════════════════════════════════════════════
    # ENGINE BRANCH
    # ═══════════════════════════════════════════════════════════════════
    if use_pdal:
        # ── PDAL path ─────────────────────────────────────────────────
        if progress_mgr:
            progress_mgr.update_current(0, "PDAL engine: starting...")
            progress_mgr.update_overall("PDAL M3C2 running...")
        try:
            (corepoints, distances, uncertainties,
             ref_points, epoch2_cloud, normals,
             significant, detection_limit, sigma1, sigma2,
             nz, nxy, dx, dy, dz, valid_mask,
             normal_scale_radii) = run_m3c2_pdal(
                las_file1, las_file2, output_file,
                cyl_radius, normal_radii,
                core_spacing if core_spacing else 0.5,
                max_depth, confidence_level,
                use_manual_sigma, sigma1_manual, sigma2_manual,
                sigma_mode=sigma_mode, registration_error=registration_error,
                progress_mgr=progress_mgr)
            if progress_mgr:
                progress_mgr.update_overall("PDAL M3C2 complete")
        except Exception as pdal_err:
            _ui_error("PDAL Error",
                      f"PDAL pipeline failed:\n{pdal_err}\n\n"
                      "Check that PDAL is installed:\n"
                      "  conda install -c conda-forge pdal\n\n"
                      "Falling back to py4dgeo is not automatic — "
                      "please switch the engine in the GUI.")
            return

        # epoch2.cloud equivalent for profile export
        class _Ep2Stub:
            def __init__(self, cloud): self.cloud = cloud
        epoch2 = _Ep2Stub(epoch2_cloud)

    else:
        # ── py4dgeo path ───────────────────────────────────────────────

        # Step 1
        if progress_mgr: progress_mgr.update_current(0, "Loading point clouds...")
        epoch1, epoch2 = py4dgeo.read_from_las(las_file1, las_file2)
        ref_points = epoch1.cloud
        if progress_mgr:
            progress_mgr.update_current(100, "Point clouds loaded")
            progress_mgr.update_overall("Point clouds loaded")

        # Step 2
        if core_spacing and core_spacing > 0:
            corepoints = subsample_by_distance(ref_points, core_spacing, progress_mgr)
        else:
            corepoints = ref_points
            if progress_mgr: progress_mgr.update_current(100, "Using all points as corepoints")
        if progress_mgr: progress_mgr.update_overall("Subsampling complete")

        # Step 3
        if progress_mgr: progress_mgr.update_current(0, "Running M3C2 algorithm...")
        m3c2 = py4dgeo.M3C2(epochs=(epoch1, epoch2), corepoints=corepoints,
                             cyl_radius=cyl_radius, normal_radii=normal_radii,
                             max_distance=max_depth if max_depth > 0 else 1e9)
        if progress_mgr: progress_mgr.update_current(50, "M3C2: Computing distances...")
        distances, uncertainties = m3c2.run()
        if progress_mgr:
            progress_mgr.update_current(100, "M3C2 complete")
            progress_mgr.update_overall("M3C2 distances computed")

        # Step 4 — use py4dgeo internal normals directly
        if progress_mgr: progress_mgr.update_current(0, "Extracting normals from py4dgeo...")
        normals = np.asarray(m3c2.directions(), dtype=np.float64)
        # Radius used for normal computation per corepoint (selected scale)
        normal_scale_radii = np.asarray(m3c2.directions_radii(), dtype=np.float32)
        # Ensure unit length
        _mag = np.linalg.norm(normals, axis=1, keepdims=True)
        _mag[_mag == 0] = 1.0
        normals = normals / _mag
        if progress_mgr: progress_mgr.update_overall("Normals extracted")

        # Step 5 — max_distance handled by py4dgeo internally (sets NaN beyond limit)
        valid_mask = ~np.isnan(distances)
        if progress_mgr: progress_mgr.update_overall("Depth filtering complete")

        # Step 6
        significant, detection_limit, sigma1, sigma2 = calculate_significance_improved(
            distances, uncertainties, confidence_level,
            sigma1_manual=sigma1_manual, sigma2_manual=sigma2_manual,
            sigma_mode=sigma_mode,
            registration_error=registration_error,
            progress_mgr=progress_mgr)
        if progress_mgr: progress_mgr.update_overall("Significance computed")

        # Step 7 — Nz, Nxy, DX, DY, DZ from py4dgeo normals
        nz  = np.abs(normals[:, 2]).astype(np.float32)
        nxy = np.sqrt(normals[:, 0]**2 + normals[:, 1]**2).astype(np.float32)
        dx  = (distances * normals[:, 0]).astype(np.float32)
        dy  = (distances * normals[:, 1]).astype(np.float32)
        dz  = (distances * normals[:, 2]).astype(np.float32)
        if progress_mgr: progress_mgr.update_overall("Normal components computed")

    # Step 8
    success = save_results_simplified(corepoints, distances, nz, nxy, significant,
                                      output_file, save_txt, progress_mgr,
                                      uncertainties=uncertainties,
                                      dx=dx, dy=dy, dz=dz,
                                      normal_scale=normal_scale_radii,
                                      las_fields=las_fields)
    if progress_mgr: progress_mgr.update_overall("Results saved")
    if not success:
        _ui_error("Chyba", "Nastala chyba pri ukladaní výsledkov.")
        return

    # Step 9
    if progress_mgr: progress_mgr.update_current(0, "Computing statistics...")
    print_comprehensive_stats(distances, significant, valid_mask, detection_limit,
                               max_depth, nz, nxy)
    if progress_mgr:
        progress_mgr.update_current(100, "Statistics computed")
        progress_mgr.update_overall("Exports running...")

    field_map = {
        "M3C2_DISTANCE": distances.astype(np.float32),
        "SIGNIFICANT":   significant.astype(np.float32),
        "Nz": nz.astype(np.float32), "Nxy": nxy.astype(np.float32),
        "DX": dx.astype(np.float32), "DY": dy.astype(np.float32), "DZ": dz.astype(np.float32),
        "NORMAL_SCALE":  normal_scale_radii.astype(np.float32),
    }
    if uncertainties is not None:
        for py4d_key, col_name in [("lodetection","LOD"),("spread1","SPREAD1"),("spread2","SPREAD2")]:
            try:
                arr = np.asarray(uncertainties[py4d_key], dtype=np.float32).ravel()
                if len(arr) == len(corepoints): field_map[col_name] = arr
            except Exception: pass

    saved_extras = []

    if make_raster:
        fields_to_export = [(fld, field_map[fld])
                            for fld in (raster_fields or ["M3C2_DISTANCE"])
                            if fld in field_map]

        def _export_one_raster(fld, values):
            rp = output_file.replace(".las", f"_{fld}.tif").replace(".laz", f"_{fld}.tif")
            ok = export_raster(corepoints, values, rp, raster_resolution, fld,
                               interp_method=raster_interp_method,
                               epsg=epsg, progress_mgr=None)  # no progress in threads
            if ok and clip_mask_path:
                clip_raster_by_mask(rp, clip_mask_path, progress_mgr=None)
            return rp if ok else None

        if len(fields_to_export) > 1:
            if progress_mgr:
                progress_mgr.update_current(0, f"Raster export: {len(fields_to_export)} fields in parallel...")
            with ThreadPoolExecutor(max_workers=min(_N_THREADS, len(fields_to_export))) as ex:
                futures = {ex.submit(_export_one_raster, fld, vals): fld
                           for fld, vals in fields_to_export}
                for fut in as_completed(futures):
                    res = fut.result()
                    if res:
                        saved_extras.append(res)
        else:
            for fld, vals in fields_to_export:
                rp = _export_one_raster(fld, vals)
                if rp: saved_extras.append(rp)

    if make_dem:
        dp = output_file.replace(".las","_DEM.tif").replace(".laz","_DEM.tif")
        if export_raster(corepoints, corepoints[:, 2], dp, dem_resolution, "Z (DEM)",
                         interp_method=dem_interp_method,
                         epsg=epsg, progress_mgr=progress_mgr):
            if clip_mask_path:
                clip_raster_by_mask(dp, clip_mask_path, progress_mgr)
            saved_extras.append(dp)

    quiver_kwargs = dict(grid_size=quiver_grid_size, export_png=quiver_png,
                         export_html=quiver_html, export_gpkg=quiver_gpkg,
                         progress_mgr=progress_mgr)
    if quiver_normal:
        result = export_quiver(corepoints, distances, normals, significant, output_file,
                               mode='normal', **quiver_kwargs)
        if result: saved_extras.extend(result)
    if quiver_motion:
        result = export_quiver(corepoints, distances, normals, significant, output_file,
                               mode='motion', dx=dx, dy=dy, **quiver_kwargs)
        if result: saved_extras.extend(result)

    # ── Profile exports (one PNG per line, parallel) ─────────────────
    if profile_lines:
        stem = os.path.splitext(output_file)[0]

        def _export_one_profile(args):
            i, lbl, pxy = args
            pp = f"{stem}_profile_{i+1:02d}.png"
            ok = export_profile_png(corepoints, distances, significant, pxy,
                                    profile_width, pp, detection_limit,
                                    profile_label=lbl, progress_mgr=None,
                                    ref_cloud=ref_points,
                                    comp_cloud=epoch2.cloud)
            return pp if ok else None

        if len(profile_lines) > 1:
            if progress_mgr:
                progress_mgr.update_current(0,
                    f"Profile export: {len(profile_lines)} lines in parallel...")
            with ThreadPoolExecutor(
                    max_workers=min(_N_THREADS, len(profile_lines))) as ex:
                futs = [ex.submit(_export_one_profile, (i, lbl, pxy))
                        for i, (lbl, pxy) in enumerate(profile_lines)]
                for fut in as_completed(futs):
                    res = fut.result()
                    if res: saved_extras.append(res)
        else:
            for i, (lbl, pxy) in enumerate(profile_lines):
                pp = f"{stem}_profile_{i+1:02d}.png"
                if export_profile_png(corepoints, distances, significant, pxy,
                                      profile_width, pp, detection_limit,
                                      profile_label=lbl, progress_mgr=progress_mgr,
                                      ref_cloud=ref_points, comp_cloud=epoch2.cloud):
                    saved_extras.append(pp)

    # ── Profile lines export (GPKG / TXT) ────────────────────────────
    if profile_lines and (profile_export_gpkg or profile_export_txt):
        stem   = os.path.splitext(output_file)[0]
        pl_res = export_profile_lines(
            profile_lines, stem,
            export_gpkg=profile_export_gpkg,
            export_txt=profile_export_txt)
        saved_extras.extend(pl_res)

    # ── Wind rose export ───────────────────────────────────────────────
    if make_windrose:
        for _wlabel, _wsig, _wsign, _wsuffix in [
            ("all points",        False, None,  '_windrose'),
            ("significant only",  True,  None,  '_windrose_sig'),
            ("positive / gain",   True,  'pos', '_windrose_pos'),
            ("negative / loss",   True,  'neg', '_windrose_neg'),
        ]:
            if progress_mgr:
                progress_mgr.update_current(0, f"Wind rose ({_wlabel})...")
            wp = export_windrose(corepoints, dx, dy, dz, significant, output_file,
                                 n_bins=windrose_bins, sig_only=_wsig,
                                 sign_filter=_wsign, suffix=_wsuffix,
                                 progress_mgr=progress_mgr)
            if wp:
                saved_extras.append(wp)

    # ── Histogram export ───────────────────────────────────────────────
    if make_histogram:
        if progress_mgr:
            progress_mgr.update_current(0, "Histogram export...")
        hp = export_histogram_png(distances, significant, detection_limit,
                                  output_file, progress_mgr=progress_mgr)
        if hp:
            saved_extras.append(hp)

    # ── Volume change estimation ───────────────────────────────────────
    _volume_result = None
    if make_volume:
        if progress_mgr:
            progress_mgr.update_current(0, "Volume change estimation...")
        _volume_result = compute_volume_change(
            corepoints, distances, significant, nz,
            cell_size=raster_resolution if raster_resolution > 0 else 1.0,
            las_file1=las_file1, las_file2=las_file2,
            dem_resolution=volume_dem_resolution,
            dem_interp=volume_dem_interp,
            detection_limit=detection_limit,
            progress_mgr=progress_mgr)

    # ── PDF Report ────────────────────────────────────────────────────
    if make_report:
        import datetime as _dt
        _end_time = _dt.datetime.now()
        report_path = os.path.splitext(output_file)[0] + '_report.pdf'
        if progress_mgr:
            progress_mgr.update_current(0, "Generating PDF report...")
        _report_ok = generate_report_pdf(
            report_path,
            report_data={
                'las_file1':         las_file1,
                'las_file2':         las_file2,
                'output_file':       output_file,
                'cyl_radius':        cyl_radius,
                'normal_radii':      normal_radii,
                'core_spacing':      core_spacing,
                'max_depth':         max_depth,
                'confidence_level':  confidence_level,
                'sigma_mode':          sigma_mode,
                'registration_error':  registration_error,
                'sigma1':            sigma1,
                'sigma2':            sigma2,
                'make_raster':       make_raster,
                'raster_resolution': raster_resolution,
                'raster_interp_method': raster_interp_method,
                'make_dem':          make_dem,
                'dem_resolution':    dem_resolution,
                'quiver_normal':     quiver_normal,
                'quiver_motion':     quiver_motion,
                'quiver_grid_size':  quiver_grid_size,
                'clip_mask_path':    clip_mask_path,
                'corepoints':        corepoints,
                'distances':         distances,
                'significant':       significant,
                'valid_mask':        valid_mask,
                'nz': nz, 'nxy': nxy,
                'dx': dx, 'dy': dy, 'dz': dz,
                'detection_limit':   detection_limit,
                'profile_lines':     profile_lines,
                'profile_width':     profile_width,
                'saved_extras':      saved_extras,
                'start_time':        _start_time,
                'end_time':          _end_time,
                'volume_result':     _volume_result,
                'make_volume':       make_volume,
                'volume_dem_interp': volume_dem_interp,
                'engine':            'PDAL' if use_pdal else 'py4dgeo',
            },
            progress_mgr=progress_mgr)
        if _report_ok:
            saved_extras.append(report_path)

    # ── Batch CSV export ──────────────────────────────────────────────
    if make_batch_csv:
        _append_batch_csv_row(output_file, {
            'distances': distances, 'significant': significant,
            'valid_mask': valid_mask, 'nz': nz,
            'detection_limit': detection_limit,
            'sigma1': sigma1, 'sigma2': sigma2,
            'corepoints': corepoints,
            'cyl_radius': cyl_radius, 'normal_radii': normal_radii,
            'core_spacing': core_spacing,
            'registration_error': registration_error,
            'sigma_mode': sigma_mode,
        }, csv_dir_override=batch_csv_dir)
        _csv_dir = batch_csv_dir if batch_csv_dir else os.path.dirname(output_file)
        saved_extras.append(os.path.join(_csv_dir, 'batch_stats.csv'))
        _xlsx_path = os.path.join(_csv_dir, EXCEL_OUTPUT_NAME)
        if os.path.exists(_xlsx_path):
            saved_extras.append(_xlsx_path)

    extras_str = "\n".join(saved_extras) if saved_extras else "—"
    if progress_mgr:
        progress_mgr.update_overall("Complete!")
        progress_mgr.update_current(100, "All done.")
    _ui_info("Hotovo",
        "M3C2 dokončené!\n"
        "Polia: X, Y, Z, M3C2_DISTANCE, SIGNIFICANT, Nz, Nxy, LOD, SPREAD1, SPREAD2\n\n"
        f"Dodatočné súbory:\n{extras_str}")


# ═══════════════════════════════════════════════════════════
# GUI
# ═══════════════════════════════════════════════════════════
def browse_file(entry, save=False):
    fp = (filedialog.asksaveasfilename(defaultextension=".las",
                                       filetypes=[("LAS/LAZ files", "*.las *.laz")])
          if save else
          filedialog.askopenfilename(filetypes=[("LAS/LAZ files", "*.las *.laz")]))
    if fp:
        entry.delete(0, tk.END); entry.insert(0, fp)

def toggle_raster_entries():
    s = "normal" if var_make_raster.get() else "disabled"
    entry_raster_res.config(state=s); option_raster_interp.config(state=s)
    for chk in raster_field_chks.values(): chk.config(state=s)

def toggle_dem_entries():
    s = "normal" if var_make_dem.get() else "disabled"
    entry_dem_res.config(state=s); option_dem_interp.config(state=s)

def toggle_quiver_entries():
    either = var_quiver_normal.get() or var_quiver_motion.get()
    s = "normal" if either else "disabled"
    entry_quiver_grid.config(state=s)
    for chk in _qfmt_chks: chk.config(state=s)

def toggle_profile_entries():
    s = "normal" if var_make_profile.get() else "disabled"
    entry_profile_width.config(state=s)
    entry_profile_txt.config(state=s)
    btn_profile_browse.config(state=s)
    btn_pick_profiles.config(state=s)

def save_settings(path=None):
    """Save all current GUI values to a JSON session file."""
    import json
    if path is None:
        path = filedialog.asksaveasfilename(
            title="Save session",
            defaultextension=".json",
            filetypes=[("M3C2 session", "*.json"), ("All files", "*.*")])
    if not path:
        return
    data = {
        # Files
        'las1':   entry_las1.get(),
        'las2':   entry_las2.get(),
        'output': entry_output.get(),
        # M3C2 params
        'cyl_radius':       entry_cyl.get(),
        'normal_radii':     entry_normals.get(),
        'core_spacing':     entry_spacing.get(),
        'use_all_points':   var_use_all_points.get(),
        'max_depth':        entry_max_depth.get(),
        'confidence':       var_confidence.get(),
        'sigma_mode':   var_sigma_mode.get(),
        'reg_error':    entry_reg_error.get(),
        'sigma1':       entry_sigma1.get(),
        'sigma2':       entry_sigma2.get(),
        'clip_raster':      var_clip_raster.get(),
        'clip_mask':        entry_clip_mask.get(),
        'epsg':             entry_epsg.get(),
        'multithreading':   var_multithreading.get(),
        'engine':           var_engine.get(),
        # Output options
        'save_txt':         var_txt.get(),
        'make_report':      var_make_report.get(),
        'make_histogram':   var_make_histogram.get(),
        'export_batch_csv': var_export_batch_csv.get(),
        'las_fields':       {k: v.get() for k, v in las_field_vars.items()},
        # Raster
        'make_raster':      var_make_raster.get(),
        'raster_res':       entry_raster_res.get(),
        'raster_interp':    var_raster_interp.get(),
        'raster_fields':    {k: v.get() for k, v in raster_field_vars.items()},
        # DEM
        'make_dem':         var_make_dem.get(),
        'dem_res':          entry_dem_res.get(),
        'dem_interp':       var_dem_interp.get(),
        # Volume
        'make_volume':      var_make_volume.get(),
        'vol_dem_res':      entry_vol_dem_res.get(),
        'vol_dem_interp':   var_vol_dem_interp.get(),
        # Quiver
        'quiver_normal':    var_quiver_normal.get(),
        'quiver_motion':    var_quiver_motion.get(),
        'quiver_grid':      entry_quiver_grid.get(),
        'quiver_png':       var_quiver_png.get(),
        'quiver_html':      var_quiver_html.get(),
        'quiver_gpkg':      var_quiver_gpkg.get(),
        # Wind rose
        'make_windrose':    var_make_windrose.get(),
        'windrose_bins':    var_windrose_bins.get(),
        # Profile
        'profile_width':    entry_profile_width.get(),
        'profile_txt':      entry_profile_txt.get(),
        'profile_gpkg':     var_profile_export_gpkg.get(),
        'profile_line_txt': var_profile_export_txt.get(),
    }
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
        messagebox.showinfo("Session saved", f"Settings saved to:\n{path}")
    except Exception as e:
        messagebox.showerror("Chyba", f"Could not save session:\n{e}")


def load_settings(path=None):
    """Load GUI values from a previously saved JSON session file."""
    import json
    if path is None:
        path = filedialog.askopenfilename(
            title="Load session",
            filetypes=[("M3C2 session", "*.json"), ("All files", "*.*")])
    if not path or not os.path.exists(path):
        return
    try:
        with open(path, 'r', encoding='utf-8') as f:
            d = json.load(f)
    except Exception as e:
        messagebox.showerror("Chyba", f"Could not read session file:\n{e}")
        return

    def _set(entry, key):
        if key in d:
            entry.delete(0, tk.END)
            entry.insert(0, str(d[key]))

    def _setvar(var, key):
        if key in d:
            try: var.set(d[key])
            except Exception: pass

    # Files
    _set(entry_las1,   'las1')
    _set(entry_las2,   'las2')
    _set(entry_output, 'output')
    # M3C2 params
    _set(entry_cyl,     'cyl_radius')
    _set(entry_normals, 'normal_radii')
    _set(entry_spacing, 'core_spacing')
    _setvar(var_use_all_points,   'use_all_points');  toggle_spacing_entry()
    _set(entry_max_depth, 'max_depth')
    _setvar(var_confidence,   'confidence')
    _setvar(var_sigma_mode,   'sigma_mode');  toggle_sigma_entries()
    _set(entry_reg_error,  'reg_error')
    _set(entry_sigma1,     'sigma1')
    _set(entry_sigma2,     'sigma2')
    _setvar(var_clip_raster, 'clip_raster');           toggle_clip_entries()
    _set(entry_clip_mask, 'clip_mask')
    _set(entry_epsg,    'epsg')
    if 'multithreading' in d:
        var_multithreading.set(bool(d['multithreading']))
    if 'engine' in d and d['engine'] in ('py4dgeo', 'pdal'):
        if d['engine'] == 'pdal' and _pdal_ok:
            var_engine.set('pdal')
        else:
            var_engine.set('py4dgeo')
    # Output options
    _setvar(var_txt,              'save_txt')
    _setvar(var_make_report,      'make_report')
    _setvar(var_make_histogram,   'make_histogram')
    _setvar(var_export_batch_csv, 'export_batch_csv')
    if 'las_fields' in d:
        for k, v in d['las_fields'].items():
            if k in las_field_vars:
                las_field_vars[k].set(v)
    # Raster
    _setvar(var_make_raster,  'make_raster');  toggle_raster_entries()
    _set(entry_raster_res,    'raster_res')
    _setvar(var_raster_interp,'raster_interp')
    if 'raster_fields' in d:
        for k, v in d['raster_fields'].items():
            if k in raster_field_vars:
                raster_field_vars[k].set(v)
    # DEM
    _setvar(var_make_dem,   'make_dem');  toggle_dem_entries()
    _set(entry_dem_res,     'dem_res')
    _setvar(var_dem_interp, 'dem_interp')
    # Volume
    _setvar(var_make_volume,    'make_volume');  toggle_volume_entries()
    _set(entry_vol_dem_res,     'vol_dem_res')
    _setvar(var_vol_dem_interp, 'vol_dem_interp')
    # Quiver
    _setvar(var_quiver_normal, 'quiver_normal')
    _setvar(var_quiver_motion, 'quiver_motion'); toggle_quiver_entries()
    _set(entry_quiver_grid,    'quiver_grid')
    _setvar(var_quiver_png,    'quiver_png')
    _setvar(var_quiver_html,   'quiver_html')
    _setvar(var_quiver_gpkg,   'quiver_gpkg')
    # Wind rose
    _setvar(var_make_windrose, 'make_windrose'); toggle_windrose_entries()
    _setvar(var_windrose_bins, 'windrose_bins')
    # Profile
    _set(entry_profile_width,   'profile_width')
    _set(entry_profile_txt,     'profile_txt')
    _setvar(var_profile_export_gpkg, 'profile_gpkg')
    _setvar(var_profile_export_txt,  'profile_line_txt')
    toggle_profile_entries()

    messagebox.showinfo("Session loaded",
                        f"Settings loaded from:\n{os.path.basename(path)}")


def start_batch_threaded():
    """
    Batch-process multiple JSON session files.

    For each selected JSON:
      1. Load GUI settings from JSON (via load_settings)
      2. Create output folder named after the JSON filename (without .json)
          next to the original output path specified in the JSON
      3. Redirect ALL outputs into that folder
      4. Run full M3C2 pipeline
      5. Move to the next JSON

    All jobs run sequentially in one worker thread.
    """
    json_paths = filedialog.askopenfilenames(
        title="Select session JSON files for batch processing",
        filetypes=[("M3C2 session JSON", "*.json"), ("All files", "*.*")])
    if not json_paths:
        return
    json_paths = list(json_paths)
    n_jobs = len(json_paths)

    # Confirm with user
    summary = "\n".join(f"  \u2022  {os.path.basename(p)}" for p in json_paths[:10])
    if n_jobs > 10:
        summary += f"\n  \u2026 and {n_jobs - 10} more"
    if not messagebox.askyesno(
            "Confirm batch processing",
            f"Found {n_jobs} session file(s):\n\n{summary}\n\n"
            f"Each job will create its own output folder next to the JSON file.\n"
            f"Existing files in those folders will be overwritten.\n\n"
            f"Start batch processing?"):
        return

    btn_start.config(state="disabled")
    btn_batch.config(state="disabled")
    progress_mgr.reset()

    def _batch_worker():
        successes = []
        failures  = []

        # All batch jobs share the parent folder of the JSON files — this is
        # where we aggregate batch_stats.csv across all runs (one row each).
        shared_csv_dir = os.path.dirname(json_paths[0])

        for i, json_path in enumerate(json_paths, start=1):
            job_name = os.path.splitext(os.path.basename(json_path))[0]
            print("\n" + "=" * 70)
            print(f"BATCH JOB {i}/{n_jobs}:  {job_name}")
            print("=" * 70)

            # Update progress label on main thread
            root.after(0, lambda n=job_name, i=i:
                       progress_mgr.update_overall(f"Batch {i}/{n_jobs}: {n}"))

            try:
                # Load session values into GUI (runs on main thread to be safe)
                loaded = {'ok': False, 'err': None}
                def _load():
                    try:
                        load_settings(json_path)
                        loaded['ok'] = True
                    except Exception as e:
                        loaded['err'] = str(e)
                root.after(0, _load)
                # Wait for GUI load to finish
                import time
                for _ in range(50):
                    if loaded['ok'] or loaded['err']:
                        break
                    time.sleep(0.1)
                if loaded['err']:
                    raise RuntimeError(f"Failed to load session: {loaded['err']}")

                time.sleep(0.3)  # ensure GUI updated

                # ── Read all GUI values from this session ─────────────────
                values = {'error': None}
                def _collect():
                    try:
                        values['cyl']          = float(entry_cyl.get())
                        values['normals_r']    = [float(x) for x in entry_normals.get().split(",")]
                        values['core_spacing'] = (0 if var_use_all_points.get()
                                                  else (float(entry_spacing.get()) if entry_spacing.get() else 0))
                        values['max_depth']    = float(entry_max_depth.get()) if entry_max_depth.get() else 0.0
                        values['conf_level']   = float(var_confidence.get())
                        values['sigma_mode']   = var_sigma_mode.get()
                        values['reg_error']    = float(entry_reg_error.get())  if entry_reg_error.get()  else 0.0
                        values['s1']           = float(entry_sigma1.get())      if entry_sigma1.get()    else 0.05
                        values['s2']           = float(entry_sigma2.get())      if entry_sigma2.get()    else 0.05
                        values['use_man_s']    = (values['sigma_mode'] == 'manual_sigma')
                        values['raster_res']   = float(entry_raster_res.get()) if entry_raster_res.get() else 1.0
                        values['dem_res']      = float(entry_dem_res.get())    if entry_dem_res.get()    else 0.1
                        values['quiver_grid']  = float(entry_quiver_grid.get()) if entry_quiver_grid.get() else 1.0
                        values['profile_w']    = float(entry_profile_width.get()) if entry_profile_width.get() else 50.0
                        values['windrose_bins']= int(var_windrose_bins.get())
                        values['vol_dem_res']  = float(entry_vol_dem_res.get()) if entry_vol_dem_res.get() else 0.5
                        values['las1']         = entry_las1.get()
                        values['las2']         = entry_las2.get()
                        values['output_orig']  = entry_output.get()
                        values['save_txt']     = var_txt.get()
                        values['make_report']  = var_make_report.get()
                        values['make_histogram']= var_make_histogram.get()
                        values['export_batch_csv'] = var_export_batch_csv.get()
                        values['make_volume']  = var_make_volume.get()
                        values['vol_dem_interp']= var_vol_dem_interp.get()
                        values['make_raster']  = var_make_raster.get()
                        values['raster_fields_sel']= [f for f, v in raster_field_vars.items() if v.get()]
                        values['raster_interp']= var_raster_interp.get()
                        values['make_dem']     = var_make_dem.get()
                        values['dem_interp']   = var_dem_interp.get()
                        values['quiver_normal']= var_quiver_normal.get()
                        values['quiver_motion']= var_quiver_motion.get()
                        values['quiver_png']   = var_quiver_png.get()
                        values['quiver_html']  = var_quiver_html.get()
                        values['quiver_gpkg']  = var_quiver_gpkg.get()
                        values['selected_las_fields']= {f for f, v in las_field_vars.items() if v.get()}
                        values['clip_mask']    = entry_clip_mask.get().strip() if var_clip_raster.get() else None
                        values['epsg']         = entry_epsg.get().strip() or None
                        values['use_pdal']     = (var_engine.get() == 'pdal')
                        values['make_windrose']= var_make_windrose.get()
                        values['windrose_sig_only']= var_windrose_sig_only.get()
                        values['profile_gpkg'] = var_profile_export_gpkg.get()
                        values['profile_txt']  = var_profile_export_txt.get()
                        # Profile lines from TXT if any
                        prof_lines = []
                        ptxt = entry_profile_txt.get().strip()
                        if var_make_profile.get() and ptxt:
                            for lbl, pxy in load_profile_lines_from_txt(ptxt):
                                prof_lines.append((lbl, pxy))
                        values['prof_lines'] = prof_lines if prof_lines else None
                    except Exception as e:
                        values['error'] = str(e)
                root.after(0, _collect)
                for _ in range(50):
                    if 'las1' in values or values['error']:
                        break
                    time.sleep(0.1)
                if values['error']:
                    raise RuntimeError(f"Failed to read settings: {values['error']}")

                # ── Create output folder next to JSON (named after JSON) ───
                batch_folder = os.path.join(os.path.dirname(json_path), job_name)
                os.makedirs(batch_folder, exist_ok=True)

                # Derive output filename from original output path, put it in new folder
                orig_out = values['output_orig']
                orig_name = os.path.basename(orig_out) if orig_out else 'vystup.laz'
                if not orig_name or not orig_name.endswith(('.las', '.laz', '.LAS', '.LAZ')):
                    orig_name = 'vystup.laz'
                new_output = os.path.join(batch_folder, orig_name)
                print(f"  Output folder: {batch_folder}")
                print(f"  Output file  : {new_output}")

                # ── Run M3C2 pipeline ─────────────────────────────────────
                run_m3c2_gui(
                    values['las1'], values['las2'], new_output,
                    values['cyl'], values['normals_r'], values['core_spacing'],
                    values['save_txt'], True, values['conf_level'], values['max_depth'],
                    values['use_man_s'], values['s1'], values['s2'], progress_mgr,
                    sigma_mode=values['sigma_mode'],
                    registration_error=values['reg_error'],
                    make_raster=values['make_raster'],
                    raster_resolution=values['raster_res'],
                    raster_fields=values['raster_fields_sel'],
                    raster_interp_method=values['raster_interp'],
                    make_dem=values['make_dem'],
                    dem_resolution=values['dem_res'],
                    dem_interp_method=values['dem_interp'],
                    quiver_normal=values['quiver_normal'],
                    quiver_motion=values['quiver_motion'],
                    quiver_grid_size=values['quiver_grid'],
                    quiver_png=values['quiver_png'],
                    quiver_html=values['quiver_html'],
                    quiver_gpkg=values['quiver_gpkg'],
                    las_fields=values['selected_las_fields']
                               if values['selected_las_fields'] else None,
                    profile_lines=values['prof_lines'],
                    profile_width=values['profile_w'],
                    profile_export_gpkg=values['profile_gpkg'],
                    profile_export_txt=values['profile_txt'],
                    make_windrose=values['make_windrose'],
                    windrose_bins=values['windrose_bins'],
                    windrose_sig_only=values['windrose_sig_only'],
                    clip_mask_path=values['clip_mask'],
                    epsg=values['epsg'],
                    use_pdal=values['use_pdal'],
                    make_report=values['make_report'],
                    make_histogram=values['make_histogram'],
                    make_volume=values['make_volume'],
                    volume_dem_resolution=values['vol_dem_res'],
                    volume_dem_interp=values['vol_dem_interp'],
                    make_batch_csv=values.get('export_batch_csv', False),
                    batch_csv_dir=shared_csv_dir,
                )

                # Copy the JSON itself into the output folder for traceability
                try:
                    import shutil
                    shutil.copy2(json_path, os.path.join(batch_folder,
                                                          os.path.basename(json_path)))
                except Exception:
                    pass

                successes.append(job_name)
                print(f"  \u2713 BATCH JOB {i}/{n_jobs} COMPLETED: {job_name}")

            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                print(f"  \u2717 BATCH JOB {i}/{n_jobs} FAILED: {job_name}")
                print(tb)
                failures.append((job_name, str(e)))
                # Continue with next job
                continue

        # ── Summary ───────────────────────────────────────────────────────
        print("\n" + "=" * 70)
        print(f"BATCH COMPLETE: {len(successes)}/{n_jobs} succeeded, "
              f"{len(failures)} failed")
        print("=" * 70)

        summary_msg = f"Batch processing finished.\n\n"
        summary_msg += f"Succeeded: {len(successes)} / {n_jobs}\n"
        if failures:
            summary_msg += f"Failed:    {len(failures)}\n\n"
            summary_msg += "Failures:\n" + "\n".join(
                f"  \u2022 {name}: {err[:80]}" for name, err in failures[:10])

        root.after(0, lambda: messagebox.showinfo("Batch processing finished",
                                                    summary_msg))
        root.after(0, lambda: btn_start.config(state="normal"))
        root.after(0, lambda: btn_batch.config(state="normal"))

    threading.Thread(target=_batch_worker, daemon=True).start()


def start_m3c2_threaded():
    """
    Called from the main thread (button click).
    ALL tkinter .get() calls happen here — never inside the worker thread.
    The worker receives only plain Python values.
    """
    # ── Read every GUI value here, in the main thread ─────────────────
    try:
        cyl          = float(entry_cyl.get())
        normals_r    = [float(x) for x in entry_normals.get().split(",")]
        core_spacing = (0 if var_use_all_points.get()
                        else (float(entry_spacing.get()) if entry_spacing.get() else 0))
        max_depth    = float(entry_max_depth.get()) if entry_max_depth.get() else 0.0
        conf_level   = float(var_confidence.get())
        sigma_mode   = var_sigma_mode.get()
        reg_error    = float(entry_reg_error.get()) if entry_reg_error.get() else 0.0
        s1           = float(entry_sigma1.get())    if entry_sigma1.get()    else 0.05
        s2           = float(entry_sigma2.get())    if entry_sigma2.get()    else 0.05
        use_man_s    = (sigma_mode == 'manual_sigma')
        raster_res   = float(entry_raster_res.get()) if entry_raster_res.get() else 1.0
        dem_res      = float(entry_dem_res.get())    if entry_dem_res.get()    else 0.1
        quiver_grid  = float(entry_quiver_grid.get()) if entry_quiver_grid.get() else 1.0
        profile_w    = float(entry_profile_width.get()) if entry_profile_width.get() else 50.0
        windrose_bins_val = int(var_windrose_bins.get())
        vol_dem_res = float(entry_vol_dem_res.get()) if entry_vol_dem_res.get() else 0.5
    except ValueError:
        messagebox.showerror("Chyba", "Skontrolujte numerické polia.")
        return

    las1_path   = entry_las1.get()
    las2_path   = entry_las2.get()
    output_path = entry_output.get()
    save_txt       = var_txt.get()
    make_report    = var_make_report.get()
    make_histogram = var_make_histogram.get()
    make_volume    = var_make_volume.get()
    vol_dem_interp = var_vol_dem_interp.get()

    make_raster        = var_make_raster.get()
    raster_fields_sel  = [f for f, v in raster_field_vars.items() if v.get()]
    raster_interp      = var_raster_interp.get()
    make_dem           = var_make_dem.get()
    dem_interp         = var_dem_interp.get()
    quiver_normal      = var_quiver_normal.get()
    quiver_motion      = var_quiver_motion.get()
    quiver_png         = var_quiver_png.get()
    quiver_html        = var_quiver_html.get()
    quiver_gpkg        = var_quiver_gpkg.get()
    selected_las_fields = {f for f, v in las_field_vars.items() if v.get()}

    clip_mask_path     = entry_clip_mask.get().strip() if var_clip_raster.get() else None
    epsg_val           = entry_epsg.get().strip() or None
    use_pdal           = (var_engine.get() == 'pdal')
    make_windrose      = var_make_windrose.get()
    windrose_sig_only  = var_windrose_sig_only.get()

    prof_lines = []
    if var_make_profile.get():
        if picked_profile_lines:
            prof_lines = list(picked_profile_lines)
        txt_path = entry_profile_txt.get().strip()
        if txt_path:
            try:
                for lbl, pxy in load_profile_lines_from_txt(txt_path):
                    prof_lines.append((lbl, pxy))
            except Exception as e:
                messagebox.showerror("Chyba", f"Profile TXT chyba: {e}")
                return

    # ── Apply thread count ────────────────────────────────────────────
    global _N_THREADS
    _N_THREADS = max(1, (os.cpu_count() or 2) - 1) if var_multithreading.get() else 1
    os.environ['OMP_NUM_THREADS'] = str(_N_THREADS)

    # ── Disable button, launch worker ────────────────────────────────
    btn_start.config(state="disabled")
    progress_mgr.reset()

    def _worker():
        try:
            run_m3c2_gui(
                las1_path, las2_path, output_path,
                cyl, normals_r, core_spacing, save_txt,
                True, conf_level, max_depth, use_man_s, s1, s2, progress_mgr,
                sigma_mode=sigma_mode, registration_error=reg_error,
                make_raster=make_raster, raster_resolution=raster_res,
                raster_fields=raster_fields_sel,
                raster_interp_method=raster_interp,
                make_dem=make_dem, dem_resolution=dem_res,
                dem_interp_method=dem_interp,
                quiver_normal=quiver_normal, quiver_motion=quiver_motion,
                quiver_grid_size=quiver_grid, quiver_png=quiver_png,
                quiver_html=quiver_html, quiver_gpkg=quiver_gpkg,
                las_fields=selected_las_fields if selected_las_fields else None,
                profile_lines=prof_lines if prof_lines else None,
                profile_width=profile_w,
                profile_export_gpkg=var_profile_export_gpkg.get(),
                profile_export_txt=var_profile_export_txt.get(),
                make_windrose=make_windrose,
                windrose_bins=windrose_bins_val,
                windrose_sig_only=windrose_sig_only,
                clip_mask_path=clip_mask_path,
                epsg=epsg_val,
                use_pdal=use_pdal,
                make_report=make_report,
                make_histogram=make_histogram,
                make_volume=make_volume,
                volume_dem_resolution=vol_dem_res,
                volume_dem_interp=vol_dem_interp,
                make_batch_csv=var_export_batch_csv.get(),
            )
        finally:
            # Must re-enable the button on the main thread
            root.after(0, lambda: btn_start.config(state="normal"))

    threading.Thread(target=_worker, daemon=True).start()


def toggle_spacing_entry():
    s = "disabled" if var_use_all_points.get() else "normal"
    entry_spacing.config(state=s)


def guess_parameters():
    path = entry_las1.get()
    if not path:
        messagebox.showerror("Chyba", "Najprv vyberte referenčný LAS/LAZ súbor."); return
    try:
        las = laspy.read(path)
        pts = np.column_stack([las.x, las.y, las.z])
    except Exception as e:
        messagebox.showerror("Chyba", f"Chyba čítania súboru: {e}"); return
    n = len(pts)
    if n > 5000:
        idx = np.random.choice(n, 5000, replace=False); pts = pts[idx]
    tree = cKDTree(pts)
    dists, _ = tree.query(pts, k=2, workers=_N_THREADS)
    d_avg = float(np.mean(dists[:, 1]))
    cyl_r  = round(2.0 * d_avg, 3); nr1 = round(3.0 * d_avg, 3)
    nr2    = round(5.0 * d_avg, 3); nr3 = round(8.0 * d_avg, 3)
    spacing = round(2.0 * d_avg, 3)
    entry_cyl.delete(0, tk.END);     entry_cyl.insert(0, str(cyl_r))
    entry_normals.delete(0, tk.END); entry_normals.insert(0, f"{nr1},{nr2},{nr3}")
    if not var_use_all_points.get():
        entry_spacing.delete(0, tk.END); entry_spacing.insert(0, str(spacing))
    messagebox.showinfo("Guess Parameters",
        "Estimated avg. point spacing: {:.4f} m\n\n"
        "Cylinder radius : {} m\n"
        "Normal radii    : {}, {}, {} m\n"
        "Core spacing    : {} m".format(d_avg, cyl_r, nr1, nr2, nr3, spacing))


# ── Profile picker callback ──────────────────────────────────────────
# Holds list of (label, np.ndarray) tuples for the current session
picked_profile_lines = []   # populated by _open_profile_picker()

def _open_profile_picker():
    f1 = entry_las1.get().strip()
    f2 = entry_las2.get().strip()
    if not f1 or not f2:
        messagebox.showerror("Chyba",
            "Please fill in both LAS/LAZ file paths before picking profiles.")
        return

    existing = [pxy for _lbl, pxy in picked_profile_lines]
    picker   = ProfilePicker(root, f1, f2, existing_lines=existing)
    result   = picker.wait()   # blocks until window is closed

    if result is None:
        return   # cancelled

    # Replace picked lines
    picked_profile_lines.clear()
    for i, pxy in enumerate(result):
        lbl = f"Line {i+1}"
        picked_profile_lines.append((lbl, pxy))

    _refresh_profile_listbox()


def _refresh_profile_listbox():
    lbox_profiles.delete(0, tk.END)
    for lbl, pxy in picked_profile_lines:
        x0, y0 = pxy[0]
        x1, y1 = pxy[1]
        length = float(np.linalg.norm(pxy[1] - pxy[0]))
        lbox_profiles.insert(tk.END,
            f"  {lbl}   ({x0:.1f}, {y0:.1f}) → ({x1:.1f}, {y1:.1f})   {length:.1f} m")
    n = len(picked_profile_lines)
    lbl_picked_count.config(
        text=f"{n} line{'s' if n != 1 else ''} picked" if n > 0 else "No lines picked yet")
    btn_clear_picked.config(state="normal" if n > 0 else "disabled")


def _clear_picked_lines():
    if not messagebox.askyesno("Clear", "Remove all interactively picked profile lines?"):
        return
    picked_profile_lines.clear()
    _refresh_profile_listbox()


# ─────────────────────────────────────────────────────────────────────
# Build main window
# ─────────────────────────────────────────────────────────────────────
root = tk.Tk()
_root_ref.append(root)   # enables _ui_error / _ui_info from worker threads
root.title("M3C2 Tool  \u2022  CHECKPOINT 14  \u2014  Samuel Lebó")
root.resizable(True, True)

# ─────────────────────────────────────────────────────────────────────
# Modern theme — color palette and fonts
# ─────────────────────────────────────────────────────────────────────
#  Primary:    #2563eb (blue)     — main accents, buttons
#  Secondary:  #0891b2 (teal)     — secondary actions
#  Success:    #059669 (emerald)  — success, positive
#  Warning:    #d97706 (amber)    — warnings, neutral attention
#  Danger:     #dc2626 (red)      — errors, destructive
#  Text:       #1e293b (slate)    — main text
#  Text muted: #64748b (slate)    — help text, hints
#  Bg:         #f8fafc (slate-50) — main background
#  Surface:    #ffffff            — card / frame background
#  Border:     #e2e8f0 (slate)    — subtle dividers
COLOR = {
    'primary':   '#2563eb',
    'primary_d': '#1d4ed8',
    'primary_l': '#dbeafe',
    'secondary': '#0891b2',
    'success':   '#059669',
    'success_l': '#d1fae5',
    'warning':   '#d97706',
    'warning_l': '#fef3c7',
    'danger':    '#dc2626',
    'text':      '#1e293b',
    'text_mut':  '#64748b',
    'text_hint': '#94a3b8',
    'bg':        '#f8fafc',
    'surface':   '#ffffff',
    'border':    '#e2e8f0',
    'hover':     '#f1f5f9',
}

# Font selection — try modern fonts first, fall back gracefully
import tkinter.font as tkfont
_available_fonts = set(tkfont.families())
if 'Segoe UI' in _available_fonts:
    FONT_FAMILY = 'Segoe UI'
elif 'SF Pro Display' in _available_fonts:
    FONT_FAMILY = 'SF Pro Display'
elif 'Inter' in _available_fonts:
    FONT_FAMILY = 'Inter'
else:
    FONT_FAMILY = 'Arial'

FONT = {
    'body':       (FONT_FAMILY, 9),
    'body_b':     (FONT_FAMILY, 9, 'bold'),
    'small':      (FONT_FAMILY, 8),
    'small_b':    (FONT_FAMILY, 8, 'bold'),
    'label':      (FONT_FAMILY, 9),
    'section':    (FONT_FAMILY, 10, 'bold'),
    'heading':    (FONT_FAMILY, 11, 'bold'),
    'title':      (FONT_FAMILY, 13, 'bold'),
    'button':     (FONT_FAMILY, 9, 'bold'),
}

# Apply root background
root.configure(bg=COLOR['bg'])
# Set min window size
root.minsize(900, 700)

style = ttk.Style()
style.theme_use("clam")

# Progressbar — slick blue
style.configure("TProgressbar",
                troughcolor=COLOR['border'],
                background=COLOR['primary'],
                borderwidth=0,
                thickness=16,
                lightcolor=COLOR['primary'],
                darkcolor=COLOR['primary'])

# Separator
style.configure("TSeparator", background=COLOR['border'])

# Scrollbar
style.configure("Vertical.TScrollbar",
                background=COLOR['border'],
                troughcolor=COLOR['bg'],
                bordercolor=COLOR['bg'],
                arrowcolor=COLOR['text_mut'])

PAD = dict(padx=8, pady=4)

_canvas    = tk.Canvas(root, highlightthickness=0, bg=COLOR['bg'])
_scrollbar = ttk.Scrollbar(root, orient="vertical", command=_canvas.yview,
                            style="Vertical.TScrollbar")
_canvas.configure(yscrollcommand=_scrollbar.set)
_scrollbar.pack(side="right", fill="y")
_canvas.pack(side="left", fill="both", expand=True)

main_frame = tk.Frame(_canvas, bg=COLOR['bg'])
_canvas_win = _canvas.create_window((0, 0), window=main_frame, anchor="nw")

def _on_frame_resize(event): _canvas.configure(scrollregion=_canvas.bbox("all"))
def _on_canvas_resize(event): _canvas.itemconfig(_canvas_win, width=event.width)
def _on_mousewheel(event): _canvas.yview_scroll(int(-1*(event.delta/120)), "units")
main_frame.bind("<Configure>", _on_frame_resize)
_canvas.bind("<Configure>", _on_canvas_resize)
_canvas.bind_all("<MouseWheel>", _on_mousewheel)

# ─────────────────────────────────────────────────────────────────────
# App header — title bar with branding
# ─────────────────────────────────────────────────────────────────────
_header = tk.Frame(main_frame, bg=COLOR['primary'], height=64)
_header.grid(row=0, column=0, columnspan=3, sticky="ew", padx=0, pady=0)
_header.grid_propagate(False)
_header.columnconfigure(0, weight=1)

_header_inner = tk.Frame(_header, bg=COLOR['primary'])
_header_inner.grid(row=0, column=0, sticky="w", padx=16, pady=10)

tk.Label(_header_inner, text="M3C2 Tool",
         font=(FONT_FAMILY, 16, 'bold'),
         bg=COLOR['primary'], fg='white'
         ).grid(row=0, column=0, sticky="w")
tk.Label(_header_inner, text="Point Cloud Change Detection  \u2022  CHECKPOINT 14",
         font=(FONT_FAMILY, 9),
         bg=COLOR['primary'], fg='#bfdbfe'
         ).grid(row=1, column=0, sticky="w")

_header_right = tk.Frame(_header, bg=COLOR['primary'])
_header_right.grid(row=0, column=1, sticky="e", padx=16, pady=10)
tk.Label(_header_right, text="Samuel Lebó  \u2022  GEOsys s.r.o.",
         font=(FONT_FAMILY, 8),
         bg=COLOR['primary'], fg='#bfdbfe'
         ).grid(row=0, column=0, sticky="e")
tk.Label(_header_right, text="github.com/Kostka22/Geology_M3C2-from-point-cloud",
         font=(FONT_FAMILY, 7),
         bg=COLOR['primary'], fg='#93c5fd'
         ).grid(row=1, column=0, sticky="e")
_header.columnconfigure(1, weight=1)


def _lf(parent, text, row, col=0, colspan=3, sticky="ew", pady=(12, 4), icon=""):
    """Modern card-style section frame with colored header bar."""
    # Outer container
    wrapper = tk.Frame(parent, bg=COLOR['bg'])
    wrapper.grid(row=row, column=col, columnspan=colspan, sticky=sticky, padx=12, pady=pady)
    wrapper.columnconfigure(0, weight=1)

    # Section title bar
    title_bar = tk.Frame(wrapper, bg=COLOR['surface'], bd=0)
    title_bar.grid(row=0, column=0, sticky="ew")
    title_bar.columnconfigure(0, weight=1)

    # Colored left accent + title
    accent = tk.Frame(title_bar, bg=COLOR['primary'], width=3)
    accent.grid(row=0, column=0, sticky="nsw", rowspan=2)
    title_inner = tk.Frame(title_bar, bg=COLOR['surface'])
    title_inner.grid(row=0, column=1, sticky="ew", padx=(10, 8), pady=(8, 4))
    title_inner.columnconfigure(1, weight=1)
    if icon:
        tk.Label(title_inner, text=icon, font=(FONT_FAMILY, 12),
                 bg=COLOR['surface'], fg=COLOR['primary']
                 ).grid(row=0, column=0, sticky="w", padx=(0, 6))
    tk.Label(title_inner, text=text, font=FONT['heading'],
             bg=COLOR['surface'], fg=COLOR['text']
             ).grid(row=0, column=1, sticky="w")

    # Content frame
    f = tk.Frame(wrapper, bg=COLOR['surface'], padx=14, pady=10, bd=0,
                  highlightbackground=COLOR['border'], highlightthickness=1)
    f.grid(row=1, column=0, sticky="ew")
    f.columnconfigure(1, weight=1)

    # Move title bar inside the border frame (visual trick — re-parent title frame)
    # Instead: pack title bar above without border
    title_bar.config(highlightbackground=COLOR['border'], highlightthickness=1,
                     bd=0)
    # Remove top border of content frame (looks like continuous card)
    return f

def _entry(parent, row, label, default, col=1, w=12, hint=None):
    tk.Label(parent, text=label, font=FONT['label'],
             bg=COLOR['surface'], fg=COLOR['text']
             ).grid(row=row, column=0, sticky="e", **PAD)
    e = tk.Entry(parent, width=w, font=FONT['body'],
                  bg='white', fg=COLOR['text'],
                  relief='solid', bd=1,
                  highlightbackground=COLOR['border'],
                  highlightcolor=COLOR['primary'],
                  highlightthickness=1,
                  insertbackground=COLOR['primary'])
    e.insert(0, default)
    e.grid(row=row, column=col, sticky="w", **PAD)
    if hint:
        tk.Label(parent, text=hint, font=FONT['small'],
                 bg=COLOR['surface'], fg=COLOR['text_mut']
                 ).grid(row=row, column=col+1, sticky="w", padx=(4, 8))
    return e

def _primary_btn(parent, text, command, **grid_kw):
    """Modern flat primary button with hover."""
    b = tk.Button(parent, text=text, command=command,
                   bg=COLOR['primary'], fg='white',
                   activebackground=COLOR['primary_d'],
                   activeforeground='white',
                   font=FONT['button'], relief='flat', bd=0,
                   padx=14, pady=6, cursor='hand2')
    b.bind('<Enter>', lambda e: b.config(bg=COLOR['primary_d']))
    b.bind('<Leave>', lambda e: b.config(bg=COLOR['primary']))
    if grid_kw:
        b.grid(**grid_kw)
    return b

def _secondary_btn(parent, text, command, **grid_kw):
    """Outlined secondary button."""
    b = tk.Button(parent, text=text, command=command,
                   bg=COLOR['surface'], fg=COLOR['primary'],
                   activebackground=COLOR['primary_l'],
                   activeforeground=COLOR['primary_d'],
                   font=FONT['button'], relief='solid', bd=1,
                   padx=12, pady=5, cursor='hand2',
                   highlightbackground=COLOR['primary'],
                   highlightthickness=0)
    b.bind('<Enter>', lambda e: b.config(bg=COLOR['primary_l']))
    b.bind('<Leave>', lambda e: b.config(bg=COLOR['surface']))
    if grid_kw:
        b.grid(**grid_kw)
    return b

def _browse_btn(parent, command, **grid_kw):
    """Small browse button."""
    b = tk.Button(parent, text="\U0001f4c1 Browse", command=command,
                   bg=COLOR['hover'], fg=COLOR['text'],
                   activebackground=COLOR['border'],
                   font=FONT['small'], relief='flat', bd=0,
                   padx=10, pady=4, cursor='hand2')
    b.bind('<Enter>', lambda e: b.config(bg=COLOR['border']))
    b.bind('<Leave>', lambda e: b.config(bg=COLOR['hover']))
    if grid_kw:
        b.grid(**grid_kw)
    return b

# ── Files ──────────────────────────────────────────────────────────
ff = _lf(main_frame, "Input / Output Files", row=1, icon="\U0001f4c2")
for r, (lbl, hint) in enumerate([
    ("Reference LAS/LAZ",       "Epoch 1 \u2014 the older / reference point cloud"),
    ("Comparison LAS/LAZ",      "Epoch 2 \u2014 the newer / comparison point cloud"),
    ("Output LAS/LAZ (save)",   "Where to save results (full M3C2 scalar fields)"),
]):
    tk.Label(ff, text=lbl, font=FONT['label'],
             bg=COLOR['surface'], fg=COLOR['text']
             ).grid(row=r*2, column=0, sticky="e", **PAD)
    tk.Label(ff, text=hint, font=FONT['small'],
             bg=COLOR['surface'], fg=COLOR['text_hint']
             ).grid(row=r*2+1, column=1, sticky="w", padx=10, pady=(0, 4))
entries = []
for r in range(3):
    e = tk.Entry(ff, width=55, font=FONT['body'],
                  bg='white', fg=COLOR['text'],
                  relief='solid', bd=1,
                  highlightbackground=COLOR['border'],
                  highlightcolor=COLOR['primary'],
                  highlightthickness=1,
                  insertbackground=COLOR['primary'])
    e.grid(row=r*2, column=1, sticky="ew", **PAD)
    entries.append(e)
    s = (r == 2)
    _browse_btn(ff, command=lambda e=e, s=s: browse_file(e, s),
                row=r*2, column=2, **PAD)
ff.columnconfigure(1, weight=1)
entry_las1, entry_las2, entry_output = entries

# ── M3C2 Parameters ────────────────────────────────────────────────
pf = _lf(main_frame, "M3C2 Parameters", row=2, icon="\U00002699")
entry_cyl = _entry(pf, 0, "Cylinder radius [m]", "2.0")
_guess_btn = tk.Button(pf, text="\u2728  Guess Parameters",
                        command=guess_parameters,
                        bg=COLOR['warning'], fg='white',
                        font=FONT['button'], relief='flat', bd=0,
                        padx=12, pady=6, cursor='hand2',
                        activebackground='#b45309',
                        activeforeground='white')
_guess_btn.grid(row=0, column=2, sticky="w", **PAD)
_guess_btn.bind('<Enter>', lambda e: _guess_btn.config(bg='#b45309'))
_guess_btn.bind('<Leave>', lambda e: _guess_btn.config(bg=COLOR['warning']))
var_multithreading = tk.BooleanVar(value=True)
tk.Checkbutton(pf, text=f"Allow multithreading  ({(os.cpu_count() or 2) - 1} threads, 1 left for OS)",
               variable=var_multithreading,
               font=FONT['small'], bg=COLOR['surface'],
               fg=COLOR['text_mut'], activebackground=COLOR['surface'],
               selectcolor=COLOR['primary_l']
               ).grid(row=1, column=2, sticky="w", padx=6, pady=(0, 2))

# ── Computation engine ──────────────────────────────────────────────
_pdal_ok, _pdal_ver = _check_pdal()
var_engine = tk.StringVar(value="py4dgeo")
eng_fr = tk.Frame(pf, bg=COLOR['surface'])
eng_fr.grid(row=2, column=2, sticky="w", padx=6, pady=(2, 4))
tk.Label(eng_fr, text="Engine:", font=FONT['small_b'],
         bg=COLOR['surface'], fg=COLOR['text_mut']
         ).grid(row=0, column=0, sticky="w")
rb_py4d = tk.Radiobutton(eng_fr, text="py4dgeo", variable=var_engine,
                          value="py4dgeo", font=FONT['small'],
                          bg=COLOR['surface'], fg=COLOR['text'],
                          activebackground=COLOR['surface'],
                          selectcolor=COLOR['primary_l'])
rb_py4d.grid(row=0, column=1, sticky="w", padx=(6, 0))
rb_pdal = tk.Radiobutton(eng_fr, text="PDAL", variable=var_engine,
                          value="pdal", font=FONT['small'],
                          bg=COLOR['surface'], fg=COLOR['text'],
                          activebackground=COLOR['surface'],
                          selectcolor=COLOR['primary_l'],
                          state="normal" if _pdal_ok else "disabled")
rb_pdal.grid(row=0, column=2, sticky="w", padx=(6, 0))
_pdal_hint = (_pdal_ver if _pdal_ok
              else "not found — conda install -c conda-forge pdal")
tk.Label(eng_fr, text=_pdal_hint, font=(FONT_FAMILY, 7),
         bg=COLOR['surface'],
         fg=COLOR['success'] if _pdal_ok else COLOR['danger']
         ).grid(row=0, column=3, sticky="w", padx=(8, 0))
entry_normals = _entry(pf, 3, "Normal radii [m] (comma)", "1.5,2.5,3.5", w=22)
entry_spacing = _entry(pf, 4, "Corepoint spacing [m]", "1.0")
var_use_all_points = tk.BooleanVar()
tk.Checkbutton(pf, text="Use all points as corepoints",
               variable=var_use_all_points, command=toggle_spacing_entry
               ).grid(row=4, column=2, sticky="w", **PAD)
entry_max_depth = _entry(pf, 5, "Max depth detection [m]", "8.0")
tk.Label(pf, text="(0 = no limit)", fg="#666666").grid(row=5, column=2, sticky="w", **PAD)
tk.Label(pf, text="Confidence level").grid(row=6, column=0, sticky="e", **PAD)
var_confidence = tk.StringVar(value="0.95")
tk.OptionMenu(pf, var_confidence, "0.95", "0.99", "0.90").grid(row=6, column=1, sticky="w", **PAD)

# ── Sigma mode — three mutually exclusive radio buttons ────────────
var_sigma_mode = tk.StringVar(value="per_point")

def toggle_sigma_entries():
    mode = var_sigma_mode.get()
    entry_reg_error.config(state="normal"   if mode == "per_point"    else "disabled")
    entry_sigma1.config(   state="normal"   if mode == "manual_sigma" else "disabled")
    entry_sigma2.config(   state="normal"   if mode == "manual_sigma" else "disabled")

tk.Label(pf, text="Sigma mode").grid(row=7, column=0, sticky="ne", **PAD)
sigma_frame = tk.Frame(pf)
sigma_frame.grid(row=7, column=1, columnspan=2, sticky="w", **PAD)
tk.Radiobutton(sigma_frame,
               text="A — Per-point LoD  (py4dgeo + registration error)",
               variable=var_sigma_mode, value="per_point",
               command=toggle_sigma_entries).pack(anchor="w")
tk.Radiobutton(sigma_frame,
               text="B — Global auto  (py4dgeo vypočíta σ1, σ2 z SPREAD1/2 sám)",
               variable=var_sigma_mode, value="global_auto",
               command=toggle_sigma_entries).pack(anchor="w")
tk.Radiobutton(sigma_frame,
               text="C — Manual σ1, σ2  (zadáš presnosti mračien)",
               variable=var_sigma_mode, value="manual_sigma",
               command=toggle_sigma_entries).pack(anchor="w")

entry_reg_error = _entry(pf, 8, "Registration error [m]", "0.05")
tk.Label(pf, text="Mode A — ALS sys. chyba (~0.05 m)", fg="#666666"
         ).grid(row=8, column=2, sticky="w", **PAD)
entry_sigma1 = _entry(pf, 9,  "Sigma 1 [m]", "0.05"); entry_sigma1.config(state="disabled")
tk.Label(pf, text="Mode C — presnosť epoch 1 / epoch 2", fg="#666666"
         ).grid(row=9, column=2, sticky="w", **PAD)
entry_sigma2 = _entry(pf, 10, "Sigma 2 [m]", "0.05"); entry_sigma2.config(state="disabled")

# ── Clip raster by extent ──────────────────────────────────────────
var_clip_raster = tk.BooleanVar()

def toggle_clip_entries():
    s = "normal" if var_clip_raster.get() else "disabled"
    entry_clip_mask.config(state=s)
    btn_clip_browse.config(state=s)

tk.Checkbutton(pf, text="Clip rasters by extent (SHP / GPKG)",
               variable=var_clip_raster, command=toggle_clip_entries,
               font=("Arial", 9)
               ).grid(row=12, column=0, columnspan=2, sticky="w", **PAD)

entry_clip_mask = tk.Entry(pf, width=40, state="disabled")
entry_clip_mask.grid(row=13, column=1, sticky="ew", **PAD)
tk.Label(pf, text="Mask file:").grid(row=13, column=0, sticky="e", **PAD)

def _browse_clip_mask():
    p = filedialog.askopenfilename(
        filetypes=[("Vector files", "*.shp *.gpkg *.geojson"),
                   ("Shapefile", "*.shp"),
                   ("GeoPackage", "*.gpkg"),
                   ("All files", "*.*")])
    if p:
        entry_clip_mask.delete(0, tk.END)
        entry_clip_mask.insert(0, p)

btn_clip_browse = tk.Button(pf, text="Browse\u2026", state="disabled",
                             command=_browse_clip_mask)
btn_clip_browse.grid(row=13, column=2, sticky="w", **PAD)
tk.Label(pf, text="Clips all GeoTIFF outputs (rasters + DEM) to the first layer of the file.",
         fg="#666666", font=("Arial", 8)
         ).grid(row=14, column=1, columnspan=2, sticky="w", padx=6)

# ── EPSG code ──────────────────────────────────────────────────────
ttk.Separator(pf, orient="horizontal").grid(
    row=15, column=0, columnspan=3, sticky="ew", padx=4, pady=(6, 4))
tk.Label(pf, text="EPSG code:").grid(row=16, column=0, sticky="e", **PAD)
entry_epsg = tk.Entry(pf, width=10)
entry_epsg.grid(row=16, column=1, sticky="w", **PAD)
tk.Label(pf,
         text="Written as CRS into all GeoTIFF outputs  "
              "(e.g. 8353 for S-JTSK/Krovak East North  ·  32634 for UTM 34N  ·  leave blank to skip)",
         fg="#666666", font=("Arial", 8), wraplength=400, justify="left"
         ).grid(row=16, column=2, sticky="w", padx=4)

# ── Output Options ──────────────────────────────────────────────────
of = _lf(main_frame, "Output Options", row=3, icon="\U0001f4be")
var_txt = tk.BooleanVar()
tk.Checkbutton(of, text="Also save results to TXT", variable=var_txt
               ).grid(row=0, column=0, columnspan=4, sticky="w", **PAD)
var_make_report = tk.BooleanVar(value=True)
tk.Checkbutton(of, text="Export PDF analysis report  (_report.pdf)",
               variable=var_make_report, font=("Arial", 9)
               ).grid(row=1, column=0, columnspan=4, sticky="w", **PAD)
var_make_histogram = tk.BooleanVar(value=True)
tk.Checkbutton(of, text="Export distance histogram  (_histogram.png)",
               variable=var_make_histogram, font=("Arial", 9)
               ).grid(row=2, column=0, columnspan=4, sticky="w", **PAD)
var_export_batch_csv = tk.BooleanVar(value=False)
tk.Checkbutton(of, text="Export batch stats  (batch_stats.csv + batch_stats_analyza.xlsx)",
               variable=var_export_batch_csv, font=("Arial", 9)
               ).grid(row=3, column=0, columnspan=4, sticky="w", **PAD)
tk.Label(of, text="LAS extra fields:").grid(row=4, column=0, sticky="ne", **PAD)
las_fld_fr = tk.Frame(of); las_fld_fr.grid(row=4, column=1, columnspan=3, sticky="w", **PAD)
LAS_OPTIONAL_FIELDS = ["Nz","Nxy","DX","DY","DZ","LOD","SPREAD1","SPREAD2","NORMAL_SCALE"]
las_field_vars = {}; las_field_chks = {}
for i, fname in enumerate(LAS_OPTIONAL_FIELDS):
    var = tk.BooleanVar(value=True); las_field_vars[fname] = var
    chk = tk.Checkbutton(las_fld_fr, text=fname, variable=var); las_field_chks[fname] = chk
    chk.grid(row=i // 4, column=i % 4, sticky="w", padx=6)
# ── Raster + DEM ────────────────────────────────────────────────────
RASTER_FIELDS  = ["M3C2_DISTANCE","LOD","SPREAD1","SPREAD2","SIGNIFICANT","Nz","Nxy","DX","DY","DZ","NORMAL_SCALE"]
INTERP_METHODS = ["IDW","TIN","Nearest","Kriging"]
raster_field_vars = {}; raster_field_chks = {}

rd_row = tk.Frame(main_frame, bg=COLOR['bg'])
rd_row.grid(row=4, column=0, columnspan=3, sticky="ew", padx=12, pady=(12, 4))
rd_row.columnconfigure(0, weight=1); rd_row.columnconfigure(1, weight=1)

# ── Raster card ──
_r_wrap = tk.Frame(rd_row, bg=COLOR['bg'])
_r_wrap.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
_r_wrap.columnconfigure(0, weight=1)
_r_title = tk.Frame(_r_wrap, bg=COLOR['surface'],
                     highlightbackground=COLOR['border'], highlightthickness=1)
_r_title.grid(row=0, column=0, sticky="ew")
_r_title.columnconfigure(0, weight=1)
tk.Frame(_r_title, bg=COLOR['secondary'], width=3
         ).grid(row=0, column=0, sticky="nsw", rowspan=2)
_r_ti = tk.Frame(_r_title, bg=COLOR['surface'])
_r_ti.grid(row=0, column=1, sticky="ew", padx=(10, 8), pady=(8, 4))
tk.Label(_r_ti, text="\U0001f5fa  Raster Export",
         font=FONT['heading'], bg=COLOR['surface'], fg=COLOR['text']
         ).grid(row=0, column=0, sticky="w")
tk.Label(_r_ti, text="Core Points \u2192 GeoTIFF",
         font=FONT['small'], bg=COLOR['surface'], fg=COLOR['text_hint']
         ).grid(row=1, column=0, sticky="w")
rf = tk.Frame(_r_wrap, bg=COLOR['surface'], padx=14, pady=10,
               highlightbackground=COLOR['border'], highlightthickness=1)
rf.grid(row=1, column=0, sticky="ew")
rf.columnconfigure(1, weight=1)
var_make_raster = tk.BooleanVar()
tk.Checkbutton(rf, text="Enable raster export", variable=var_make_raster,
               command=toggle_raster_entries,
               font=FONT['body'], bg=COLOR['surface'], fg=COLOR['text'],
               activebackground=COLOR['surface'], selectcolor=COLOR['primary_l']
               ).grid(row=0, column=0, columnspan=3, sticky="w", padx=4, pady=2)
tk.Label(rf, text="Resolution [m]", font=FONT['label'],
         bg=COLOR['surface'], fg=COLOR['text']
         ).grid(row=1, column=0, sticky="e", **PAD)
entry_raster_res = tk.Entry(rf, width=8, state="disabled", font=FONT['body'],
                             bg='white', relief='solid', bd=1,
                             highlightbackground=COLOR['border'], highlightthickness=1)
entry_raster_res.insert(0, "0.5")
entry_raster_res.grid(row=1, column=1, sticky="w", **PAD)
tk.Label(rf, text="Interpolation", font=FONT['label'],
         bg=COLOR['surface'], fg=COLOR['text']
         ).grid(row=2, column=0, sticky="e", **PAD)
var_raster_interp = tk.StringVar(value="IDW")
option_raster_interp = tk.OptionMenu(rf, var_raster_interp, *INTERP_METHODS)
option_raster_interp.config(state="disabled", width=10, font=FONT['body'],
                             bg=COLOR['surface'], fg=COLOR['text'],
                             activebackground=COLOR['hover'],
                             relief='solid', bd=1,
                             highlightbackground=COLOR['border'], highlightthickness=1)
option_raster_interp.grid(row=2, column=1, sticky="w", **PAD)
tk.Label(rf, text="Fields:", font=FONT['label'],
         bg=COLOR['surface'], fg=COLOR['text']
         ).grid(row=3, column=0, sticky="ne", **PAD)
rf_fld = tk.Frame(rf, bg=COLOR['surface'])
rf_fld.grid(row=3, column=1, columnspan=2, sticky="w", **PAD)
for i, fname in enumerate(RASTER_FIELDS):
    var = tk.BooleanVar(value=(fname == "M3C2_DISTANCE"))
    raster_field_vars[fname] = var
    chk = tk.Checkbutton(rf_fld, text=fname, variable=var, state="disabled",
                          font=FONT['small'], bg=COLOR['surface'], fg=COLOR['text'],
                          activebackground=COLOR['surface'],
                          selectcolor=COLOR['primary_l'])
    raster_field_chks[fname] = chk
    chk.grid(row=i//3, column=i%3, sticky="w", padx=4)

# ── DEM card ──
_d_wrap = tk.Frame(rd_row, bg=COLOR['bg'])
_d_wrap.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
_d_wrap.columnconfigure(0, weight=1)
_d_title = tk.Frame(_d_wrap, bg=COLOR['surface'],
                     highlightbackground=COLOR['border'], highlightthickness=1)
_d_title.grid(row=0, column=0, sticky="ew")
_d_title.columnconfigure(0, weight=1)
tk.Frame(_d_title, bg=COLOR['secondary'], width=3
         ).grid(row=0, column=0, sticky="nsw", rowspan=2)
_d_ti = tk.Frame(_d_title, bg=COLOR['surface'])
_d_ti.grid(row=0, column=1, sticky="ew", padx=(10, 8), pady=(8, 4))
tk.Label(_d_ti, text="\u26f0  DEM Export",
         font=FONT['heading'], bg=COLOR['surface'], fg=COLOR['text']
         ).grid(row=0, column=0, sticky="w")
tk.Label(_d_ti, text="Z values \u2192 GeoTIFF",
         font=FONT['small'], bg=COLOR['surface'], fg=COLOR['text_hint']
         ).grid(row=1, column=0, sticky="w")
df = tk.Frame(_d_wrap, bg=COLOR['surface'], padx=14, pady=10,
               highlightbackground=COLOR['border'], highlightthickness=1)
df.grid(row=1, column=0, sticky="ew")
df.columnconfigure(1, weight=1)
var_make_dem = tk.BooleanVar()
tk.Checkbutton(df, text="Enable DEM export", variable=var_make_dem,
               command=toggle_dem_entries,
               font=FONT['body'], bg=COLOR['surface'], fg=COLOR['text'],
               activebackground=COLOR['surface'], selectcolor=COLOR['primary_l']
               ).grid(row=0, column=0, columnspan=2, sticky="w", padx=4, pady=2)
tk.Label(df, text="Resolution [m]", font=FONT['label'],
         bg=COLOR['surface'], fg=COLOR['text']
         ).grid(row=1, column=0, sticky="e", **PAD)
entry_dem_res = tk.Entry(df, width=8, state="disabled", font=FONT['body'],
                          bg='white', relief='solid', bd=1,
                          highlightbackground=COLOR['border'], highlightthickness=1)
entry_dem_res.insert(0, "0.1")
entry_dem_res.grid(row=1, column=1, sticky="w", **PAD)
tk.Label(df, text="Interpolation", font=FONT['label'],
         bg=COLOR['surface'], fg=COLOR['text']
         ).grid(row=2, column=0, sticky="e", **PAD)
var_dem_interp = tk.StringVar(value="IDW")
option_dem_interp = tk.OptionMenu(df, var_dem_interp, *INTERP_METHODS)
option_dem_interp.config(state="disabled", width=10, font=FONT['body'],
                          bg=COLOR['surface'], fg=COLOR['text'],
                          activebackground=COLOR['hover'],
                          relief='solid', bd=1,
                          highlightbackground=COLOR['border'], highlightthickness=1)
option_dem_interp.grid(row=2, column=1, sticky="w", **PAD)

# ── Quiver Export ───────────────────────────────────────────────────
qf = _lf(main_frame, "Quiver Export", row=5, icon="\U000027a1"); qf.columnconfigure(1, weight=1)
var_quiver_normal = tk.BooleanVar(); var_quiver_motion = tk.BooleanVar()
var_quiver_png    = tk.BooleanVar(value=True)
var_quiver_html   = tk.BooleanVar(value=True)
var_quiver_gpkg   = tk.BooleanVar(value=True)
_qfmt_chks = []

mode_fr = tk.Frame(qf); mode_fr.grid(row=0, column=0, columnspan=4, sticky="w", padx=4, pady=(2,4))
tk.Checkbutton(mode_fr, text="Export normal-direction quiver",
               variable=var_quiver_normal, command=toggle_quiver_entries,
               font=("Arial",9)).grid(row=0, column=0, sticky="w", padx=(0,20))
tk.Checkbutton(mode_fr, text="Export motion vector  (DX · DY)",
               variable=var_quiver_motion, command=toggle_quiver_entries,
               font=("Arial",9)).grid(row=0, column=1, sticky="w")

tk.Label(qf, text="Grid cell [m]").grid(row=1, column=0, sticky="e", **PAD)
entry_quiver_grid = tk.Entry(qf, width=8, state="disabled"); entry_quiver_grid.insert(0,"1.0")
entry_quiver_grid.grid(row=1, column=1, sticky="w", **PAD)
tk.Label(qf, text="(independent of corepoint spacing)", fg="#666666").grid(row=1, column=2, sticky="w", **PAD)

tk.Label(qf, text="Formats:").grid(row=2, column=0, sticky="e", **PAD)
qfmt_fr = tk.Frame(qf); qfmt_fr.grid(row=2, column=1, columnspan=3, sticky="w", **PAD)
for i, (var, lbl, tip) in enumerate([(var_quiver_png,"PNG","static image"),
                                       (var_quiver_html,"HTML","interactive Plotly"),
                                       (var_quiver_gpkg,"GPKG","QGIS vector arrows")]):
    chk = tk.Checkbutton(qfmt_fr, text=f"{lbl} ({tip})", variable=var, state="disabled")
    chk.grid(row=0, column=i, sticky="w", padx=10); _qfmt_chks.append(chk)

tk.Label(qf, text="Both modes can run together.  Normal → _quiver_normal.*  ·  Motion → _quiver_motion.*",
         fg="#666666", font=("Arial",8), wraplength=580, justify="left"
         ).grid(row=3, column=0, columnspan=4, sticky="w", padx=4, pady=(0,2))

# ══════════════════════════════════════════════════════════════════════
# ── Profile Cut Export ─────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════
# ── Volume Change Estimation ─────────────────────────────────────────
vcf = _lf(main_frame, "Volume Change Estimation", row=6, icon="\U0001f4ca")
vcf.columnconfigure(1, weight=1)

var_make_volume      = tk.BooleanVar()
var_vol_dem_interp   = tk.StringVar(value="IDW")
_vol_controls        = []

def toggle_volume_entries():
    s = "normal" if var_make_volume.get() else "disabled"
    for w in _vol_controls:
        w.config(state=s)

tk.Checkbutton(vcf, text="Enable volume change estimation",
               variable=var_make_volume, command=toggle_volume_entries
               ).grid(row=0, column=0, columnspan=3, sticky="w", padx=4, pady=(2, 4))

tk.Label(vcf, text="DoD grid res [m]:").grid(row=1, column=0, sticky="e", **PAD)
entry_vol_dem_res = tk.Entry(vcf, width=8, state="disabled")
entry_vol_dem_res.insert(0, "0.5")
entry_vol_dem_res.grid(row=1, column=1, sticky="w", **PAD)
_vol_controls.append(entry_vol_dem_res)

tk.Label(vcf, text="DoD interpolation:").grid(row=1, column=2, sticky="e", **PAD)
om_vol_interp = tk.OptionMenu(vcf, var_vol_dem_interp, "IDW", "TIN", "Nearest")
om_vol_interp.config(state="disabled", width=8)
om_vol_interp.grid(row=1, column=3, sticky="w", **PAD)
_vol_controls.append(om_vol_interp)

tk.Label(vcf,
         text="Computes volume gain/loss by two methods: M3C2 column integration and "
              "DEM difference (DoD). Both results + % difference appear in the PDF report.",
         fg="#666666", font=("Arial", 8), wraplength=580, justify="left"
         ).grid(row=2, column=0, columnspan=4, sticky="w", padx=4, pady=(0, 2))

# ── Wind Rose Export ────────────────────────────────────────────────
wrf = _lf(main_frame, "Wind Rose Export", row=7, icon="\U0001f9ed")
wrf.columnconfigure(1, weight=1)

var_make_windrose    = tk.BooleanVar()
var_windrose_sig_only = tk.BooleanVar(value=True)
var_windrose_bins    = tk.StringVar(value="16")
_windrose_controls   = []

def toggle_windrose_entries():
    s = "normal" if var_make_windrose.get() else "disabled"
    for w in _windrose_controls:
        w.config(state=s)

tk.Checkbutton(wrf, text="Enable wind rose export",
               variable=var_make_windrose, command=toggle_windrose_entries
               ).grid(row=0, column=0, columnspan=4, sticky="w", padx=4, pady=(2, 4))

tk.Label(wrf, text="Direction bins:").grid(row=1, column=0, sticky="e", **PAD)
om_bins = tk.OptionMenu(wrf, var_windrose_bins, "8", "16", "32", "36")
om_bins.config(state="disabled", width=5)
om_bins.grid(row=1, column=1, sticky="w", **PAD)
_windrose_controls.append(om_bins)

tk.Label(wrf,
         text="Two-panel PNG: (a) horizontal azimuth rose  ·  (b) vertical angle rose  ·  "
              "grey = not significant  ·  coloured by 3-D displacement magnitude  →  _windrose.png",
         fg="#666666", font=("Arial", 8), wraplength=580, justify="left"
         ).grid(row=2, column=0, columnspan=4, sticky="w", padx=4, pady=(0, 2))

# ── Profile Cut Export ───────────────────────────────────────────────
prf = _lf(main_frame, "Profile Cut Export", row=8, icon="\U0001f4cf")
prf.columnconfigure(1, weight=1)

var_make_profile = tk.BooleanVar()
tk.Checkbutton(prf, text="Enable profile export",
               variable=var_make_profile, command=toggle_profile_entries
               ).grid(row=0, column=0, columnspan=3, sticky="w", padx=4, pady=(2, 6))

# ── Corridor width (shared) ─────────────────────────────────────────
tk.Label(prf, text="Corridor width [m]").grid(row=1, column=0, sticky="e", **PAD)
entry_profile_width = tk.Entry(prf, width=8, state="disabled")
entry_profile_width.insert(0, "50.0"); entry_profile_width.grid(row=1, column=1, sticky="w", **PAD)
tk.Label(prf, text="points within ± width/2 of each line",
         fg="#666666").grid(row=1, column=2, sticky="w", **PAD)

# ── Interactive picker ──────────────────────────────────────────────
sep1 = ttk.Separator(prf, orient="horizontal")
sep1.grid(row=2, column=0, columnspan=3, sticky="ew", padx=4, pady=(6, 4))

pick_fr = tk.Frame(prf)
pick_fr.grid(row=3, column=0, columnspan=3, sticky="ew", padx=4, pady=2)
pick_fr.columnconfigure(2, weight=1)

btn_pick_profiles = tk.Button(
    pick_fr,
    text="\u25b6   Pick profiles interactively…",
    command=_open_profile_picker,
    state="disabled",
    font=("Arial", 9, "bold"), relief="flat",
    bg="#1565c0", fg="white", padx=12, pady=5,
    activebackground="#0d47a1", activeforeground="white",
    cursor="hand2")
btn_pick_profiles.grid(row=0, column=0, sticky="w", padx=(0, 12))

tk.Label(pick_fr, text="reference = blue   ·   comparison = red",
         fg="#666666", font=("Arial", 8)).grid(row=0, column=1, sticky="w")

# Picked-lines listbox
lbox_fr = tk.Frame(prf)
lbox_fr.grid(row=4, column=0, columnspan=3, sticky="ew", padx=4, pady=(4, 2))
lbox_fr.columnconfigure(0, weight=1)

lbl_picked_count = tk.Label(lbox_fr, text="No lines picked yet",
                              font=("Arial", 8), fg="#888888")
lbl_picked_count.grid(row=0, column=0, sticky="w")

btn_clear_picked = tk.Button(lbox_fr, text="Clear picked lines",
                              command=_clear_picked_lines, state="disabled",
                              font=("Arial", 8), relief="flat", bg="#eeeeee", padx=6)
btn_clear_picked.grid(row=0, column=1, sticky="e")

lbox_profiles = tk.Listbox(lbox_fr, height=4, font=("Courier", 8),
                             selectmode="browse", relief="sunken", bd=1)
lbox_profiles.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(2, 0))
_lbox_sb = ttk.Scrollbar(lbox_fr, orient="vertical", command=lbox_profiles.yview)
_lbox_sb.grid(row=1, column=2, sticky="ns")
lbox_profiles.config(yscrollcommand=_lbox_sb.set)

# ── TXT fallback ────────────────────────────────────────────────────
sep2 = ttk.Separator(prf, orient="horizontal")
sep2.grid(row=5, column=0, columnspan=3, sticky="ew", padx=4, pady=(6, 4))

tk.Label(prf, text="Or load from TXT  (Y X per line):",
         fg="#555555", font=("Arial", 8)).grid(row=6, column=0, columnspan=3,
                                                sticky="w", padx=4, pady=(0, 2))
tk.Label(prf, text="Profile TXT").grid(row=7, column=0, sticky="e", **PAD)
entry_profile_txt = tk.Entry(prf, width=46, state="disabled")
entry_profile_txt.grid(row=7, column=1, sticky="ew", **PAD)

def _browse_profile():
    p = filedialog.askopenfilename(
        filetypes=[("Text files", "*.txt *.csv *.dat"), ("All files", "*.*")])
    if p:
        entry_profile_txt.delete(0, tk.END); entry_profile_txt.insert(0, p)

btn_profile_browse = tk.Button(prf, text="Browse\u2026", state="disabled",
                                command=_browse_profile)
btn_profile_browse.grid(row=7, column=2, **PAD)

tk.Label(prf,
         text="TXT adds one extra profile on top of any interactively picked lines.  "
              "Output files: _profile_01.png, _profile_02.png …",
         fg="#666666", font=("Arial", 8), wraplength=560, justify="left"
         ).grid(row=8, column=0, columnspan=3, sticky="w", padx=4, pady=(0, 2))

# ── Export picked lines ────────────────────────────────────────────
sep_pl = ttk.Separator(prf, orient="horizontal")
sep_pl.grid(row=9, column=0, columnspan=3, sticky="ew", padx=4, pady=(4, 3))
tk.Label(prf, text="Export picked lines:",
         fg="#555555", font=("Arial", 8)
         ).grid(row=10, column=0, sticky="e", **PAD)
var_profile_export_gpkg = tk.BooleanVar(value=False)
var_profile_export_txt  = tk.BooleanVar(value=True)
chk_pl_fr = tk.Frame(prf)
chk_pl_fr.grid(row=10, column=1, columnspan=2, sticky="w", **PAD)
tk.Checkbutton(chk_pl_fr, text="GPKG  (loadable in QGIS)",
               variable=var_profile_export_gpkg, font=("Arial", 9)
               ).grid(row=0, column=0, sticky="w", padx=(0, 16))
tk.Checkbutton(chk_pl_fr, text="TXT trajectory  (reloadable into this tool)",
               variable=var_profile_export_txt, font=("Arial", 9)
               ).grid(row=0, column=1, sticky="w")

# ── Progress bars ────────────────────────────────────────────────────
_progress_wrap = tk.Frame(main_frame, bg=COLOR['surface'],
                            highlightbackground=COLOR['border'],
                            highlightthickness=1)
_progress_wrap.grid(row=9, column=0, columnspan=3, sticky="ew",
                     padx=12, pady=(14, 6))
_progress_wrap.columnconfigure(0, weight=1)

# Progress header
_ph = tk.Frame(_progress_wrap, bg=COLOR['surface'])
_ph.grid(row=0, column=0, sticky="ew", padx=14, pady=(10, 4))
tk.Label(_ph, text="\u25cb  Processing Status", font=FONT['heading'],
         bg=COLOR['surface'], fg=COLOR['text']
         ).pack(side="left")

lbl_current_task = tk.Label(_progress_wrap, text="Current Task: Ready", anchor="w",
                             font=FONT['body'],
                             bg=COLOR['surface'], fg=COLOR['text_mut'])
lbl_current_task.grid(row=1, column=0, sticky="w", padx=14, pady=(4, 2))
progress_current = ttk.Progressbar(_progress_wrap, length=600, mode="determinate")
progress_current.grid(row=2, column=0, padx=14, pady=(0, 6), sticky="ew")

lbl_overall = tk.Label(_progress_wrap, text="Overall Progress: Ready", anchor="w",
                        font=FONT['body'],
                        bg=COLOR['surface'], fg=COLOR['text_mut'])
lbl_overall.grid(row=3, column=0, sticky="w", padx=14, pady=(4, 2))
progress_overall = ttk.Progressbar(_progress_wrap, length=600, mode="determinate")
progress_overall.grid(row=4, column=0, padx=14, pady=(0, 12), sticky="ew")

progress_mgr = ProgressManager(
    current_bar=progress_current, current_label=lbl_current_task,
    overall_bar=progress_overall, overall_label=lbl_overall)

# ── Action bar — session buttons + run button ────────────────────────
_action_bar = tk.Frame(main_frame, bg=COLOR['bg'])
_action_bar.grid(row=10, column=0, columnspan=3, sticky="ew",
                  padx=12, pady=(4, 18))
_action_bar.columnconfigure(1, weight=1)

# Session buttons — left side
_sess_fr = tk.Frame(_action_bar, bg=COLOR['bg'])
_sess_fr.grid(row=0, column=0, sticky="w")

def _session_btn(parent, text, command):
    b = tk.Button(parent, text=text, command=command,
                   bg=COLOR['surface'], fg=COLOR['text'],
                   activebackground=COLOR['hover'],
                   font=FONT['button'], relief='solid', bd=1,
                   padx=14, pady=7, cursor='hand2',
                   highlightbackground=COLOR['border'],
                   highlightthickness=0)
    b.bind('<Enter>', lambda e: b.config(bg=COLOR['hover']))
    b.bind('<Leave>', lambda e: b.config(bg=COLOR['surface']))
    return b

_session_btn(_sess_fr, "\U0001f4be   Save session",
             command=save_settings).pack(side="left", padx=(0, 8))
_session_btn(_sess_fr, "\U0001f4c2   Load session",
             command=load_settings).pack(side="left", padx=0)

tk.Label(_sess_fr, text="\u2192  Save / restore settings as .json",
         font=FONT['small'], bg=COLOR['bg'], fg=COLOR['text_hint']
         ).pack(side="left", padx=14)

# Batch button — middle, prominent purple/indigo
btn_batch = tk.Button(_action_bar,
                       text="\u2630    Batch Process",
                       command=start_batch_threaded,
                       bg=COLOR['secondary'], fg='white',
                       font=(FONT_FAMILY, 11, 'bold'),
                       relief='flat', bd=0,
                       padx=22, pady=11, cursor='hand2',
                       activebackground='#0e7490',
                       activeforeground='white')
btn_batch.grid(row=0, column=1, sticky="e", padx=(0, 10))
btn_batch.bind('<Enter>', lambda e: btn_batch.config(bg='#0e7490'))
btn_batch.bind('<Leave>', lambda e: btn_batch.config(bg=COLOR['secondary']))

# Start button — right side, prominent
btn_start = tk.Button(_action_bar,
                       text="\u25b6    Run M3C2 Analysis",
                       command=start_m3c2_threaded,
                       bg=COLOR['success'], fg='white',
                       font=(FONT_FAMILY, 12, 'bold'),
                       relief='flat', bd=0,
                       padx=36, pady=12, cursor='hand2',
                       activebackground='#047857',
                       activeforeground='white')
btn_start.grid(row=0, column=2, sticky="e")
btn_start.bind('<Enter>', lambda e: btn_start.config(bg='#047857'))
btn_start.bind('<Leave>', lambda e: btn_start.config(bg=COLOR['success']))

# Footer
_footer = tk.Frame(main_frame, bg=COLOR['bg'])
_footer.grid(row=11, column=0, columnspan=3, sticky="ew", padx=12, pady=(0, 12))
tk.Label(_footer,
         text="M3C2 Tool \u2022 CHECKPOINT 14 \u2022 "
              "Built with py4dgeo \u2022 Samuel Lebó \u00a9 2026",
         font=FONT['small'], bg=COLOR['bg'], fg=COLOR['text_hint']
         ).pack(side="left")

main_frame.columnconfigure(1, weight=1)
root.mainloop()